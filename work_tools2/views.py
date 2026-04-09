from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
import json
import sqlparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from work_tools2.models import Menu, FormConfig, FormQueryItem, FormUpdateItem


def home(request):
    """Home page."""
    return render(request, "home.html", {"active_page": "home"})


def form_merge(request):
    """Runoob tutorial page."""
    return render(request, "form_merge.html", {"active_page": "form_merge"})


def table_config(request):
    """Runoob tutorial page."""
    return render(request, "table_config.html", {"active_page": "table_config"})


def dashboard(request):
    """Dashboard page with sidebar layout."""
    return render(request, "dashboard.html", {"active_page": "dashboard"})


def dynamic(request, form_id: str):
    """Orders list page."""
    return render(request, "dynamic.html", {"active_page": "dynamic"})


@csrf_exempt
def dynamic_submit(request):
    """处理动态表单提交"""
    if request.method == 'POST':
        try:
            # 解析 JSON 数据
            data = json.loads(request.body)

            # 获取配置信息
            config = data.get('config', {})
            form_values = data.get('formValues', {})

            # 执行后端校验
            validation_result = validate_form_data(config, form_values)

            if not validation_result['success']:
                return JsonResponse({
                    'success': False,
                    'message': validation_result['message'],
                    'errors': validation_result.get('errors', [])
                }, status=400)

            # 生成 SQL 更新语句（包括执行语句和回退语句）
            sql_result = generate_update_sql(config, form_values)

            print("=== 接收到表单数据 ===")
            print(f"表单名称：{config.get('formName')}")
            print(f"查询字段数量：{len(config.get('queryItems', []))}")
            print(f"更新字段数量：{len(config.get('updateItems', []))}")
            print(f"\n所有表单值:")
            for key, value in form_values.items():
                print(f"  {key}: {value}")
            print(f"\n生成的执行 SQL 语句:")
            for i, formatted_sql in enumerate(sql_result['forward_sqls'], 1):
                print(f"{i}. \n{formatted_sql}\n")
            print(f"\n生成的回退 SQL 语句:")
            for i, formatted_sql in enumerate(sql_result['backward_sqls'], 1):
                print(f"{i}. \n{formatted_sql}\n")

            # 从公共字段获取 dynamicNo
            dynamic_no = form_values.get('dynamicNo', {}).get('value', '')

            # 创建保存目录 D:\临时文件\YYYYMM\DD\
            import os
            now = datetime.now()
            year_month = now.strftime('%Y%m')
            day = now.strftime('%d')
            save_dir = f"D:\\临时文件\\{year_month}\\{day}"

            print(f"保存目录: {save_dir}")

            # 确保目录存在
            try:
                os.makedirs(save_dir, exist_ok=True)
                print(f"目录创建成功: {os.path.exists(save_dir)}")
            except Exception as e:
                print(f"创建目录失败: {e}")
                raise

            # 生成 SQL 文件内容
            sql_content = []
            sql_content.append(f"-- 表单：{config.get('formName', '未知')}")
            sql_content.append(f"-- 生成时间：{now.strftime('%Y-%m-%d %H:%M:%S')}")
            sql_content.append(f"-- 文件名：{dynamic_no}")
            sql_content.append("")

            # 添加执行语句
            if sql_result['forward_sqls']:
                sql_content.append("-- ==================== 执行语句 ====================")
                sql_content.append("")
                for i, sql in enumerate(sql_result['forward_sqls'], 1):
                    sql_content.append(f"-- 执行语句 {i}")
                    sql_content.append(sql)
                    sql_content.append("")

            # 添加回退语句
            if sql_result['backward_sqls']:
                sql_content.append("-- ==================== 回退语句 ====================")
                sql_content.append("")
                for i, sql in enumerate(sql_result['backward_sqls'], 1):
                    sql_content.append(f"-- 回退语句 {i}")
                    sql_content.append(sql)
                    sql_content.append("")

            # 写入文件
            filename = f"{dynamic_no}.sql"
            filepath = os.path.join(save_dir, filename)

            print(f"SQL 文件路径: {filepath}")

            with open(filepath, 'w', encoding='utf-8') as f:
                f.write('\n'.join(sql_content))

            print(f"SQL 文件写入成功: {os.path.exists(filepath)}")

            # 返回本地完整路径
            return JsonResponse({
                'success': True,
                'message': 'SQL 文件生成成功',
                'filePath': filepath,  # 返回本地完整路径
                'sql_count': len(sql_result['forward_sqls']),
                'forward_sqls': sql_result['forward_sqls'],
                'backward_sqls': sql_result['backward_sqls']
            })

        except json.JSONDecodeError as e:
            return JsonResponse({
                'success': False,
                'message': f'JSON 解析失败：{str(e)}'
            }, status=400)
        except Exception as e:
            print(f"处理异常：{e}")
            return JsonResponse({
                'success': False,
                'message': f'服务器错误：{str(e)}'
            }, status=500)

    return JsonResponse({
        'success': False,
        'message': '仅支持 POST 请求'
    }, status=405)


def generate_update_sql(config, form_values):
    """
    根据配置生成 SQL UPDATE 语句（包括执行语句和回退语句）
    ValidRule 规则说明：
    - required: 必填项，如果值为空则不拼接该字段
    - requiredReverse: 不必填（反向必填），如果值不为空才拼接
    - defaultNull: 不填时为空值，使用 NULL
    - defaultField: 不填时使用字段名本身

    返回：
    {
        'forward_sqls': [执行 SQL 语句列表],
        'backward_sqls': [回退 SQL 语句列表]
    }
    """
    forward_sqls = []
    backward_sqls = []

    table_name_list = config.get('tableNameList', [])
    query_items = config.get('queryItems', [])
    update_items = config.get('updateItems', [])

    # 对每个表生成 UPDATE 语句
    for table_name in table_name_list:
        # 1. 构建 WHERE 条件
        where_conditions = []
        for item in query_items:
            connected_tables = item.get('connectedTable', [])
            valid_rule = item.get('ValidRule', '')

            # 如果该字段关联到当前表
            if table_name in connected_tables:
                binding_key = item.get('bindingKey')
                value_data = form_values.get(binding_key, {})
                value = value_data.get('value', '')

                # 根据 ValidRule 判断是否添加到 WHERE 条件
                if valid_rule == 'required':
                    # 必填项：只有值不为空时才添加
                    if value:
                        where_conditions.append(f"{binding_key} = '{value}'")
                elif valid_rule == 'requiredReverse':
                    # 不必填：值不为空时才添加
                    if value:
                        where_conditions.append(f"{binding_key} = '{value}'")
                else:
                    # 其他情况默认添加
                    where_conditions.append(f"{binding_key} = '{value}'")

        # 2. 构建执行语句的 SET 子句（使用 newValue）
        forward_set_clauses = []
        # 3. 构建回退语句的 SET 子句（使用 originValue）
        backward_set_clauses = []

        for item in update_items:
            connected_tables = item.get('connectedTable', [])
            new_valid_rule = item.get('newValidRule', '')
            origin_valid_rule = item.get('originValidRule', '')

            # 如果是补充框，需要处理子字段
            if item.get('inputType') == 'supplement':
                parent_key = item.get('bindingKey')

                # 处理主字段 - 执行语句（newValue）
                value_data = form_values.get(parent_key, {})
                new_value = value_data.get('newValue', '')
                origin_value = value_data.get('originValue', '')

                # 执行语句：根据 newValidRule 处理 newValue
                forward_set_value = handle_field_value(parent_key, new_value, new_valid_rule)
                if forward_set_value is not None and table_name in connected_tables:
                    forward_set_clauses.append(forward_set_value)

                # 回退语句：根据 originValidRule 处理 originValue
                backward_set_value = handle_field_value(parent_key, origin_value, origin_valid_rule)
                if backward_set_value is not None and table_name in connected_tables:
                    backward_set_clauses.append(backward_set_value)

                # 处理子字段（继承主字段的 connectedTable）
                sub_fields = item.get('subFields', [])
                for sub_field in sub_fields:
                    sub_binding_key = sub_field.get('bindingKey')
                    sub_value_data = form_values.get(sub_binding_key, {})
                    sub_new_value = sub_value_data.get('newValue', '')
                    sub_origin_value = sub_value_data.get('originValue', '')

                    # 执行语句：子字段继承主字段的表关联和验证规则
                    if table_name in connected_tables:
                        forward_sub_set_value = handle_field_value(sub_binding_key, sub_new_value, new_valid_rule)
                        if forward_sub_set_value is not None:
                            forward_set_clauses.append(forward_sub_set_value)

                        # 回退语句：子字段特殊处理 - 输入什么保留什么，为空就拼接=''
                        if table_name in connected_tables:
                            # 子字段的回退语句不使用 ValidRule，直接处理
                            if sub_origin_value is None or sub_origin_value == '':
                                backward_sub_set_value = f"{sub_binding_key} = ''"
                            else:
                                backward_sub_set_value = f"{sub_binding_key} = '{sub_origin_value}'"

                            if backward_sub_set_value is not None:
                                backward_set_clauses.append(backward_sub_set_value)
            else:
                # 普通字段直接处理
                binding_key = item.get('bindingKey')
                value_data = form_values.get(binding_key, {})
                new_value = value_data.get('newValue', '')
                origin_value = value_data.get('originValue', '')

                # 执行语句：使用 newValue
                if table_name in connected_tables:
                    forward_set_value = handle_field_value(binding_key, new_value, new_valid_rule)
                    if forward_set_value is not None:
                        forward_set_clauses.append(forward_set_value)

                    # 回退语句：使用 originValue
                    backward_set_value = handle_field_value(binding_key, origin_value, origin_valid_rule)
                    if backward_set_value is not None:
                        backward_set_clauses.append(backward_set_value)

        # 4. 只有当有 SET 子句和 WHERE 条件时才生成 SQL
        if (forward_set_clauses or backward_set_clauses) and where_conditions:
            where_clause_str = ' AND '.join(where_conditions)

            # 生成执行语句
            if forward_set_clauses:
                forward_set_clause_str = ', '.join(forward_set_clauses)
                forward_sql = f"UPDATE {table_name} SET {forward_set_clause_str} WHERE {where_clause_str}"
                forward_sqls.append(format_sql(forward_sql))

            # 生成回退语句
            if backward_set_clauses:
                backward_set_clause_str = ', '.join(backward_set_clauses)
                backward_sql = f"UPDATE {table_name} SET {backward_set_clause_str} WHERE {where_clause_str}"
                backward_sqls.append(format_sql(backward_sql))

    return {
        'forward_sqls': forward_sqls,
        'backward_sqls': backward_sqls
    }


def format_sql(sql):
    """
    格式化 SQL 语句（模仿 sql-formatter 风格）
    参考：https://github.com/sql-formatter-org/sql-formatter
    """
    # 使用 sqlparse 做基础格式化
    formatted = sqlparse.format(
        sql,
        reindent=True,
        keyword_case='upper',
        identifier_case='lower',
        use_space_around_operators=True,
    )

    # 进一步优化格式
    lines = formatted.split('\n')
    result_lines = []

    indent = '  '  # 2 空格缩进

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # UPDATE 语句特殊处理
        if line.upper().startswith('UPDATE'):
            result_lines.append(line)
        elif 'SET' in line.upper() and line.strip().upper().startswith('SET'):
            # SET 关键字单独一行
            result_lines.append(line)
        elif line.startswith(',') or line.startswith('AND') or line.startswith('OR'):
            # 操作符前置，带缩进
            result_lines.append(indent + line)
        elif '=' in line and (',' in line or 'AND' in line.upper()):
            # SET 子句的字段行
            if line.startswith('='):
                result_lines.append(indent + line)
            else:
                result_lines.append(indent + line)
        else:
            result_lines.append(indent + line)

    return '\n'.join(result_lines)


def handle_field_value(field_name, value, valid_rule):
    """
    根据 ValidRule 处理字段值
    返回：格式化后的 SQL 片段 或 None（表示不拼接该字段）

    规则说明：
    - required: 必填项，如果值为空返回 None（不拼接）
    - requiredReverse: 不必填，如果值不为空才拼接，否则返回 None
    - defaultNull: 不填时为空值，值为空时使用 NULL
    - defaultField: 不填时使用字段名本身，值为空时返回 field_name = field_name
    """
    # 检查值是否为空
    is_empty = (value is None or value == '' or value == 'null' or value == 'NULL')

    if valid_rule == 'required':
        # 必填项：如果值为空，不拼接该字段
        if is_empty:
            return None
        else:
            return f"{field_name} = '{value}'"

    elif valid_rule == 'requiredReverse':
        # 不必填：值不为空才拼接
        if not is_empty:
            return f"{field_name} = '{value}'"
        else:
            return None

    elif valid_rule == 'defaultNull':
        # 不填时为空值
        if is_empty:
            return f"{field_name} = NULL"
        else:
            return f"{field_name} = '{value}'"

    elif valid_rule == 'defaultField':
        # 不填时使用字段名本身
        if is_empty:
            return f"{field_name} = {field_name}"
        else:
            return f"{field_name} = '{value}'"

    else:
        # 默认情况：直接拼接
        if is_empty:
            return None
        else:
            return f"{field_name} = '{value}'"


@csrf_exempt
def download_template(request):
    """
    根据配置动态生成 Excel 导入模板
    """
    if request.method == 'POST':
        try:
            # 解析 JSON 数据
            data = json.loads(request.body)
            config = data.get('config', {})

            form_name = config.get('formName', '模板')
            query_items = config.get('queryItems', [])
            update_items = config.get('updateItems', [])

            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active
            ws.title = form_name[:31]  # Excel 工作表名最长 31 字符

            # 设置表头样式
            header_font = Font(bold=True, size=12)
            header_alignment = Alignment(horizontal='center', vertical='center')
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # 构建表头列表
            headers = []

            # 1. 添加查询字段表头（仅一个 label）
            for item in query_items:
                headers.append({
                    'label': item.get('label', ''),
                    'bindingKey': item.get('bindingKey', ''),
                    'type': 'query'
                })

            # 2. 添加更新字段表头
            for item in update_items:
                if item.get('inputType') == 'supplement':
                    # 补充框特殊处理
                    parent_label = item.get('label', '')
                    parent_binding_key = item.get('bindingKey', '')
                    sub_fields = item.get('subFields', [])

                    # 新值：只添加主字段 label
                    headers.append({
                        'label': f'新{parent_label}',
                        'bindingKey': parent_binding_key,
                        'type': 'update_new'
                    })

                    # 原值：添加主字段 label + 所有子字段 label
                    # 主字段原值
                    headers.append({
                        'label': f'原{parent_label}',
                        'bindingKey': parent_binding_key,
                        'type': 'update_origin'
                    })

                    # 子字段原值
                    for sub_field in sub_fields:
                        sub_label = sub_field.get('label', '')
                        sub_binding_key = sub_field.get('bindingKey', '')

                        headers.append({
                            'label': f'原{sub_label}',
                            'bindingKey': sub_binding_key,
                            'type': 'update_origin_sub'
                        })
                else:
                    # 普通字段的新值和原值
                    label = item.get('label', '')
                    binding_key = item.get('bindingKey', '')

                    headers.append({
                        'label': f'新{label}',
                        'bindingKey': binding_key,
                        'type': 'update_new'
                    })
                    headers.append({
                        'label': f'原{label}',
                        'bindingKey': binding_key,
                        'type': 'update_origin'
                    })

            # 写入表头
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header['label'])
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill

                # 设置列宽
                col_letter = chr(64 + (col_num % 26)) if col_num <= 26 else chr(64 + (col_num // 26)) + chr(
                    64 + (col_num % 26))
                ws.column_dimensions[col_letter].width = 15

            # 生成文件名：formName_时间戳.xlsx
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{form_name}_{timestamp}.xlsx".replace('/', '-').replace('\\', '-')
            # 保存到内存
            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # 返回文件
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            # 使用 quote 处理中文文件名
            from urllib.parse import quote
            encoded_filename = quote(filename)
            response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'

            return response


        except json.JSONDecodeError as e:
            return JsonResponse({
                'success': False,
                'message': f'JSON 解析失败：{str(e)}'
            }, status=400)
        except Exception as e:
            print(f"处理异常：{e}")
            return JsonResponse({
                'success': False,
                'message': f'服务器错误：{str(e)}'
            }, status=500)

    return JsonResponse({
        'success': False,
        'message': '仅支持 POST 请求'
    }, status=405)


from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import json
import sqlparse
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Color
from io import BytesIO

from work_tools2.models import Menu


@csrf_exempt
def batch_import(request):
    """
    批量导入数据
    逻辑：
    1. 接收 Excel 文件、配置和公共字段
    2. 校验公共字段（filePrefix、onesLink、dynamicNo）
    3. 读取 Excel 数据
    4. 检查是否有有效的数据列
    5. 对每一行数据进行校验和 SQL 生成
    6. 如果校验失败，在失败列记录错误信息
    7. 全部成功：将所有 SQL 写入 .sql 文件到 D:\临时文件\YYYYMM\DD\
    8. 有失败：将包含失败信息的 Excel 保存到 D:\临时文件\YYYYMM\DD\
    9. 返回文件的本地完整路径
    """
    if request.method == 'POST':
        try:
            import os

            # 获取上传的文件
            file = request.FILES.get('file')
            config_json = request.POST.get('config')
            query_values_json = request.POST.get('queryValues')

            if not file or not config_json:
                return JsonResponse({
                    'success': False,
                    'message': '缺少文件或配置参数'
                }, status=400)

            # 解析配置
            config = json.loads(config_json)
            form_name = config.get('formName', '模板')
            query_items = config.get('queryItems', [])
            update_items = config.get('updateItems', [])

            # 解析公共字段值
            query_values = {}
            if query_values_json:
                try:
                    query_values = json.loads(query_values_json)
                except json.JSONDecodeError:
                    query_values = {}

            # 校验公共字段（始终校验）
            common_fields = ['filePrefix', 'onesLink', 'dynamicNo']
            for field_name in common_fields:
                value_data = query_values.get(field_name, {})

                # 处理 value_data 可能是字符串或字典的情况
                if isinstance(value_data, dict):
                    value = value_data.get('value', '')
                elif isinstance(value_data, str):
                    value = value_data
                else:
                    value = ''

                if not value or str(value).strip() == '':
                    return JsonResponse({
                        'success': False,
                        'message': f'{field_name}不能为空'
                    }, status=400)

            # 加载 Excel 文件
            wb = load_workbook(file)
            ws = wb.active

            # 构建表头映射（从第一行读取）
            headers = {}
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                # 去除首尾空格并检查是否为有效字符串
                if cell_value is not None and str(cell_value).strip():
                    headers[str(cell_value).strip()] = col

            print(f"headers: {headers}")

            # 检查是否有有效的数据列
            required_columns = []
            # 查询字段至少需要一个
            for item in query_items:
                required_columns.append(item.get('label', ''))

            # 更新字段需要新值列（补充框只检查主字段，子字段通过匹配自动填充）
            for item in update_items:
                if item.get('inputType') == 'supplement':
                    # 补充框只添加主字段的新值，不添加子字段
                    required_columns.append(f'新{item.get("label", "")}')
                else:
                    required_columns.append(f'新{item.get("label", "")}')

            print(f"required_columns: {required_columns}")

            # 检查是否有任何一个必需的列存在
            has_valid_data = any(col in headers for col in required_columns)

            print(f"has_valid_data: {has_valid_data}")

            if not has_valid_data or len(headers) == 0:
                return JsonResponse({
                    'success': False,
                    'message': '数据表中无有效的数据，请检查 Excel 文件格式是否正确'
                }, status=400)

            print(f"Excel 总行数：{ws.max_row}")

            # 检查是否有实际的数据行（至少要有表头 +1 行数据）
            if ws.max_row < 2:
                return JsonResponse({
                    'success': False,
                    'message': 'Excel 文件没有数据行，请至少填写一行数据'
                }, status=400)

            # 统计有多少行包含有效的必填数据
            valid_data_rows = 0
            for row_idx in range(2, ws.max_row + 1):
                has_required_value = False
                for col_name in required_columns:
                    col_num = headers.get(col_name)
                    if col_num:
                        cell_value = ws.cell(row=row_idx, column=col_num).value
                        if cell_value is not None and str(cell_value).strip():
                            has_required_value = True
                            break
                if has_required_value:
                    valid_data_rows += 1

            if valid_data_rows == 0:
                return JsonResponse({
                    'success': False,
                    'message': f'Excel 中没有有效的必填数据（共{ws.max_row - 1}行，但都没有必填字段的值）'
                }, status=400)

            print(f"有效数据行数：{valid_data_rows}")

            # 添加失败原因列
            fail_column = ws.max_column + 1
            ws.cell(row=1, column=fail_column, value='失败原因')
            ws.cell(row=1, column=fail_column).font = Font(bold=True)
            ws.cell(row=1, column=fail_column).alignment = Alignment(horizontal='center')
            ws.cell(row=1, column=fail_column).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC",
                                                                  fill_type="solid")

            # 统计信息
            total_rows = ws.max_row - 1  # 减去表头
            success_count = 0
            fail_count = 0
            all_sql_statements = []  # 存储所有成功的 SQL 语句

            # 从第二行开始处理数据
            for row_idx in range(2, ws.max_row + 1):
                # 构建表单值
                form_values, missing_columns = build_form_values_from_excel(ws, row_idx, headers, query_items,
                                                                            update_items)

                # 如果有缺失的列，直接记录错误
                if missing_columns:
                    missing_cols_str = ', '.join(missing_columns)
                    fail_count += 1

                    ws.cell(row=row_idx, column=fail_column, value=f'缺少必需的列：{missing_cols_str}')
                    ws.cell(row=row_idx, column=fail_column).fill = PatternFill(start_color="FFFFCC",
                                                                                end_color="FFFFCC", fill_type="solid")
                    continue

                # 执行校验（传入 query_values 用于校验公共字段）
                validation_result = validate_form_data(config, form_values, query_values)

                if not validation_result['success']:
                    fail_count += 1

                    # 写入失败原因
                    fail_reason = '; '.join(validation_result.get('errors', []))
                    ws.cell(row=row_idx, column=fail_column, value=fail_reason)
                    ws.cell(row=row_idx, column=fail_column).fill = PatternFill(start_color="FFFFCC",
                                                                                end_color="FFFFCC", fill_type="solid")
                else:
                    # 生成 SQL 语句
                    sql_result = generate_update_sql(config, form_values)

                    if sql_result['forward_sqls'] and sql_result['backward_sqls']:
                        # 收集 SQL 语句
                        all_sql_statements.append({
                            'row': row_idx,
                            'forward_sqls': sql_result['forward_sqls'],
                            'backward_sqls': sql_result['backward_sqls']
                        })
                        success_count += 1
                    else:
                        fail_count += 1

                        fail_reason = '未生成有效的 SQL 语句'
                        ws.cell(row=row_idx, column=fail_column, value=fail_reason)
                        ws.cell(row=row_idx, column=fail_column).fill = PatternFill(start_color="FFFFCC",
                                                                                    end_color="FFFFCC",
                                                                                    fill_type="solid")

            # 添加统计信息工作表
            stats_ws = wb.create_sheet(title='导入统计')
            stats_ws.cell(row=1, column=1, value='总行数').font = Font(bold=True)
            stats_ws.cell(row=1, column=2, value=total_rows).font = Font(bold=True)
            stats_ws.cell(row=2, column=1, value='成功数').font = Font(bold=True)
            stats_ws.cell(row=2, column=2, value=success_count).font = Font(bold=True)
            stats_ws.cell(row=3, column=1, value='失败数').font = Font(bold=True)
            stats_ws.cell(row=3, column=2, value=fail_count).font = Font(bold=True)
            stats_ws.cell(row=4, column=1, value='成功率').font = Font(bold=True)
            stats_ws.cell(row=4, column=2,
                          value=f'{success_count / total_rows * 100:.2f}%' if total_rows > 0 else '0%').font = Font(
                bold=True)

            # 从公共字段获取 dynamicNo
            dynamic_no = ''
            if query_values:
                dynamic_no_data = query_values.get('dynamicNo', '')
                if isinstance(dynamic_no_data, dict):
                    dynamic_no = dynamic_no_data.get('value', '')
                elif isinstance(dynamic_no_data, str):
                    dynamic_no = dynamic_no_data

            # 创建保存目录 D:\临时文件\YYYYMM\DD\
            now = datetime.now()
            year_month = now.strftime('%Y%m')
            day = now.strftime('%d')
            save_dir = f"D:\\临时文件\\{year_month}\\{day}"

            print(f"保存目录: {save_dir}")

            # 确保目录存在
            try:
                os.makedirs(save_dir, exist_ok=True)
                print(f"目录创建成功: {os.path.exists(save_dir)}")
            except Exception as e:
                print(f"创建目录失败: {e}")
                raise

            # 判断是否全部成功
            all_success = (fail_count == 0 and success_count > 0)

            # 只有全部成功时才生成 SQL 文件
            sql_file_path = None
            if all_success and all_sql_statements:
                sql_filename = f"{dynamic_no}.sql"
                sql_filepath = os.path.join(save_dir, sql_filename)

                print(f"SQL 文件路径: {sql_filepath}")

                # 生成 SQL 文件内容
                sql_content = []
                sql_content.append(f"-- 批量导入 SQL 文件")
                sql_content.append(f"-- 表单：{form_name}")
                sql_content.append(f"-- 生成时间：{now.strftime('%Y-%m-%d %H:%M:%S')}")
                sql_content.append(f"-- 总行数：{total_rows}，成功：{success_count}，失败：{fail_count}")
                sql_content.append("")

                for idx, stmt in enumerate(all_sql_statements, 1):
                    sql_content.append(f"-- ==================== 第 {stmt['row']} 行数据 ====================")
                    sql_content.append("")
                    sql_content.append("-- 执行语句")
                    for sql in stmt['forward_sqls']:
                        sql_content.append(sql)
                        sql_content.append("")
                    sql_content.append("-- 回退语句")
                    for sql in stmt['backward_sqls']:
                        sql_content.append(sql)
                        sql_content.append("")

                # 写入文件
                try:
                    with open(sql_filepath, 'w', encoding='utf-8') as f:
                        f.write('\n'.join(sql_content))
                    print(f"SQL 文件写入成功: {os.path.exists(sql_filepath)}")
                    sql_file_path = sql_filepath
                except Exception as e:
                    print(f"SQL 文件写入失败: {e}")
                    raise

            # 只有有失败记录时才保存 Excel 结果文件
            excel_file_path = None
            if fail_count > 0:
                excel_filename = f"{dynamic_no}_导入结果_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
                excel_filepath = os.path.join(save_dir, excel_filename)

                print(f"Excel 文件路径: {excel_filepath}")

                try:
                    wb.save(excel_filepath)
                    print(f"Excel 文件写入成功: {os.path.exists(excel_filepath)}")
                    excel_file_path = excel_filepath
                except Exception as e:
                    print(f"Excel 文件写入失败: {e}")
                    raise

            # 返回文件路径
            return JsonResponse({
                'success': True,
                'message': f'批量导入完成，成功{success_count}条，失败{fail_count}条',
                'sqlFilePath': sql_file_path,  # 只有全部成功时才有值
                'excelFilePath': excel_file_path,  # 只有有失败时才有值
                'totalRows': total_rows,
                'successCount': success_count,
                'failCount': fail_count
            })

        except Exception as e:
            print(f"批量导入异常：{e}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'message': f'服务器错误：{str(e)}'
            }, status=500)

    return JsonResponse({
        'success': False,
        'message': '仅支持 POST 请求'
    }, status=405)


@csrf_exempt
def download_failed_file(request):
    """
    下载失败的 Excel 结果文件
    """
    if request.method == 'GET':
        try:
            import os
            from django.http import FileResponse

            file_path = request.GET.get('path', '')

            if not file_path or not os.path.exists(file_path):
                return JsonResponse({
                    'success': False,
                    'message': '文件不存在'
                }, status=404)

            # 返回文件
            response = FileResponse(open(file_path, 'rb'))
            response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            # 设置文件名
            filename = os.path.basename(file_path)
            from urllib.parse import quote
            encoded_filename = quote(filename)
            response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'

            return response

        except Exception as e:
            print(f"下载文件异常：{e}")
            return JsonResponse({
                'success': False,
                'message': f'服务器错误：{str(e)}'
            }, status=500)

    return JsonResponse({
        'success': False,
        'message': '仅支持 GET 请求'
    }, status=405)



def build_form_values_from_excel(ws, row_idx, headers, query_items, update_items):
    """
    从 Excel 行构建表单值
    返回：(form_values, missing_columns)
    - form_values: 表单值字典
    - missing_columns: 缺失的列名列表

    注意：
    - 补充框的主字段新值是必填的
    - 补充框的子字段是可选的（Excel 中有则读取，没有则留空，后续通过主字段匹配填充）
    """
    form_values = {}
    missing_columns = []

    # 处理查询字段
    for item in query_items:
        label = item.get('label', '')
        binding_key = item.get('bindingKey', '')

        if label in headers:
            col = headers[label]
            value = ws.cell(row=row_idx, column=col).value
            form_values[binding_key] = {
                'label': label,
                'value': str(value) if value is not None else '',
                'inputType': 'query',
                'fieldType': item.get('type', 'text'),
                'ValidRule': item.get('ValidRule', '')
            }
        else:
            missing_columns.append(label)
            form_values[binding_key] = {
                'label': label,
                'value': '',
                'inputType': 'query',
                'fieldType': item.get('type', 'text'),
                'ValidRule': item.get('ValidRule', '')
            }

    # 处理更新字段
    for item in update_items:
        label = item.get('label', '')
        binding_key = item.get('bindingKey', '')
        input_type = item.get('inputType', '')

        if input_type == 'supplement':
            # 补充框特殊处理
            sub_fields = item.get('subFields', [])

            # 主字段新值（必填）
            new_label = f'新{label}'
            origin_label = f'原{label}'

            new_value = ''
            origin_value = ''

            if new_label in headers:
                col = headers[new_label]
                new_value = ws.cell(row=row_idx, column=col).value
            else:
                missing_columns.append(new_label)

            # 原值（可选，有则读取）
            if origin_label in headers:
                col = headers[origin_label]
                origin_value = ws.cell(row=row_idx, column=col).value

            form_values[binding_key] = {
                'label': label,
                'newValue': str(new_value) if new_value is not None else '',
                'originValue': str(origin_value) if origin_value is not None else '',
                'inputType': 'supplement',
                'fieldType': 'supplement',
                'newValidRule': item.get('newValidRule', ''),
                'originValidRule': item.get('originValidRule', '')
            }

            # 子字段（可选，有则读取，没有留空后续通过匹配填充）
            for sub_field in sub_fields:
                sub_label = sub_field.get('label', '')
                sub_binding_key = sub_field.get('bindingKey', '')

                new_sub_label = f'新{sub_label}'
                origin_sub_label = f'原{sub_label}'

                new_sub_value = ''
                origin_sub_value = ''

                # 新子字段（可选）
                if new_sub_label in headers:
                    col = headers[new_sub_label]
                    new_sub_value = ws.cell(row=row_idx, column=col).value

                # 原子字段（可选）
                if origin_sub_label in headers:
                    col = headers[origin_sub_label]
                    origin_sub_value = ws.cell(row=row_idx, column=col).value

                form_values[sub_binding_key] = {
                    'newValue': str(new_sub_value) if new_sub_value is not None else '',
                    'originValue': str(origin_sub_value) if origin_sub_value is not None else '',
                    'inputType': 'supplement-sub',
                    'fieldType': 'supplement-sub',
                    'parentKey': binding_key,
                    'label': sub_label
                }
        else:
            # 普通字段
            new_label = f'新{label}'
            origin_label = f'原{label}'

            new_value = ''
            origin_value = ''

            if new_label in headers:
                col = headers[new_label]
                new_value = ws.cell(row=row_idx, column=col).value
            else:
                missing_columns.append(new_label)

            if origin_label in headers:
                col = headers[origin_label]
                origin_value = ws.cell(row=row_idx, column=col).value
            else:
                missing_columns.append(origin_label)

            form_values[binding_key] = {
                'label': label,
                'newValue': str(new_value) if new_value is not None else '',
                'originValue': str(origin_value) if origin_value is not None else '',
                'inputType': input_type,
                'fieldType': item.get('type', 'text'),
                'newValidRule': item.get('newValidRule', ''),
                'originValidRule': item.get('originValidRule', '')
            }

    return form_values, missing_columns


def validate_form_data(config, form_values, query_values=None):
    """
    后端表单校验（与前端逻辑一致）
    返回：{'success': bool, 'message': str, 'errors': list}

    参数：
    - config: 配置对象
    - form_values: 表单值字典
    - query_values: 查询字段值（包含公共字段），可选
    """
    errors = []

    query_items = config.get('queryItems', [])
    update_items = config.get('updateItems', [])

    # 1. 校验文本、数值、日期、下拉框的新值和原值
    for item in update_items:
        if item.get('inputType') in ['select', 'input'] and item.get('type') in ['text', 'number', 'date', 'string']:
            binding_key = item.get('bindingKey')
            value_data = form_values.get(binding_key, {})

            # 新值校验
            new_value = value_data.get('newValue')
            new_valid_rule = item.get('newValidRule', '')
            if (new_value is None or new_value == '') and new_valid_rule == 'required':
                errors.append(f"新{item.get('label')}不能为空")

            # 原值校验
            origin_value = value_data.get('originValue')
            origin_valid_rule = item.get('originValidRule', '')
            if (origin_value is None or origin_value == '') and origin_valid_rule == 'required':
                errors.append(f"原{item.get('label')}不能为空")

    # 2. 校验单选框
    for item in update_items:
        if item.get('inputType') == 'radio':
            binding_key = item.get('bindingKey')
            value_data = form_values.get(binding_key, {})
            new_value = value_data.get('newValue')
            valid_rule = item.get('newValidRule', '')

            if (new_value is None or new_value == '') and valid_rule == 'required':
                errors.append(f"新{item.get('label')}不能为空")

    # 3. 校验查询字段
    for item in query_items:
        binding_key = item.get('bindingKey')
        value_data = form_values.get(binding_key, {})
        value = value_data.get('value')
        valid_rule = item.get('ValidRule', '')

        if (value is None or value == '') and valid_rule == 'required':
            errors.append(f"{item.get('label')}不能为空")

    # 4. 校验公共字段（优先使用 query_values，否则从 form_values 中读取）
    common_fields = ['filePrefix', 'onesLink', 'dynamicNo']
    for field_name in common_fields:
        value = None

        # 先从 query_values 中读取（批量导入场景）
        if query_values:
            value_data = query_values.get(field_name, {})
            value = value_data.get('value') if isinstance(value_data, dict) else value_data

        # 如果 query_values 中没有，再从 form_values 中读取（普通提交场景）
        if not value:
            value_data = form_values.get(field_name, {})
            value = value_data.get('value')

        if value is None or str(value).strip() == '':
            errors.append(f"{field_name}不能为空")

    # 5. 校验查询字段至少有一个不为空
    query_field_values = [
        form_values.get(item.get('bindingKey'), {}).get('value')
        for item in query_items
    ]
    has_non_empty_query = any(
        v is not None and v != ''
        for v in query_field_values
    )

    if query_items and not has_non_empty_query:
        errors.append("查询字段至少需要填写一个条件")

    # 返回校验结果
    if errors:
        return {
            'success': False,
            'message': f'共有{len(errors)}个校验错误',
            'errors': errors
        }

    return {
        'success': True,
        'message': '校验通过',
        'errors': []
    }


def get_menus():
    """获取菜单数据，用于侧边栏"""
    # 获取所有一级菜单（is_visible=True）
    parent_menus = Menu.objects.filter(parent__isnull=True, is_visible=True).order_by(
        "group_name", "sort_order"
    )

    menus_by_group = {}
    for menu in parent_menus:
        group = menu.group_name or "其他"
        if group not in menus_by_group:
            menus_by_group[group] = []

        # 获取子菜单
        children = Menu.objects.filter(parent=menu, is_visible=True).order_by(
            "sort_order"
        )

        menus_by_group[group].append(
            {
                "id": menu.id,
                "name": menu.name,
                "pinyin": menu.pinyin or "",
                "icon": menu.icon or "bi-circle",
                "url": menu.url,
                "has_children": children.exists(),
                "children": [
                    {
                        "id": child.id,
                        "name": child.name,
                        "pinyin": child.pinyin or "",
                        "url": child.url,
                    }
                    for child in children
                ],
            }
        )

    return menus_by_group


def render_with_menus(request, template_name, context=None):
    """Render template with menu data"""
    if context is None:
        context = {}

    # 添加菜单数据
    context["menus"] = get_menus()

    return render(request, template_name, context)


@csrf_exempt
def get_form_configs(request):
    """获取所有表单配置列表"""
    if request.method == 'GET':
        try:
            configs = FormConfig.objects.all().order_by('-created_at')
            config_list = []

            for config in configs:
                config_list.append({
                    'id': config.id,
                    'form_id': config.form_id,
                    'form_name': config.form_name,
                    'table_name_list': config.table_name_list,
                    'is_active': config.is_active,
                    'created_at': config.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': config.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                })

            return JsonResponse({
                'success': True,
                'data': config_list
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'获取失败：{str(e)}'
            }, status=500)

    return JsonResponse({
        'success': False,
        'message': '仅支持 GET 请求'
    }, status=405)


@csrf_exempt
def get_form_config_detail(request, form_id):
    """获取单个表单配置详情"""
    if request.method == 'GET':
        try:
            config = FormConfig.objects.get(form_id=form_id)

            # 构建配置对象
            config_data = {
                'formName': config.form_name,
                'tableNameList': config.table_name_list,
                'queryItems': [],
                'updateItems': []
            }

            # 获取查询字段
            query_items = FormQueryItem.objects.filter(form_config=config).order_by('sort_order')
            for item in query_items:
                config_data['queryItems'].append({
                    'label': item.label,
                    'type': item.field_type,
                    'defaultValue': item.default_value,
                    'bindingKey': item.binding_key,
                    'sortOrder': item.sort_order,
                    'connectedTable': item.connected_table,
                    'ValidRule': item.valid_rule,
                })

            # 获取更新字段
            update_items = FormUpdateItem.objects.filter(form_config=config).order_by('sort_order')
            for item in update_items:
                update_item_data = {
                    'label': item.label,
                    'type': item.field_type,
                    'originDefaultValue': item.origin_default_value,
                    'newDefaultValue': item.new_default_value,
                    'bindingKey': item.binding_key,
                    'sortOrder': item.sort_order,
                    'inputType': item.input_type,
                    'connectedTable': item.connected_table,
                    'newValidRule': item.new_valid_rule,
                    'originValidRule': item.origin_valid_rule,
                }

                # 如果有选项配置
                if item.options:
                    update_item_data['options'] = item.options

                # 如果是补充框
                if item.input_type == 'supplement':
                    update_item_data['mainField'] = item.main_field
                    update_item_data['subFields'] = item.sub_fields

                config_data['updateItems'].append(update_item_data)

            return JsonResponse({
                'success': True,
                'data': config_data
            })
        except FormConfig.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': '表单配置不存在'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'获取失败：{str(e)}'
            }, status=500)

    return JsonResponse({
        'success': False,
        'message': '仅支持 GET 请求'
    }, status=405)


@csrf_exempt
def save_form_config(request):
    """保存表单配置（新增或更新）"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            form_id = data.get('formId')  # 编辑时传入
            form_name = data.get('formName')
            table_name_list = data.get('tableNameList', [])
            query_items = data.get('queryItems', [])
            update_items = data.get('updateItems', [])

            if not form_name:
                return JsonResponse({
                    'success': False,
                    'message': '表单名称不能为空'
                }, status=400)

            if not table_name_list or len(table_name_list) == 0:
                return JsonResponse({
                    'success': False,
                    'message': '至少需要配置一个表名'
                }, status=400)

            # 检查是否已存在（编辑模式）
            if form_id:
                try:
                    config = FormConfig.objects.get(form_id=form_id)
                    config.form_name = form_name
                    config.table_name_list = table_name_list
                    config.save()
                except FormConfig.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'message': '表单配置不存在'
                    }, status=404)
            else:
                # 新增模式，自动生成 form_id
                config = FormConfig.objects.create(
                    form_name=form_name,
                    table_name_list=table_name_list,
                    is_active=True,
                )
                form_id = config.form_id

            # 删除旧的查询字段
            FormQueryItem.objects.filter(form_config=config).delete()

            # 创建新的查询字段
            for item_data in query_items:
                FormQueryItem.objects.create(
                    form_config=config,
                    label=item_data.get('label'),
                    field_type=item_data.get('type', 'text'),
                    binding_key=item_data.get('bindingKey'),
                    sort_order=item_data.get('sortOrder', 0),
                    connected_table=item_data.get('connectedTable', []),
                    valid_rule=item_data.get('ValidRule', 'required'),
                    default_value=item_data.get('defaultValue', ''),
                )

            # 删除旧的更新字段
            FormUpdateItem.objects.filter(form_config=config).delete()

            # 创建新的更新字段
            for item_data in update_items:
                FormUpdateItem.objects.create(
                    form_config=config,
                    label=item_data.get('label'),
                    field_type=item_data.get('type', 'text'),
                    binding_key=item_data.get('bindingKey'),
                    sort_order=item_data.get('sortOrder', 0),
                    input_type=item_data.get('inputType', 'input'),
                    connected_table=item_data.get('connectedTable', []),
                    new_valid_rule=item_data.get('newValidRule', 'required'),
                    origin_valid_rule=item_data.get('originValidRule', 'required'),
                    origin_default_value=item_data.get('originDefaultValue', ''),
                    new_default_value=item_data.get('newDefaultValue', ''),
                    main_field=item_data.get('mainField', ''),
                    sub_fields=item_data.get('subFields', []),
                    options=item_data.get('options', []),
                )

            return JsonResponse({
                'success': True,
                'message': '保存成功',
                'form_id': form_id
            })
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'JSON 解析失败'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'保存失败：{str(e)}'
            }, status=500)

    return JsonResponse({
        'success': False,
        'message': '仅支持 POST 请求'
    }, status=405)


@csrf_exempt
def delete_form_config(request, form_id):
    """删除表单配置"""
    if request.method == 'DELETE':
        try:
            config = FormConfig.objects.get(form_id=form_id)
            config.delete()

            return JsonResponse({
                'success': True,
                'message': '删除成功'
            })
        except FormConfig.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': '表单配置不存在'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'删除失败：{str(e)}'
            }, status=500)

    return JsonResponse({
        'success': False,
        'message': '仅支持 DELETE 请求'
    }, status=405)

