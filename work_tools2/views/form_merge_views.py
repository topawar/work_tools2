import json
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from work_tools2.models import FormConfig, FormQueryItem, FormUpdateItem
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import load_workbook
from datetime import datetime
from io import BytesIO
import os


@csrf_exempt
def download_merge_template(request):
    """下载合并模板（多Sheet Excel）"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            form_ids = data.get('formIds', [])
            file_prefix = data.get('filePrefix', 'merge')

            if not form_ids or len(form_ids) == 0:
                return JsonResponse({
                    'success': False,
                    'message': '请至少选择一个表单'
                }, status=400)

            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            # 删除默认创建的sheet
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

            # 为每个选中的表单创建一个Sheet
            for form_id in form_ids:
                try:
                    config = FormConfig.objects.get(id=form_id)

                    # 创建Sheet，使用表单名称作为Sheet名
                    ws = wb.create_sheet(title=config.form_name)

                    # 设置表头样式
                    header_font_white = Font(bold=True, color='FFFFFF')  # 白色字体（用于蓝色背景）
                    header_font_black = Font(bold=True, color='000000')  # 黑色字体（用于黄色背景）
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    header_alignment = Alignment(horizontal='center', vertical='center')
                    # 黄色背景，用于标注有默认值的字段
                    default_value_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

                    # 构建表头：查询字段 + 更新字段
                    headers = []

                    # 添加查询字段
                    query_items = FormQueryItem.objects.filter(form_config=config).order_by('sort_order')
                    for item in query_items:
                        headers.append({
                            'label': f'{item.label}',
                            'hasDefaultValue': bool(item.default_value)
                        })

                    # 添加更新字段
                    update_items = FormUpdateItem.objects.filter(form_config=config).order_by('sort_order')
                    for item in update_items:
                        # 跳过计算字段，计算字段由表达式自动生成
                        if item.input_type == 'calculated':
                            continue

                        if item.input_type == 'supplement':
                            # 补充框：主字段（新值和原值都需要用户填写）
                            headers.append({
                                'label': f'新{item.label}',
                                'hasDefaultValue': bool(item.new_default_value)
                            })
                            headers.append({
                                'label': f'原{item.label}',
                                'hasDefaultValue': bool(item.origin_default_value)
                            })

                            # 添加子字段（只有原值需要填写，新值自动查询）
                            sub_fields = item.sub_fields or []
                            for sub_field in sub_fields:
                                if isinstance(sub_field, dict):
                                    sub_label = sub_field.get('label', sub_field.get('bindingKey', ''))
                                else:
                                    sub_label = str(sub_field)
                                headers.append({
                                    'label': f'原{sub_label}',
                                    'hasDefaultValue': False
                                })
                        else:
                            # 普通字段：新值和原值
                            headers.append({
                                'label': f'新{item.label}',
                                'hasDefaultValue': bool(item.new_default_value)
                            })
                            headers.append({
                                'label': f'原{item.label}',
                                'hasDefaultValue': bool(item.origin_default_value)
                            })

                    # 写入表头
                    for col_idx, header_info in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header_info['label'])
                        cell.alignment = header_alignment
                        
                        # 根据是否有默认值设置背景色和字体颜色
                        if header_info['hasDefaultValue']:
                            # 有默认值：黄色背景 + 黑色字体
                            cell.font = header_font_black
                            cell.fill = default_value_fill
                        else:
                            # 无默认值：蓝色背景 + 白色字体
                            cell.font = header_font_white
                            cell.fill = header_fill

                        # 设置列宽 - 根据表头自适应
                        from openpyxl.utils import get_column_letter
                        column_letter = get_column_letter(col_idx)
                        
                        # 获取表头的长度（中文按2个字符计算）
                        header_value = header_info['label']
                        max_length = len(str(header_value).encode('utf-8'))
                        
                        # 设置列宽：最大长度 + 额外空间（2-4个字符），最小10，最大50
                        adjusted_width = min(max(max_length + 4, 10), 50)
                        ws.column_dimensions[column_letter].width = adjusted_width

                except FormConfig.DoesNotExist:
                    continue

            # 保存Excel到内存
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # 返回Excel文件
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename={file_prefix}_合并模板.xlsx'

            return response

        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'JSON 解析失败'
            }, status=400)
        except Exception as e:
            import traceback
            print(f"生成模板异常: {str(e)}")
            print(traceback.format_exc())
            return JsonResponse({
                'success': False,
                'message': f'生成失败：{str(e)}'
            }, status=500)

    return JsonResponse({
        'success': False,
        'message': '仅支持 POST 请求'
    }, status=405)
