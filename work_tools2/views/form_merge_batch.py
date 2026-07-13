import json
import os
from datetime import datetime
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook, Workbook
from work_tools2.models import FormConfig, FormQueryItem, FormUpdateItem, DatabaseIPConfig
from work_tools2.path_utils import get_save_path_from_config
from work_tools2.views.usage_views import record_form_usage


def _build_form_config_data(config):
    """从 FormConfig ORM 对象构建 query/update 配置数据"""
    query_items_data = []
    for item in FormQueryItem.objects.filter(form_config=config).order_by('sort_order'):
        query_items_data.append({
            'label': item.label,
            'bindingKey': item.binding_key,
            'type': item.field_type,
            'defaultValue': item.default_value,
            'ValidRule': item.valid_rule,
            'connectedTable': item.connected_table or [],
            'expressions': item.expressions or {},
        })

    update_items_data = []
    for item in FormUpdateItem.objects.filter(form_config=config).order_by('sort_order'):
        update_item = {
            'label': item.label,
            'bindingKey': item.binding_key,
            'inputType': item.input_type,
            'type': item.field_type,
            'newDefaultValue': item.new_default_value,
            'originDefaultValue': item.origin_default_value,
            'newValidRule': item.new_valid_rule,
            'originValidRule': item.origin_valid_rule,
            'mainTable': item.main_table,
            'mainField': item.main_field,
            'subFields': item.sub_fields or [],
            'connectedTable': item.connected_table or [],
        }
        if item.input_type == 'calculated':
            update_item['expressions'] = item.expressions or {}
            update_item['splitExpression'] = item.split_expression or False
            update_item['backwardExpressions'] = item.backward_expressions or {}
        if item.component_name:
            from work_tools2.models import ComponentConfig
            component = ComponentConfig.objects.filter(name=item.component_name).first()
            if component:
                update_item['options'] = component.options
            else:
                update_item['options'] = item.options or []
        else:
            update_item['options'] = item.options or []
        update_items_data.append(update_item)

    return query_items_data, update_items_data


def _build_worksheet_from_rows(ws, query_items, update_items, rows):
    """根据配置和在线编辑行数据构造 openpyxl worksheet"""
    # 构造表头：查询字段 -> 所有新值列 -> 所有原值列
    headers = []
    for item in query_items:
        headers.append(item.get('label', ''))

    new_headers = []
    origin_headers = []
    for item in update_items:
        input_type = item.get('inputType', '')
        if input_type == 'calculated':
            continue
        if input_type == 'supplement':
            new_headers.append(f'新{item.get("label", "")}')
            origin_headers.append(f'原{item.get("label", "")}')
            for sf in item.get('subFields', []):
                if isinstance(sf, dict) and sf.get('type') == 'auxiliary':
                    sub_label = sf.get('label', sf.get('bindingKey', ''))
                    new_headers.append(f'新{sub_label}')
                    origin_headers.append(f'原{sub_label}')
        else:
            new_headers.append(f'新{item.get("label", "")}')
            origin_headers.append(f'原{item.get("label", "")}')
    headers.extend(new_headers)
    headers.extend(origin_headers)

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # 写入数据行
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, '')
            ws.cell(row=row_idx, column=col_idx, value=value)

    return headers


@csrf_exempt
def batch_import_merge(request):
    """批量导入并合并多个表单的SQL（支持文件上传和在线编辑器JSON数据）"""
    if request.method == 'POST':
        try:
            content_type = request.content_type or ''
            is_json = 'application/json' in content_type

            if is_json:
                # 在线编辑器模式
                data = json.loads(request.body)
                query_values = data.get('queryValues', {})
                rows_per_sheet = data.get('rowsPerSheet', {})
                if not rows_per_sheet:
                    return JsonResponse({'success': False, 'message': '没有数据可导入'}, status=400)
                wb = None
            else:
                # 文件上传模式
                file = request.FILES.get('file')
                form_ids_json = request.POST.get('formIds')
                query_values_json = request.POST.get('queryValues')

                if not file or not form_ids_json:
                    return JsonResponse({'success': False, 'message': '缺少文件或表单ID参数'}, status=400)

                form_ids = json.loads(form_ids_json)
                if not form_ids or len(form_ids) == 0:
                    return JsonResponse({'success': False, 'message': '请至少选择一个表单'}, status=400)

                query_values = {}
                if query_values_json:
                    try:
                        query_values = json.loads(query_values_json)
                    except json.JSONDecodeError:
                        query_values = {}

                wb = load_workbook(file)
                rows_per_sheet = None

            # 验证公共字段
            common_fields = ['filePrefix', 'onesLink', 'dynamicNo', 'ops_remark']
            for field_name in common_fields:
                value_data = query_values.get(field_name, {})
                if isinstance(value_data, dict):
                    value = value_data.get('value', '')
                elif isinstance(value_data, str):
                    value = value_data
                else:
                    value = ''

                if not value or str(value).strip() == '':
                    return JsonResponse({'success': False, 'message': f'{field_name}不能为空'}, status=400)

            # 存储所有SQL语句
            all_forward_sqls = []
            all_backward_sqls = []
            failed_sheets = []
            success_count = 0
            total_processed = 0

            # 在线编辑器模式下保留每个 Sheet 的 worksheet，用于生成错误报告
            json_mode_sheets = {}
            # 在线编辑器模式下记录实际处理到的表单ID（用于生成SQL文件时写入数据库信息）
            json_form_ids = []

            def _extract_row_errors(ws):
                """从 worksheet 的失败原因列提取行级错误"""
                headers = {}
                for col in range(1, ws.max_column + 1):
                    val = ws.cell(row=1, column=col).value
                    if val is not None:
                        headers[str(val).strip()] = col
                error_col = headers.get('失败原因')
                if not error_col:
                    return []
                errors = []
                for row_idx in range(2, ws.max_row + 1):
                    val = ws.cell(row=row_idx, column=error_col).value
                    if val is not None and str(val).strip():
                        errors.append({
                            'rowIndex': row_idx,
                            'failReason': str(val).strip()
                        })
                return errors

            if is_json:
                # 在线编辑器模式：每个 Sheet 名对应一个表单名
                for sheet_name, rows in rows_per_sheet.items():
                    total_processed += 1
                    config = FormConfig.objects.filter(form_name=sheet_name).first()
                    if not config:
                        failed_sheets.append({
                            'sheet': sheet_name,
                            'error': f'未找到表单配置：{sheet_name}'
                        })
                        continue

                    try:
                        query_items_data, update_items_data = _build_form_config_data(config)

                        # 构造虚拟 worksheet
                        temp_wb = Workbook()
                        ws = temp_wb.active
                        ws.title = sheet_name
                        _build_worksheet_from_rows(ws, query_items_data, update_items_data, rows)
                        json_mode_sheets[sheet_name] = ws

                        from work_tools2.views.dynamic_views import process_single_sheet_import
                        result = process_single_sheet_import(
                            ws,
                            query_items_data,
                            update_items_data,
                            query_values,
                            config.form_name,
                            config.table_name_list,
                            config.query_mode,
                            config.append_ops_remark,
                            config.table_aliases
                        )

                        row_errors = _extract_row_errors(ws)

                        if result['success']:
                            all_forward_sqls.extend(result['forward_sqls'])
                            all_backward_sqls.extend(result['backward_sqls'])
                            success_count += 1
                            json_form_ids.append(config.id)
                            record_form_usage(config, source='merge')
                        else:
                            failed_sheets.append({
                                'sheet': sheet_name,
                                'error': result.get('message', '处理失败'),
                                'rowErrors': row_errors
                            })
                    except Exception as e:
                        import traceback
                        error_detail = traceback.format_exc()
                        print(f"处理Sheet '{sheet_name}' 失败: {error_detail}")
                        failed_sheets.append({
                            'sheet': sheet_name,
                            'error': f'处理失败：{str(e)}'
                        })
            else:
                # 文件上传模式
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    total_processed += 1

                    config = FormConfig.objects.filter(form_name=sheet_name).first()
                    if not config:
                        failed_sheets.append({
                            'sheet': sheet_name,
                            'error': f'未找到表单配置：{sheet_name}'
                        })
                        continue

                    try:
                        query_items_data, update_items_data = _build_form_config_data(config)

                        from work_tools2.views.dynamic_views import process_single_sheet_import
                        result = process_single_sheet_import(
                            ws,
                            query_items_data,
                            update_items_data,
                            query_values,
                            config.form_name,
                            config.table_name_list,
                            config.query_mode,
                            config.append_ops_remark,
                            config.table_aliases
                        )

                        row_errors = _extract_row_errors(ws)

                        if result['success']:
                            all_forward_sqls.extend(result['forward_sqls'])
                            all_backward_sqls.extend(result['backward_sqls'])
                            success_count += 1
                            record_form_usage(config, source='merge')
                        else:
                            failed_sheets.append({
                                'sheet': sheet_name,
                                'error': result.get('message', '处理失败'),
                                'rowErrors': row_errors
                            })
                    except Exception as e:
                        import traceback
                        error_detail = traceback.format_exc()
                        print(f"处理Sheet '{sheet_name}' 失败: {error_detail}")
                        failed_sheets.append({
                            'sheet': sheet_name,
                            'error': f'处理失败：{str(e)}'
                        })
            
            # 合并所有SQL语句
            now = datetime.now()
            
            # 使用路径配置获取保存路径
            save_dir = get_save_path_from_config()
            print(f"[DEBUG] 表单合并SQL文件保存路径: {save_dir}")
            
            # 获取公共字段值
            dynamic_no = ''
            file_prefix = 'merge'
            if query_values:
                dynamic_no_data = query_values.get('dynamicNo', '')
                if isinstance(dynamic_no_data, dict):
                    dynamic_no = dynamic_no_data.get('value', '')
                elif isinstance(dynamic_no_data, str):
                    dynamic_no = dynamic_no_data
                
                file_prefix_data = query_values.get('filePrefix', '')
                if isinstance(file_prefix_data, dict):
                    file_prefix = file_prefix_data.get('value', 'merge')
                elif isinstance(file_prefix_data, str):
                    file_prefix = file_prefix_data
            
            fail_count = len(failed_sheets)
            
            # 判断是否全部成功
            all_success = (fail_count == 0 and success_count > 0)
            
            # 如果有失败的Sheet，生成带有错误信息的Excel文件（包含所有Sheet）
            excel_file_path = None
            if not all_success:
                from openpyxl.styles import Font, Alignment, PatternFill

                # 创建结果Workbook
                wb_result = Workbook()
                
                # 删除默认的sheet
                if 'Sheet' in wb_result.sheetnames:
                    del wb_result['Sheet']

                # 复制所有Sheet（包括成功和失败的）- process_single_sheet_import已经在每行标记了具体的错误信息
                source_sheets = json_mode_sheets if is_json else {name: wb[name] for name in wb.sheetnames}
                for sheet_name, original_ws in source_sheets.items():
                    # 复制Sheet（Sheet名称最多31字符）
                    result_ws = wb_result.create_sheet(title=sheet_name[:31])

                    # 复制所有数据和样式（保留process_single_sheet_import写入的行级错误信息）
                    for row in original_ws.iter_rows():
                        for cell in row:
                            new_cell = result_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                            if cell.font:
                                new_cell.font = Font(
                                    name=cell.font.name,
                                    size=cell.font.size,
                                    bold=cell.font.bold,
                                    italic=cell.font.italic
                                )
                            if cell.fill and cell.fill.patternType:
                                from openpyxl.styles.colors import Color
                                # 安全地获取颜色值，处理不同的颜色类型
                                def get_color_obj(color_attr):
                                    """将颜色属性转换为 Color 对象"""
                                    if color_attr is None:
                                        return None
                                    # 如果已经是 Color 对象，直接返回
                                    if isinstance(color_attr, Color):
                                        return color_attr
                                    # 如果有 rgb 属性且是字符串，使用它
                                    if hasattr(color_attr, 'rgb'):
                                        rgb_value = color_attr.rgb
                                        if isinstance(rgb_value, str):
                                            return Color(rgb=rgb_value)
                                    # 如果是字符串，直接使用
                                    if isinstance(color_attr, str):
                                        return Color(rgb=color_attr)
                                    # 其他情况返回 None
                                    return None
                                
                                start_color = get_color_obj(cell.fill.start_color)
                                end_color = get_color_obj(cell.fill.end_color)
                                
                                new_cell.fill = PatternFill(
                                    start_color=start_color,
                                    end_color=end_color,
                                    fill_type=cell.fill.fill_type or 'solid'
                                )
                            if cell.alignment:
                                new_cell.alignment = Alignment(
                                    horizontal=cell.alignment.horizontal,
                                    vertical=cell.alignment.vertical,
                                    wrap_text=cell.alignment.wrapText
                                )

                # 保存Excel文件
                excel_filename = f"{dynamic_no}_导入失败_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
                excel_filepath = os.path.join(save_dir, excel_filename)
                wb_result.save(excel_filepath)
                excel_file_path = excel_filepath
            
            # 如果全部成功，生成SQL文件
            sql_file_path = None
            if all_success and (all_forward_sqls or all_backward_sqls):
                # 文件名格式：编号_文件名.sql
                sql_filename = f"{dynamic_no}_{file_prefix}.sql"
                sql_filepath = os.path.join(save_dir, sql_filename)
                
                # 生成合并的SQL文件内容
                timestamp = now.strftime('%Y%m%d_%H%M%S')
                forward_sql_content = generate_merged_sql_file(all_forward_sqls, file_prefix, timestamp, is_forward=True)
                backward_sql_content = generate_merged_sql_file(all_backward_sqls, file_prefix, timestamp, is_forward=False)
                
                # 收集所有表单的databaseIpIds（去重）
                all_database_ip_ids = set()
                source_form_ids = json_form_ids if is_json else form_ids
                for form_id in source_form_ids:
                    form_config = FormConfig.objects.filter(id=form_id).first()
                    if form_config and form_config.database_ip_ids:
                        all_database_ip_ids.update(form_config.database_ip_ids)
                
                # 写入SQL文件
                sql_content = []
                
                sql_content.append("1.执行语句")
                sql_content.append(forward_sql_content)
                sql_content.append("")
                sql_content.append("2.回退语句")
                sql_content.append(backward_sql_content)
                
                # 添加数据库信息
                if all_database_ip_ids:
                    db_configs = DatabaseIPConfig.objects.filter(id__in=all_database_ip_ids).order_by('id')
                    if db_configs:
                        sql_content.append("3.数据库")
                        for db_config in db_configs:
                            sql_content.append(f"ip：{db_config.ip_address}")
                            sql_content.append(f"库名：{db_config.database_name}")
                            sql_content.append("")
                
                with open(sql_filepath, 'w', encoding='utf-8') as f:
                    f.write('\n'.join(sql_content))
                
                sql_file_path = sql_filepath
            
            # 返回结果
            if all_success:
                return JsonResponse({
                    'success': True,
                    'message': f'批量导入成功！共处理{total_processed}个Sheet，全部成功',
                    'sqlFilePath': sql_file_path,
                    'excelFilePath': None,
                    'totalSheets': total_processed,
                    'successCount': success_count,
                    'failCount': 0,
                    'failed_sheets': []
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': f'批量导入失败！共处理{total_processed}个Sheet，成功{success_count}个，失败{fail_count}个',
                    'sqlFilePath': None,
                    'excelFilePath': excel_file_path,
                    'totalSheets': total_processed,
                    'successCount': success_count,
                    'failCount': fail_count,
                    'failed_sheets': [{'sheet': f['sheet'], 'error': f['error']} for f in failed_sheets]
                })

        except json.JSONDecodeError as e:
            return JsonResponse({'success': False, 'message': f'JSON 解析失败：{str(e)}'}, status=400)
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(f"批量导入合并异常: {error_detail}")
            return JsonResponse({'success': False, 'message': f'服务器错误：{str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'message': '仅支持 POST 请求'}, status=405)


def generate_merged_sql_file(sql_statements, file_prefix, timestamp, is_forward=True):
    """生成合并的SQL文件内容"""
    lines = []
    for idx, sql_item in enumerate(sql_statements, 1):
        # 如果sql_item是字典，使用formatted字段，否则直接使用
        if isinstance(sql_item, dict):
            sql = sql_item.get('formatted', sql_item.get('raw', ''))
        else:
            sql = str(sql_item)

        lines.append(sql + ';')
        lines.append('')
    
    return '\n'.join(lines)
