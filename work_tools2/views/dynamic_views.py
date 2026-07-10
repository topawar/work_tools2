import json
import os
import re
import time
import sqlparse
from datetime import datetime
from io import BytesIO
from collections import defaultdict

from django.http import HttpResponse, JsonResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from work_tools2.models import DatabaseIPConfig, FormConfig
from work_tools2.path_utils import get_save_path_from_config
from work_tools2.views.usage_views import record_form_usage


# ==================== 工具函数 ====================
def format_sql(sql):
    """格式化 SQL 语句 - 优化可读性"""
    try:
        import sqlparse
        import re

        # 先使用 sqlparse 进行基本格式化，但不重新缩进
        formatted = sqlparse.format(
            sql,
            reindent=False,  # 不自动重新缩进，我们自己控制
            keyword_case='upper',
            identifier_case='upper',
            strip_comments=True
        )

        # 处理 IN 子句：如果值很多，进行换行格式化
        def format_in_clause(match):
            """格式化 IN 子句，当值超过3个时换行显示"""
            field_name = match.group(1)
            values_str = match.group(2)

            # 提取所有值（需要处理元组格式）
            values = []
            current_value = ""
            paren_depth = 0
            
            for char in values_str:
                if char == '(':
                    paren_depth += 1
                    current_value += char
                elif char == ')':
                    paren_depth -= 1
                    current_value += char
                elif char == ',' and paren_depth == 0:
                    # 只有在括号外层的逗号才分割
                    if current_value.strip():
                        values.append(current_value.strip())
                    current_value = ""
                else:
                    current_value += char
            
            # 添加最后一个值
            if current_value.strip():
                values.append(current_value.strip())

            if len(values) <= 3:
                # 值不多，保持在一行
                return f"{field_name} IN ({values_str})"
            else:
                # 值很多，换行显示
                indent = " " * 4  # 基础缩进
                values_indent = " " * 6  # 值的缩进

                # 每行显示3个值
                formatted_values = []
                for i in range(0, len(values), 3):
                    chunk = values[i:i+3]
                    formatted_values.append(f"{values_indent}{', '.join(chunk)}")

                result = f"{field_name} IN (\n"
                result += ",\n".join(formatted_values)
                result += f"\n{indent})"
                return result

        # 匹配 IN 子句：支持单字段和联合条件，以及嵌套括号
        # 使用非贪婪匹配和手动处理嵌套括号
        def find_in_clauses(text):
            """查找所有 IN 子句，支持嵌套括号"""
            results = []
            pattern = r'(\([A-Z_,\s]+\)|[A-Z_][A-Z0-9_]*)\s+IN\s*\('
            
            for match in re.finditer(pattern, text, re.IGNORECASE):
                start_pos = match.end()  # IN ( 之后的位置
                field_name = match.group(1)
                
                # 手动查找匹配的右括号
                paren_depth = 1
                pos = start_pos
                while pos < len(text) and paren_depth > 0:
                    if text[pos] == '(':
                        paren_depth += 1
                    elif text[pos] == ')':
                        paren_depth -= 1
                    pos += 1
                
                if paren_depth == 0:
                    values_str = text[start_pos:pos-1]
                    results.append((match.start(), pos, field_name, values_str))
            
            return results
        
        # 查找并替换所有 IN 子句
        in_clauses = find_in_clauses(formatted)
        # 从后往前替换，避免位置偏移
        for start, end, field_name, values_str in reversed(in_clauses):
            replacement = format_in_clause(type('obj', (object,), {'group': lambda self, x: field_name if x == 1 else values_str})())
            formatted = formatted[:start] + replacement + formatted[end:]

        # 手动格式化：按关键字分割并重新组织
        lines = []

        # 将SQL按关键字分割（注意：IN 子句可能已经包含换行）
        # 先按行分割，再处理每行
        raw_lines = formatted.split('\n')

        in_set_clause = False  # 标记是否在 SET 子句中

        for line in raw_lines:
            stripped = line.strip()
            if not stripped:
                continue

            # 如果这一行已经是 IN 子句的一部分（以 ) 结尾且前面有 IN），直接保留
            if stripped.startswith(')') or (stripped.endswith(',') and '(' in line):
                lines.append(line)
                continue

            # 处理普通行
            parts = stripped.split()
            i = 0
            current_line = ""

            while i < len(parts):
                word = parts[i]
                word_upper = word.upper()

                # UPDATE 关键字：新起一行
                if word_upper == 'UPDATE':
                    if current_line:
                        lines.append(current_line)
                    current_line = word
                    in_set_clause = False
                    i += 1
                    continue

                # SET 关键字：新起一行，缩进
                if word_upper == 'SET':
                    if current_line:
                        lines.append(current_line)
                    current_line = f"  {word}"
                    in_set_clause = True
                    i += 1
                    continue

                # WHERE 关键字：新起一行
                if word_upper == 'WHERE':
                    if current_line:
                        lines.append(current_line)
                    current_line = f"  {word}"
                    in_set_clause = False
                    i += 1
                    continue

                # OR 关键字：新起一行，增加缩进
                if word_upper == 'OR':
                    if current_line:
                        lines.append(current_line)
                    current_line = f"    {word}"
                    i += 1
                    continue

                # 如果在 SET 子句中，遇到逗号说明是下一个字段，需要换行
                if in_set_clause and word.endswith(','):
                    # 添加当前字段（带逗号）
                    current_line += f" {word}"
                    lines.append(current_line)
                    current_line = "    "  # 下一行缩进
                    i += 1
                    continue

                # 其他内容：添加到当前行
                current_line += f" {word}"
                i += 1

            if current_line:
                lines.append(current_line)

        return '\n'.join(lines)

    except ImportError:
        # 如果 sqlparse 未安装，返回原始SQL
        return sql


def handle_field_value(field_name, value, valid_rule):
    """根据 ValidRule 处理字段值"""
    is_empty = (value is None or value == '' or value == 'null' or value == 'NULL')

    if valid_rule == 'required':
        if is_empty:
            return None
        else:
            return f"{field_name} = '{value}'"
    elif valid_rule == 'requiredReverse':
        if not is_empty:
            return f"{field_name} = '{value}'"
        else:
            return None
    elif valid_rule == 'defaultNull':
        # 空值时设置为空字符串，而不是 NULL
        if is_empty:
            return f"{field_name} = ''"
        else:
            return f"{field_name} = '{value}'"
    elif valid_rule == 'defaultField':
        if is_empty:
            return f"{field_name} = {field_name}"
        else:
            return f"{field_name} = '{value}'"
    else:
        if is_empty:
            return None
        else:
            return f"{field_name} = '{value}'"


def evaluate_calculated_expression(expression, binding_key, form_values, update_items, query_items, is_forward=True, context_table=''):
    """计算字段/计算查询条件表达式求值

    将表达式中的 ${VAR} 替换为对应字段的值。
    - 普通查询字段使用 value 作为双向值
    - 差异条件查询字段根据 is_forward 使用 newValue/originValue
    - 更新字段根据 is_forward 使用 newValue/originValue
    - context_table 用于同名字段多表时，优先选择关联到当前表的查询字段
    返回: (求值后的表达式字符串, 是否有效)
    """
    import re

    if not expression or not binding_key:
        return None, False

    variables = re.findall(r'\$\{\s*([A-Za-z_][A-Za-z0-9_]*)\s*\}', expression)
    result_expression = expression
    has_any_valid_value = False
    all_variables_processed = True

    # 没有占位变量的表达式直接视为有效（如纯子查询）
    if not variables and expression.strip():
        has_any_valid_value = True

    for var_name in variables:
        field_item = None
        var_name_upper = var_name.upper()

        for ui in update_items:
            if ui.get('bindingKey', '').upper() == var_name_upper:
                field_item = ui
                break

        if not field_item:
            query_candidates = [
                qi for qi in query_items
                if qi.get('bindingKey', '').upper() == var_name_upper
                and qi.get('type') != 'difference_condition'
            ]
            if query_candidates:
                # 同名字段多表时，优先选择关联到当前计算表的查询字段
                if context_table:
                    matched = [qi for qi in query_candidates if context_table in qi.get('connectedTable', [])]
                    if matched:
                        field_item = matched[0]
                # 若当前计算表未直接关联，则根据表达式中引用的表名进行二次匹配
                if not field_item:
                    expr_tables = _extract_table_names_from_expression(expression)
                    if expr_tables:
                        matched = [
                            qi for qi in query_candidates
                            if any(t in expr_tables for t in qi.get('connectedTable', []))
                        ]
                        if matched:
                            field_item = matched[0]
                if not field_item:
                    field_item = query_candidates[0]

        if not field_item:
            print(f"[警告] 计算字段 {binding_key} 的表达式中引用了未定义的变量: {var_name}")
            continue

        field_key = field_item['bindingKey']
        field_type = field_item.get('type', 'text')
        field_connected_tables = field_item.get('connectedTable', [])

        # 构造表查找顺序：优先当前计算上下文表
        lookup_tables = list(field_connected_tables)
        if context_table and context_table in lookup_tables:
            lookup_tables.remove(context_table)
            lookup_tables.insert(0, context_table)

        value_data = None
        for table in lookup_tables:
            # 更新字段使用 {bindingKey}_{table}
            unique_key = f"{field_key}_{table}"
            if unique_key in form_values:
                value_data = form_values[unique_key]
                break
            # 查询字段使用 query_{bindingKey}_{table}
            query_key = f"query_{field_key}_{table}"
            if query_key in form_values:
                value_data = form_values[query_key]
                break
        if value_data is None:
            # 兼容旧数据/未分表存储
            value_data = form_values.get(field_key, {}) or form_values.get(f"query_{field_key}", {})

        if 'value' in value_data:
            # 普通查询字段：没有验证规则概念，直接使用值
            new_value = value_data.get('value', '')
            origin_value = value_data.get('value', '')
            valid_rule = 'defaultField'
        else:
            new_value = value_data.get('newValue', '')
            origin_value = value_data.get('originValue', '')
            # 使用被引用字段自身的验证规则
            valid_rule = field_item.get('newValidRule', 'defaultField')

        is_number_type = (field_type == 'number')
        value = new_value if is_forward else origin_value

        if value:
            has_any_valid_value = True
            if is_number_type:
                result_expression = re.sub(
                    r'\$\{\s*' + re.escape(var_name) + r'\s*\}',
                    value,
                    result_expression
                )
            else:
                result_expression = re.sub(
                    r'\$\{\s*' + re.escape(var_name) + r'\s*\}',
                    f"'{value}'",
                    result_expression
                )
        else:
            if valid_rule == 'defaultField':
                result_expression = re.sub(
                    r'\$\{\s*' + re.escape(var_name) + r'\s*\}',
                    var_name.upper(),
                    result_expression
                )
            elif valid_rule == 'required':
                all_variables_processed = False
                has_any_valid_value = False
                break
            else:
                replacement = '0' if is_number_type else "''"
                result_expression = re.sub(
                    r'\$\{\s*' + re.escape(var_name) + r'\s*\}',
                    replacement,
                    result_expression
                )

    if all_variables_processed and has_any_valid_value:
        return result_expression, True
    return None, False


def _get_expression_by_table(expressions, table_name):
    """从表达式数组中按 tableName 查找表达式，兼容旧 dict 格式"""
    if not expressions:
        return ''
    if isinstance(expressions, dict):
        return expressions.get(table_name, '') or ''
    if isinstance(expressions, list):
        for entry in expressions:
            if isinstance(entry, dict) and entry.get('tableName') == table_name:
                return entry.get('expression', '') or ''
    return ''


def generate_update_sql(config, form_values):
    """根据配置生成 SQL UPDATE 语句"""
    forward_sqls = []
    backward_sqls = []
    missing_field_labels = set()  # 收集所有缺失字段的label（去重）
    missing_update_table_labels = {}  # 查询条件满足但无有效更新字段的表: {table_name: [labels]}

    table_name_list = config.get('tableNameList', [])
    table_aliases = config.get('tableAliases', {}) or {}
    table_joins = config.get('tableJoins', []) or []
    query_items = config.get('queryItems', [])
    update_items = config.get('updateItems', [])

    # 构建 JOIN 映射
    join_main_map = {ji['main_table']: ji for ji in table_joins}
    join_table_set = {ji['join_table'] for ji in table_joins}

    # 获取查询模式：strict(严格) 或 loose(宽松)
    query_mode = config.get('queryMode', 'strict')
    
    # 获取是否拼接操作备注的配置，默认为 True
    append_ops_remark = config.get('appendOpsRemark', True)

    # 获取操作备注
    ops_remark = form_values.get('ops_remark', '')
    if isinstance(ops_remark, dict):
        ops_remark = ops_remark.get('value', '')
    elif not ops_remark:
        ops_remark = ''

    for table_name in table_name_list:
        # 跳过仅作为 JOIN 从表的表，避免生成独立 UPDATE
        if table_name in join_table_set:
            continue

        join_info = join_main_map.get(table_name)

        # 当前主表的别名（如果有）
        main_alias = table_aliases.get(table_name, '')
        if join_info:
            main_alias = join_info.get('main_alias') or main_alias

        main_col_prefix = f"{main_alias}." if main_alias else ""

        # JOIN 相关信息
        join_table = ''
        join_alias = ''
        join_col_prefix = ''
        join_clause = ''
        on_clause = ''
        if join_info:
            join_table = join_info['join_table']
            join_alias = join_info.get('join_alias', '')
            join_col_prefix = f"{join_alias}." if join_alias else ""
            join_type = join_info.get('join_type', 'JOIN')
            join_clause = f"{join_type} {join_table} {join_alias}".strip()
            on_conditions = join_info.get('on_conditions', [])
            on_clause = ' AND '.join(on_conditions) if on_conditions else ''

        # 第一步：找出该表及关联 JOIN 表的所有查询字段
        table_query_fields = []
        for item in query_items:
            connected_tables = item.get('connectedTable', [])
            if table_name in connected_tables:
                table_query_fields.append(item)
            elif join_info and join_table in connected_tables:
                table_query_fields.append(item)

        # 第二步：先收集该表所有的SET子句，判断是否真的有更新操作
        forward_set_clauses = []
        backward_set_clauses = []

        for item in update_items:
            connected_tables = item.get('connectedTable', [])
            new_valid_rule = item.get('newValidRule', '')
            origin_valid_rule = item.get('originValidRule', '')
            input_type = item.get('inputType', '')
            binding_key = item.get('bindingKey')

            # JOIN 场景下，SET 只更新主表字段
            if join_info and table_name not in connected_tables:
                continue
            # 非 JOIN 场景：如果该表不在此更新字段的关联表中，跳过
            if not join_info and table_name not in connected_tables:
                continue

            # 处理计算字段类型
            if input_type == 'calculated':
                expressions = item.get('expressions', []) or []
                split_expression = item.get('splitExpression', False)
                backward_expressions = item.get('backwardExpressions', []) or []
                binding_key = item.get('bindingKey', '')

                forward_expression_raw = _get_expression_by_table(expressions, table_name)
                backward_expression_raw = (
                    _get_expression_by_table(backward_expressions, table_name)
                    if split_expression
                    else forward_expression_raw
                )

                forward_expression, forward_valid = evaluate_calculated_expression(
                    forward_expression_raw.strip(), binding_key, form_values, update_items, query_items, is_forward=True, context_table=table_name
                )
                backward_expression, backward_valid = evaluate_calculated_expression(
                    backward_expression_raw.strip(), binding_key, form_values, update_items, query_items, is_forward=False, context_table=table_name
                )

                if forward_valid:
                    forward_set_clauses.append(f"{main_col_prefix}{binding_key} = {forward_expression}")
                if backward_valid:
                    backward_set_clauses.append(f"{main_col_prefix}{binding_key} = {backward_expression}")

            # 处理补充框类型
            elif input_type == 'supplement':
                parent_key = item.get('bindingKey')

                # 关键修复：对于同名字段绑定不同表的情况，需要从 form_values 中查找正确的值
                value_data = None
                if connected_tables:
                    for table in connected_tables:
                        unique_key = f"{parent_key}_{table}"
                        if unique_key in form_values:
                            value_data = form_values[unique_key]
                            break
                
                # 如果没找到，使用原始的 parent_key（向后兼容）
                if value_data is None:
                    value_data = form_values.get(parent_key, {})

                # 如果 form_values 中没有这个补充框，说明前端没有传输（原值为空），跳过
                if not value_data:
                    continue

                new_value = value_data.get('newValue', '')
                origin_value = value_data.get('originValue', '')

                forward_set_value = handle_field_value(f"{main_col_prefix}{parent_key}", new_value, new_valid_rule)
                if forward_set_value is not None and table_name in connected_tables:
                    forward_set_clauses.append(forward_set_value)

                backward_set_value = handle_field_value(f"{main_col_prefix}{parent_key}", origin_value, origin_valid_rule)
                if backward_set_value is not None and table_name in connected_tables:
                    backward_set_clauses.append(backward_set_value)

                sub_fields = item.get('subFields', [])
                main_table_for_sub = item.get('mainTable', '')
                for sub_field in sub_fields:
                    sub_binding_key = sub_field.get('bindingKey')
                    
                    # 子字段使用 main_table 查找（数据来自 main_table）
                    sub_value_data = None
                    if main_table_for_sub:
                        sub_unique_key = f"{sub_binding_key}_{main_table_for_sub}"
                        sub_value_data = form_values.get(sub_unique_key)
                    
                    # 回退：尝试 connected_tables
                    if sub_value_data is None and connected_tables:
                        for table in connected_tables:
                            unique_key = f"{sub_binding_key}_{table}"
                            if unique_key in form_values:
                                sub_value_data = form_values[unique_key]
                                break
                    
                    if sub_value_data is None:
                        sub_value_data = form_values.get(sub_binding_key, {})
                    
                    sub_new_value = sub_value_data.get('newValue', '')
                    sub_origin_value = sub_value_data.get('originValue', '')

                    if table_name in connected_tables:
                        forward_sub_set_value = handle_field_value(f"{main_col_prefix}{sub_binding_key}", sub_new_value, new_valid_rule)
                        if forward_sub_set_value is not None:
                            forward_set_clauses.append(forward_sub_set_value)

                        if sub_origin_value is None or sub_origin_value == '':
                            backward_sub_set_value = f"{main_col_prefix}{sub_binding_key} = ''"
                        else:
                            backward_sub_set_value = f"{main_col_prefix}{sub_binding_key} = '{sub_origin_value}'"

                        if backward_sub_set_value is not None:
                            backward_set_clauses.append(backward_sub_set_value)

            # 处理普通字段类型
            else:
                binding_key = item.get('bindingKey')

                # 关键修复：对于同名字段绑定不同表的情况，需要从 form_values 中查找正确的值
                # 优先尝试 bindingKey_tableName 格式，然后尝试原始的 bindingKey
                value_data = None
                if connected_tables:
                    # 先尝试 bindingKey_tableName 格式
                    for table in connected_tables:
                        unique_key = f"{binding_key}_{table}"
                        if unique_key in form_values:
                            value_data = form_values[unique_key]
                            break

                # 如果没找到，使用原始的 bindingKey（向后兼容）
                if value_data is None:
                    value_data = form_values.get(binding_key, {})

                new_value = value_data.get('newValue', '')
                origin_value = value_data.get('originValue', '')

                if table_name in connected_tables:
                    forward_set_value = handle_field_value(f"{main_col_prefix}{binding_key}", new_value, new_valid_rule)
                    if forward_set_value is not None:
                        forward_set_clauses.append(forward_set_value)

                    backward_set_value = handle_field_value(f"{main_col_prefix}{binding_key}", origin_value, origin_valid_rule)
                    if backward_set_value is not None:
                        backward_set_clauses.append(backward_set_value)

        # 第三步：如果该表没有任何SET子句，记录并跳过
        if not forward_set_clauses and not backward_set_clauses:
            # 收集该表关联的更新字段标签
            table_update_labels = []
            for item in update_items:
                if table_name in item.get('connectedTable', []):
                    table_update_labels.append(item.get('label', item.get('bindingKey', '')))
            labels_str = '、'.join(table_update_labels) if table_update_labels else '无'
            missing_update_table_labels[table_name] = table_update_labels
            continue

        # 第四步：有SET子句时，才根据查询模式收集WHERE条件
        forward_where_conditions = []
        backward_where_conditions = []

        if query_mode == 'loose':
            # 宽松模式：只收集有值的字段
            for item in table_query_fields:
                binding_key = item.get('bindingKey')
                field_type = item.get('type', 'text')
                connected_tables = item.get('connectedTable', [])
                # JOIN 场景下，仅关联 JOIN 表的字段使用 JOIN 表别名前缀
                if join_info and join_table in connected_tables and table_name not in connected_tables:
                    field_col_prefix = join_col_prefix
                else:
                    field_col_prefix = main_col_prefix
                value_data = _get_query_value_data(item.get('bindingKey'), connected_tables, form_values)

                if field_type == 'calculated' or field_type == 'subquery':
                    # 计算查询条件 / 子查询条件：按表达式求值
                    expressions = item.get('expressions', []) or []
                    split_expression = item.get('splitExpression', False)
                    backward_expressions = item.get('backwardExpressions', []) or []
                    expr_table = table_name if table_name in connected_tables else (join_table if join_info and join_table in connected_tables else table_name)
                    forward_expression_raw = _get_expression_by_table(expressions, expr_table)
                    backward_expression_raw = (
                        _get_expression_by_table(backward_expressions, expr_table)
                        if split_expression
                        else forward_expression_raw
                    )
                    if forward_expression_raw and binding_key:
                        fw_expr, fw_valid = evaluate_calculated_expression(
                            forward_expression_raw.strip(), binding_key, form_values, update_items, query_items, is_forward=True, context_table=expr_table
                        )
                        bw_expr, bw_valid = evaluate_calculated_expression(
                            backward_expression_raw.strip(), binding_key, form_values, update_items, query_items, is_forward=False, context_table=expr_table
                        )
                        if fw_valid:
                            if field_type == 'subquery':
                                forward_where_conditions.append(f"{field_col_prefix}{binding_key} IN ({fw_expr})")
                            else:
                                forward_where_conditions.append(fw_expr)
                        if bw_valid:
                            if field_type == 'subquery':
                                backward_where_conditions.append(f"{field_col_prefix}{binding_key} IN ({bw_expr})")
                            else:
                                backward_where_conditions.append(bw_expr)
                elif field_type == 'difference_condition':
                    # 差异条件：与页面输入框标签/占位符保持一致
                    # “新值”输入框占位符为“执行语句用”，用于正向 WHERE
                    # “原值”输入框占位符为“回退语句用”，用于反向 WHERE
                    new_value = value_data.get('newValue', '')
                    origin_value = value_data.get('originValue', '')
                    if new_value:
                        forward_where_conditions.append(f"{field_col_prefix}{binding_key} = '{new_value}'")
                    if origin_value:
                        backward_where_conditions.append(f"{field_col_prefix}{binding_key} = '{origin_value}'")
                else:
                    value = value_data.get('value', '')
                    if value:
                        forward_where_conditions.append(f"{field_col_prefix}{binding_key} = '{value}'")
                        backward_where_conditions.append(f"{field_col_prefix}{binding_key} = '{value}'")
        else:
            # 严格模式（默认）：要求所有关联字段都有值
            all_fields_have_value = True

            for item in table_query_fields:
                binding_key = item.get('bindingKey')
                field_type = item.get('type', 'text')
                connected_tables = item.get('connectedTable', [])
                # JOIN 场景下，仅关联 JOIN 表的字段使用 JOIN 表别名前缀
                if join_info and join_table in connected_tables and table_name not in connected_tables:
                    field_col_prefix = join_col_prefix
                else:
                    field_col_prefix = main_col_prefix
                value_data = _get_query_value_data(item.get('bindingKey'), connected_tables, form_values)

                if field_type == 'calculated' or field_type == 'subquery':
                    expressions = item.get('expressions', []) or []
                    split_expression = item.get('splitExpression', False)
                    backward_expressions = item.get('backwardExpressions', []) or []
                    # JOIN 场景下，字段可能关联到从表，需按实际关联表查找表达式
                    expr_table = table_name if table_name in connected_tables else (join_table if join_info and join_table in connected_tables else table_name)
                    forward_expression_raw = _get_expression_by_table(expressions, expr_table)
                    backward_expression_raw = (
                        _get_expression_by_table(backward_expressions, expr_table)
                        if split_expression
                        else forward_expression_raw
                    )
                    if not (forward_expression_raw and binding_key):
                        all_fields_have_value = False
                        break
                    fw_expr, fw_valid = evaluate_calculated_expression(
                        forward_expression_raw.strip(), binding_key, form_values, update_items, query_items, is_forward=True, context_table=expr_table
                    )
                    bw_expr, bw_valid = evaluate_calculated_expression(
                        backward_expression_raw.strip(), binding_key, form_values, update_items, query_items, is_forward=False, context_table=expr_table
                    )
                    if not fw_valid or not bw_valid:
                        all_fields_have_value = False
                        break
                    if field_type == 'subquery':
                        forward_where_conditions.append(f"{field_col_prefix}{binding_key} IN ({fw_expr})")
                        backward_where_conditions.append(f"{field_col_prefix}{binding_key} IN ({bw_expr})")
                    else:
                        forward_where_conditions.append(fw_expr)
                        backward_where_conditions.append(bw_expr)
                elif field_type == 'difference_condition':
                    new_value = value_data.get('newValue', '')
                    origin_value = value_data.get('originValue', '')
                    if not new_value or not origin_value:
                        all_fields_have_value = False
                        break
                    # 与页面标签/占位符保持一致：新值用于执行语句，原值用于回退语句
                    forward_where_conditions.append(f"{field_col_prefix}{binding_key} = '{new_value}'")
                    backward_where_conditions.append(f"{field_col_prefix}{binding_key} = '{origin_value}'")
                else:
                    value = value_data.get('value', '')
                    if not value:
                        all_fields_have_value = False
                        break
                    forward_where_conditions.append(f"{field_col_prefix}{binding_key} = '{value}'")
                    backward_where_conditions.append(f"{field_col_prefix}{binding_key} = '{value}'")

            # 严格模式下，如果有字段为空则处理
            if not all_fields_have_value:
                # 关键逻辑：只有当表有多个查询字段时，才记录缺失字段
                # 如果只有一个查询字段，静默跳过（不生成SQL即可）
                if len(table_query_fields) > 1:
                    for item in table_query_fields:
                        label = item.get('label', item.get('bindingKey'))
                        field_type = item.get('type', 'text')
                        value_data = _get_query_value_data(item.get('bindingKey'), item.get('connectedTable', []), form_values)

                        if field_type in ('calculated', 'subquery'):
                            # 计算字段/子查询字段为后端表达式，不提示用户填写
                            continue

                        has_value = False
                        if field_type == 'difference_condition':
                            if value_data.get('newValue') or value_data.get('originValue'):
                                has_value = True
                        else:
                            if value_data.get('value'):
                                has_value = True

                        if not has_value:
                            missing_field_labels.add(label)
                else:
                    # 只有一个查询字段且为空，静默跳过
                    pass
                continue

        # 第五步：至少有一个查询字段有值时才生成SQL
        # 宽松模式：如果没有任何查询字段有值，静默跳过
        if not forward_where_conditions and not backward_where_conditions:
            continue

        # 第六步：生成SQL语句
        forward_where_clause_str = ' AND '.join(forward_where_conditions)
        backward_where_clause_str = ' AND '.join(backward_where_conditions)

        if forward_set_clauses and forward_where_conditions:
            forward_set_clause_str = ', '.join(forward_set_clauses)
            # 根据配置决定是否添加操作备注
            if append_ops_remark and ops_remark:
                forward_set_clause_str += f", {main_col_prefix}ops_remark = '{ops_remark}'"
            # 构建表引用：单表或 JOIN
            main_table_ref = f"{table_name} {main_alias}".strip() if main_alias else table_name
            if join_info and join_clause and on_clause:
                table_ref = f"{main_table_ref} {join_clause} ON {on_clause}"
            else:
                table_ref = main_table_ref
            forward_sql = f"UPDATE {table_ref} SET {forward_set_clause_str} WHERE {forward_where_clause_str}"
            # 返回原始SQL和格式化后的SQL
            forward_sqls.append({
                'raw': forward_sql,
                'formatted': format_sql(forward_sql)
            })

        if backward_set_clauses and backward_where_conditions:
            backward_set_clause_str = ', '.join(backward_set_clauses)
            # 回退语句：根据配置决定是否清空操作备注
            if append_ops_remark:
                backward_set_clause_str += f", {main_col_prefix}ops_remark = ''"
            # 构建表引用：单表或 JOIN
            main_table_ref = f"{table_name} {main_alias}".strip() if main_alias else table_name
            if join_info and join_clause and on_clause:
                table_ref = f"{main_table_ref} {join_clause} ON {on_clause}"
            else:
                table_ref = main_table_ref
            backward_sql = f"UPDATE {table_ref} SET {backward_set_clause_str} WHERE {backward_where_clause_str}"
            # 返回原始SQL和格式化后的SQL
            backward_sqls.append({
                'raw': backward_sql,
                'formatted': format_sql(backward_sql)
            })

    # 返回结果
    has_sql = bool(forward_sqls or backward_sqls)
    
    # 将收集的缺失字段标签合并为错误信息
    # 关键逻辑：只有当没有任何SQL生成时，才返回错误信息
    missing_field_errors = []
    if missing_field_labels and not has_sql:
        missing_field_errors = [f"{label}未填写" for label in missing_field_labels]
    
    # 添加无更新字段的表的错误信息
    if not has_sql:
        for table_name, labels in missing_update_table_labels.items():
            labels_str = '、'.join(labels) if labels else '无'
            missing_field_errors.append(f"表 {table_name} 无有效更新字段（关联字段: {labels_str}）")

    return {
        'forward_sqls': forward_sqls,
        'backward_sqls': backward_sqls,
        'missing_field_errors': missing_field_errors
    }


def merge_where_clauses(where_clauses):
    """
    智能合并WHERE子句 - 使用元组IN子句
    - 将多个WHERE条件合并为 (field1, field2) IN ((val1, val2), (val3, val4))
    - 避免笛卡尔积问题，保持精确配对
    - 多字段时每个元组单独一行，提高可读性
    - 当某个字段在所有条件中取值相同时，会提取为独立的等值条件，使SQL更简短

    示例1（单字段）：
      输入: ["PURCHASE_SCHEME_NO = 'A'", "PURCHASE_SCHEME_NO = 'B'"]
      输出: "PURCHASE_SCHEME_NO IN ('A', 'B')"

    示例2（多字段）：
      输入: [
        "PURCHASE_SCHEME_NO = 'A' AND INQ_ID = 'X'",
        "PURCHASE_SCHEME_NO = 'B' AND INQ_ID = 'Y'"
      ]
      输出: |
        (PURCHASE_SCHEME_NO, INQ_ID) IN (
            ('A', 'X'),
            ('B', 'Y')
        )

    示例3（含相同值字段）：
      输入: [
        "CREATE_USER = 'U1' AND INQ_ID = 'A' AND ALIVE_FLAG = '1'",
        "CREATE_USER = 'U1' AND INQ_ID = 'B' AND ALIVE_FLAG = '1'"
      ]
      输出: "ALIVE_FLAG = '1' AND (CREATE_USER, INQ_ID) IN (\n    ('U1', 'A'),\n    ('U1', 'B')\n)"
    """
    if not where_clauses:
        return ''

    if len(where_clauses) == 1:
        return where_clauses[0]

    # 去重并保持顺序
    unique_clauses = list(dict.fromkeys(where_clauses))

    if len(unique_clauses) == 1:
        return unique_clauses[0]

    # 解析每个WHERE子句，提取字段和值
    parsed_conditions = []
    for clause in unique_clauses:
        conditions = parse_where_clause(clause)
        if conditions:
            parsed_conditions.append(conditions)

    if not parsed_conditions:
        # 如果无法解析，回退到OR连接
        return ' OR '.join([f"({wc})" for wc in unique_clauses])

    # 检查所有条件是否包含相同的字段
    first_fields = set(parsed_conditions[0].keys())
    all_same_fields = all(set(cond.keys()) == first_fields for cond in parsed_conditions)

    if not all_same_fields:
        # 字段不一致，回退到OR连接
        return ' OR '.join([f"({wc})" for wc in unique_clauses])

    # 所有条件字段一致，可以合并为IN子句
    fields = list(parsed_conditions[0].keys())

    # 识别所有值相同的字段，提取为等值条件（如 ALIVE_FLAG = '1'）
    common_conditions = {}
    variable_fields = []

    for field in fields:
        values = [cond[field] for cond in parsed_conditions]
        if all(v == values[0] for v in values):
            common_conditions[field] = values[0]
        else:
            variable_fields.append(field)

    # 构建等值条件部分
    equality_parts = [f"{field} = '{value}'" for field, value in common_conditions.items()]

    # 构建IN条件部分
    in_parts = []
    if len(variable_fields) == 1:
        # 单字段：field IN ('val1', 'val2', ...)
        field = variable_fields[0]
        values = [cond[field] for cond in parsed_conditions]
        values_str = ', '.join([f"'{v}'" for v in values])
        in_parts.append(f"{field} IN ({values_str})")
    elif len(variable_fields) > 1:
        # 多字段：(field1, field2) IN (
        #     ('val1a', 'val1b'),
        #     ('val2a', 'val2b')
        # )
        fields_str = ', '.join(variable_fields)
        tuples = []
        for cond in parsed_conditions:
            tuple_values = ', '.join([f"'{cond[f]}'" for f in variable_fields])
            tuples.append(f"    ({tuple_values})")
        tuples_str = ',\n'.join(tuples)
        in_parts.append(f"({fields_str}) IN (\n{tuples_str}\n)")

    # 组合条件：等值条件在前，IN条件在后
    all_parts = equality_parts + in_parts
    if all_parts:
        return ' AND '.join(all_parts)
    else:
        # 兜底：返回第一个条件
        return unique_clauses[0]


def parse_where_clause(clause):
    """
    解析WHERE子句，提取字段和值的映射
    例如: "PURCHASE_SCHEME_NO = 'A' AND INQ_ID = 'X'" 
    返回: {'PURCHASE_SCHEME_NO': 'A', 'INQ_ID': 'X'}
    """
    conditions = {}

    # 按 AND 分割
    parts = clause.split(' AND ')

    for part in parts:
        part = part.strip()
        # 匹配 [alias.]field = 'value' 或 [alias.]field = value
        import re
        match = re.match(r"(?:([a-zA-Z_][a-zA-Z0-9_]*)\.)?([a-zA-Z_][a-zA-Z0-9_]*)\s*=\s*'([^']*)'", part)
        if not match:
            match = re.match(r"(?:([a-zA-Z_][a-zA-Z0-9_]*)\.)?([a-zA-Z_][a-zA-Z0-9_]*)\s*=\s*(\S+)", part)

        if match:
            alias = match.group(1)
            field = match.group(2)
            value = match.group(3)
            conditions[f"{alias}.{field}" if alias else field] = value
        else:
            # 无法解析，返回None
            return None

    return conditions if conditions else None


def merge_sql_statements(all_sql_statements):
    """
    合并相同修改但不同查询条件的SQL语句
    - 同一字段的不同值使用 IN
    - 不同字段的条件使用 OR
    - 完全相同的语句（包括WHERE）进行去重
    - 汇总表更新语句（多条明细共用一个汇总条件）放置在最后
    
    关键修复：对于包含子查询的SET子句，需要提取其结构特征而非完整内容作为分组键
    """
    if not all_sql_statements:
        return []

    # 按SET子句和表名分组
    forward_groups = defaultdict(list)
    backward_groups = defaultdict(list)

    for stmt in all_sql_statements:
        for forward_sql_item in stmt['forward_sqls']:
            # 使用原始SQL进行提取
            raw_sql = forward_sql_item['raw'] if isinstance(forward_sql_item, dict) else forward_sql_item

            # 提取表名和SET子句作为key
            table_name = extract_table_name(raw_sql)
            set_clause = extract_set_clause(raw_sql)
            where_clause = extract_where_clause(raw_sql)
            
            # 关键修复：对于汇总SQL，使用规范化后的SET子句作为分组键
            # 对于普通SQL，使用完整的SET子句作为分组键（包含具体值）
            is_summary = _is_summary_sql(set_clause)
            if is_summary:
                # 汇总SQL：使用结构特征作为分组键
                set_key = _normalize_set_clause_for_grouping(set_clause)
            else:
                # 普通SQL：使用完整的SET子句作为分组键（确保相同值的才能合并）
                set_key = set_clause

            key = f"{table_name}|{set_key}"
            forward_groups[key].append({
                'table_name': table_name,
                'set_clause': set_clause,
                'where_clause': where_clause
            })

        for backward_sql_item in stmt['backward_sqls']:
            # 使用原始SQL进行提取
            raw_sql = backward_sql_item['raw'] if isinstance(backward_sql_item, dict) else backward_sql_item

            table_name = extract_table_name(raw_sql)
            set_clause = extract_set_clause(raw_sql)
            where_clause = extract_where_clause(raw_sql)
            
            # 同样处理回退语句
            is_summary = _is_summary_sql(set_clause)
            if is_summary:
                set_key = _normalize_set_clause_for_grouping(set_clause)
            else:
                set_key = set_clause

            key = f"{table_name}|{set_key}"
            backward_groups[key].append({
                'table_name': table_name,
                'set_clause': set_clause,
                'where_clause': where_clause
            })

    merged_forward_normal = []  # 普通SQL（明细表更新）
    merged_forward_summary = []  # 汇总SQL（主表汇总更新）
    merged_backward_normal = []
    merged_backward_summary = []

    # 合并执行语句
    for key, items in forward_groups.items():
        if len(items) == 1:
            # 只有一个条件，不需要合并
            item = items[0]
            sql = f"UPDATE {item['table_name']} SET {item['set_clause']} WHERE {item['where_clause']}"
            formatted_sql = format_sql(sql)

            # 判断是否为汇总SQL：如果包含SELECT SUM等聚合函数，则为汇总SQL
            if _is_summary_sql(item['set_clause']):
                merged_forward_summary.append(formatted_sql)
            else:
                merged_forward_normal.append(formatted_sql)
        else:
            # 多个条件，需要智能合并
            # 先对WHERE子句去重（完全相同的WHERE只保留一个）
            unique_where_clauses = list(dict.fromkeys([item['where_clause'] for item in items]))

            if len(unique_where_clauses) == 1:
                # 所有WHERE条件都相同，说明是重复的汇总SQL，只生成一条
                item = items[0]
                sql = f"UPDATE {item['table_name']} SET {item['set_clause']} WHERE {unique_where_clauses[0]}"
                formatted_sql = format_sql(sql)

                # 汇总SQL放到最后
                if _is_summary_sql(item['set_clause']):
                    merged_forward_summary.append(formatted_sql)
                else:
                    merged_forward_normal.append(formatted_sql)
            else:
                # 关键修复：检查是否为汇总SQL
                is_summary = _is_summary_sql(items[0]['set_clause'])
                
                if is_summary:
                    # 汇总SQL：不能合并！因为每个合同的SET子句中的子查询使用了不同的合同号
                    # 必须为每个合同生成独立的UPDATE语句，但需要去重
                    seen_sqls = set()  # 用于去重
                    for item in items:
                        sql = f"UPDATE {item['table_name']} SET {item['set_clause']} WHERE {item['where_clause']}"
                        formatted_sql = format_sql(sql)
                        # 只有未出现过的SQL才添加
                        if formatted_sql not in seen_sqls:
                            seen_sqls.add(formatted_sql)
                            merged_forward_summary.append(formatted_sql)
                else:
                    # 普通SQL：可以合并WHERE条件
                    merged_where = merge_where_clauses(unique_where_clauses)
                    sql = f"UPDATE {items[0]['table_name']} SET {items[0]['set_clause']} WHERE {merged_where}"
                    formatted_sql = format_sql(sql)
                    merged_forward_normal.append(formatted_sql)

    # 合并回退语句
    for key, items in backward_groups.items():
        if len(items) == 1:
            item = items[0]
            sql = f"UPDATE {item['table_name']} SET {item['set_clause']} WHERE {item['where_clause']}"
            formatted_sql = format_sql(sql)

            # 判断是否为汇总SQL
            if _is_summary_sql(item['set_clause']):
                merged_backward_summary.append(formatted_sql)
            else:
                merged_backward_normal.append(formatted_sql)
        else:
            # 先对WHERE子句去重
            unique_where_clauses = list(dict.fromkeys([item['where_clause'] for item in items]))

            if len(unique_where_clauses) == 1:
                # 所有WHERE条件都相同，只生成一条
                item = items[0]
                sql = f"UPDATE {item['table_name']} SET {item['set_clause']} WHERE {unique_where_clauses[0]}"
                formatted_sql = format_sql(sql)

                if _is_summary_sql(item['set_clause']):
                    merged_backward_summary.append(formatted_sql)
                else:
                    merged_backward_normal.append(formatted_sql)
            else:
                # 关键修复：检查是否为汇总SQL
                is_summary = _is_summary_sql(items[0]['set_clause'])
                
                if is_summary:
                    # 汇总SQL：不能合并！因为每个合同的SET子句中的子查询使用了不同的合同号
                    # 必须为每个合同生成独立的UPDATE语句，但需要去重
                    seen_sqls = set()  # 用于去重
                    for item in items:
                        sql = f"UPDATE {item['table_name']} SET {item['set_clause']} WHERE {item['where_clause']}"
                        formatted_sql = format_sql(sql)
                        # 只有未出现过的SQL才添加
                        if formatted_sql not in seen_sqls:
                            seen_sqls.add(formatted_sql)
                            merged_backward_summary.append(formatted_sql)
                else:
                    # 普通SQL：可以合并WHERE条件
                    merged_where = merge_where_clauses(unique_where_clauses)
                    sql = f"UPDATE {items[0]['table_name']} SET {items[0]['set_clause']} WHERE {merged_where}"
                    formatted_sql = format_sql(sql)
                    merged_backward_normal.append(formatted_sql)

    # 合并结果：普通SQL在前，汇总SQL在后
    merged_forward = merged_forward_normal + merged_forward_summary
    merged_backward = merged_backward_normal + merged_backward_summary

    return {
        'forward_sqls': merged_forward,
        'backward_sqls': merged_backward
    }


def _is_summary_sql(set_clause):
    """
    判断SET子句是否包含汇总计算（聚合函数）
    如果包含 SELECT SUM、SELECT COUNT、SELECT AVG 等，则认为是汇总SQL
    """
    if not set_clause:
        return False

    set_upper = set_clause.upper()
    summary_keywords = ['SELECT SUM', 'SELECT COUNT', 'SELECT AVG', 'SELECT MAX', 'SELECT MIN']

    for keyword in summary_keywords:
        if keyword in set_upper:
            return True

    return False


def _normalize_set_clause_for_grouping(set_clause):
    """
    规范化SET子句用于分组，将具体的值替换为占位符，保留结构特征
    
    例如：
    - BPO_QTY = '10', BPO_PRICE = '100' -> BPO_QTY = ?, BPO_PRICE = ?
    - BPO_AMT = (SELECT SUM(BPO_AMT) FROM tphct02 WHERE BPO_ID = 'xxx' ...) 
      -> BPO_AMT = (SELECT SUM(BPO_AMT) FROM tphct02 WHERE BPO_ID = ? ...)
    
    这样可以确保相同结构的SET子句被分到同一组，即使具体值不同
    """
    import re
    
    if not set_clause:
        return ''
    
    # 匹配引号包裹的字符串值，替换为 ?
    normalized = re.sub(r"'[^']*'", '?', set_clause)
    
    # 匹配数字值（独立存在的数字），替换为 ?
    # 注意：不要替换字段名中的数字
    normalized = re.sub(r'\b(\d+)\b', '?', normalized)
    
    return normalized


def extract_set_clause(sql):
    """
    从 SQL 中提取SET子句
    关键修复：需要正确处理子查询中的WHERE关键字
    """
    sql_upper = sql.upper()
    set_start = sql_upper.find(' SET ')
    
    if set_start == -1:
        return ''
    
    # 从 SET 之后开始查找外层的 WHERE
    # 需要跳过子查询中的 WHERE（通过括号层级判断）
    where_start = _find_outer_where(sql, set_start + 5)
    
    if where_start != -1:
        return sql[set_start + 5:where_start].strip()
    else:
        return sql[set_start + 5:].strip()


def _find_outer_where(sql, start_pos):
    """
    从指定位置开始查找外层的 WHERE 关键字
    跳过括号内的 WHERE（子查询中的WHERE）
    
    返回 WHERE 的位置，如果没找到返回 -1
    """
    paren_depth = 0
    i = start_pos
    sql_upper = sql.upper()
    
    while i < len(sql_upper):
        char = sql_upper[i]
        
        if char == '(':
            paren_depth += 1
        elif char == ')':
            paren_depth -= 1
        elif paren_depth == 0 and sql_upper[i:i+7] == ' WHERE ':
            # 只有在括号外层（paren_depth == 0）的 WHERE 才是外层的
            return i
        
        i += 1
    
    return -1


def extract_where_clause(sql):
    """
    从 SQL 中提取WHERE子句
    关键修复：需要正确处理子查询中的WHERE关键字
    """
    sql_upper = sql.upper()
    
    # 先找到 SET 的位置
    set_start = sql_upper.find(' SET ')
    if set_start == -1:
        return ''
    
    # 从 SET 之后查找外层的 WHERE
    where_start = _find_outer_where(sql, set_start + 5)
    
    if where_start != -1:
        return sql[where_start + 7:].strip()
    return ''


def extract_table_name(sql):
    """从SQL中提取主表名（兼容 JOIN UPDATE）"""
    sql_upper = sql.upper()
    update_start = sql_upper.find('UPDATE ')
    set_start = sql_upper.find(' SET ')

    if update_start != -1 and set_start != -1:
        table_ref = sql[update_start + 7:set_start].strip()
        upper_ref = table_ref.upper()
        # 找到 JOIN 关键字位置，只取主表部分
        join_pos = upper_ref.find(' JOIN ')
        if join_pos == -1:
            join_pos = upper_ref.find(' INNER ')
        if join_pos == -1:
            join_pos = upper_ref.find(' LEFT ')
        if join_pos != -1:
            table_ref = table_ref[:join_pos].strip()
        # 去除别名
        parts = table_ref.split()
        return parts[0] if parts else table_ref
    return ''


def _get_query_value_data(binding_key, connected_tables, form_values):
    """从 form_values 中查找查询字段的值，优先使用 query_{bindingKey}_{tableName} 表级独立键。"""
    if connected_tables:
        for table_name in connected_tables:
            query_key = f"query_{binding_key}_{table_name}"
            if query_key in form_values:
                return form_values.get(query_key, {})
    query_key = f"query_{binding_key}"
    return form_values.get(query_key, {})


def _extract_table_names_from_expression(expression):
    """从 SQL 表达式中提取显式表名（忽略别名与 schema 前缀），用于变量解析。"""
    if not expression:
        return set()
    tables = set()
    # 匹配 FROM/JOIN/UPDATE/INTO/TABLE 后面的表名（含 schema 前缀）
    pattern = re.compile(
        r'(?:FROM|JOIN|UPDATE|INTO|TABLE)\s+([a-zA-Z_][a-zA-Z0-9_]*(?:\.[a-zA-Z_][a-zA-Z0-9_]*)?)',
        re.IGNORECASE
    )
    for match in pattern.finditer(expression):
        name = match.group(1)
        tables.add(name.split('.')[-1])
    return tables


def _get_actual_key(binding_key, connected_tables, form_values):
    """根据 connectedTable 返回 form_values 中实际使用的 key（仅用于更新字段）"""
    if connected_tables:
        for table in connected_tables:
            unique_key = f"{binding_key}_{table}"
            if unique_key in form_values:
                return unique_key
    return binding_key


def _get_value_data(binding_key, connected_tables, form_values):
    """从 form_values 中查找更新字段值，支持 bindingKey_tableName 格式"""
    actual_key = _get_actual_key(binding_key, connected_tables, form_values)
    return form_values.get(actual_key, {})


def _is_empty(value):
    return value is None or str(value).strip() == ''


def validate_form_data(config, form_values, query_values=None):
    """后端表单校验：统一在线编辑、单条提交、文件导入、表单合并的校验规则"""
    errors = []
    import re
    from decimal import Decimal, InvalidOperation

    query_items = config.get('queryItems', [])
    update_items = config.get('updateItems', [])

    # ========== 辅助函数 ==========
    def _check_required(value, valid_rule, label_prefix, label):
        """必填校验，返回错误字符串或 None"""
        if valid_rule == 'required' and _is_empty(value):
            return f"{label_prefix}{label}不能为空"
        return None

    def _validate_select_value(value, options, actual_key, field_key, label_prefix, label):
        """
        下拉框/单选框值范围校验，支持 label 自动转换为 value。
        field_key 为 'newValue' 或 'originValue'。
        返回错误字符串或 None。
        """
        if _is_empty(value):
            return None

        value_str = str(value).strip()
        label_to_value = {}
        valid_values = set()
        for opt in options:
            opt_value = str(opt.get('value', ''))
            opt_label = str(opt.get('label', ''))
            valid_values.add(opt_value)
            if opt_label:
                label_to_value[opt_label] = opt_value

        if value_str in label_to_value:
            converted = label_to_value[value_str]
            if actual_key in form_values:
                form_values[actual_key][field_key] = converted
            return None

        if value_str in valid_values:
            return None

        return f"{label_prefix}{label}的值'{value_str}'不在可选范围内"

    def _validate_number_value(value, label_prefix, label):
        if _is_empty(value):
            return None
        try:
            Decimal(str(value))
            return None
        except (InvalidOperation, ValueError, TypeError):
            return f"{label_prefix}{label}的值'{value}'不是有效的数值"

    def _validate_date_value(value, label_prefix, label):
        if _is_empty(value):
            return None
        if re.match(r'^\d{8}$', str(value)):
            return None
        return f"{label_prefix}{label}的日期格式不正确，应为yyyyMMdd格式（如：20260125）"

    # ========== 更新字段校验 ==========
    for item in update_items:
        binding_key = item.get('bindingKey')
        connected_tables = item.get('connectedTable', [])
        input_type = item.get('inputType', '')
        field_type = item.get('type', 'text')
        label = item.get('label', '')

        actual_key = _get_actual_key(binding_key, connected_tables, form_values)
        value_data = form_values.get(actual_key, {})

        # 跳过补充框（后面单独处理）和计算字段（由后端自动计算，不需要用户填写）
        if input_type == 'supplement' or input_type == 'calculated':
            continue

        new_value = value_data.get('newValue')
        origin_value = value_data.get('originValue')
        new_valid_rule = item.get('newValidRule', '')
        origin_valid_rule = item.get('originValidRule', '')

        if input_type == 'select' or input_type == 'radio':
            options = item.get('options', [])
            err = _validate_select_value(new_value, options, actual_key, 'newValue', '新', label)
            if err:
                errors.append(err)
            err = _validate_select_value(origin_value, options, actual_key, 'originValue', '原', label)
            if err:
                errors.append(err)
            # 转换后重新取值
            value_data = form_values.get(actual_key, {})
            new_value = value_data.get('newValue')
            origin_value = value_data.get('originValue')

        elif field_type == 'number' or input_type == 'number':
            err = _validate_number_value(new_value, '新', label)
            if err:
                errors.append(err)
            err = _validate_number_value(origin_value, '原', label)
            if err:
                errors.append(err)

        elif field_type == 'date' or input_type == 'date':
            err = _validate_date_value(new_value, '新', label)
            if err:
                errors.append(err)
            err = _validate_date_value(origin_value, '原', label)
            if err:
                errors.append(err)

        # 普通文本/其他类型只做必填校验
        err = _check_required(new_value, new_valid_rule, '新', label)
        if err:
            errors.append(err)
        err = _check_required(origin_value, origin_valid_rule, '原', label)
        if err:
            errors.append(err)

    # ========== 补充框校验 ==========
    supplement_items = [item for item in update_items if item.get('inputType') == 'supplement']

    # 检查整组补充框是否至少有一个有值
    has_any_supplement_value = False
    for item in supplement_items:
        binding_key = item.get('bindingKey')
        connected_tables = item.get('connectedTable', [])
        value_data = _get_value_data(binding_key, connected_tables, form_values)
        if not _is_empty(value_data.get('newValue')) or not _is_empty(value_data.get('originValue')):
            has_any_supplement_value = True
            break

    for item in supplement_items:
        binding_key = item.get('bindingKey')
        connected_tables = item.get('connectedTable', [])
        label = item.get('label', '')
        actual_key = _get_actual_key(binding_key, connected_tables, form_values)
        value_data = form_values.get(actual_key, {})

        if not value_data:
            continue

        new_value = value_data.get('newValue')
        origin_value = value_data.get('originValue')
        main_table = item.get('mainTable', '')
        main_field = item.get('mainField', '')
        sub_fields = item.get('subFields', [])

        # 校验新值/原值是否在数据库中存在（含辅助字段联合精确条件）
        def _validate_supplement_main_value(main_value, is_origin):
            if _is_empty(main_value) or not main_table or not main_field:
                return
            try:
                auxiliary_values = []
                for sf in get_auxiliary_sub_fields(sub_fields):
                    sub_binding_key = sf.get('bindingKey', '')
                    db_field = sf.get('dbField') or sub_binding_key
                    if not sub_binding_key or not db_field:
                        continue
                    aux_value_data = _get_value_data(sub_binding_key, connected_tables, form_values)
                    aux_value = str(aux_value_data.get('originValue' if is_origin else 'newValue', '')).strip() if isinstance(aux_value_data, dict) else ''
                    if aux_value:
                        auxiliary_values.append({'dbField': db_field, 'value': aux_value})

                where_conditions = [f"{main_field} = %s"]
                params = [str(main_value)]
                for av in auxiliary_values:
                    where_conditions.append(f"{av['dbField']} = %s")
                    params.append(av['value'])
                sql = f"SELECT COUNT(*) FROM {main_table} WHERE {' AND '.join(where_conditions)}"

                with connection.cursor() as cursor:
                    cursor.execute(sql, params)
                    count = cursor.fetchone()[0]
                    if count == 0:
                        prefix = '原' if is_origin else '新'
                        errors.append(f"{prefix}{label}的值'{main_value}'在数据库中不存在")
            except Exception as e:
                print(f"校验补充框{'原值' if is_origin else '新值'}失败: {e}")

        if not _is_empty(new_value) and main_table and main_field:
            from django.db import connection
            _validate_supplement_main_value(new_value, False)
        if not _is_empty(origin_value) and main_table and main_field:
            from django.db import connection
            _validate_supplement_main_value(origin_value, True)

        # 整组有值时跳过单个补充框的必填校验
        if has_any_supplement_value:
            continue

        new_valid_rule = item.get('newValidRule', '')
        origin_valid_rule = item.get('originValidRule', '')
        err = _check_required(new_value, new_valid_rule, '新', label)
        if err:
            errors.append(err)
        err = _check_required(origin_value, origin_valid_rule, '原', label)
        if err:
            errors.append(err)

    # ========== 查询字段校验 ==========
    for item in query_items:
        binding_key = item.get('bindingKey')
        connected_tables = item.get('connectedTable', [])
        label = item.get('label', '')
        valid_rule = item.get('ValidRule', '')
        field_type = item.get('type', 'text')

        value_data = _get_query_value_data(binding_key, connected_tables, form_values)

        if field_type in ('calculated', 'subquery'):
            # 计算字段/子查询字段：校验表达式是否已配置
            expressions = item.get('expressions', []) or []
            has_expression = False
            if isinstance(expressions, dict):
                has_expression = any((expr or '').strip() for expr in expressions.values())
            elif isinstance(expressions, list):
                has_expression = any((entry.get('expression') or '').strip() for entry in expressions)
            if not has_expression:
                errors.append(f"{label}（{'子查询' if field_type == 'subquery' else '计算字段'}）未配置表达式")
            # 若启用执行/回退不一致，需同时校验回退表达式
            if item.get('splitExpression'):
                backward_expressions = item.get('backwardExpressions', []) or []
                has_backward = False
                if isinstance(backward_expressions, dict):
                    has_backward = any((expr or '').strip() for expr in backward_expressions.values())
                elif isinstance(backward_expressions, list):
                    has_backward = any((entry.get('expression') or '').strip() for entry in backward_expressions)
                if not has_backward:
                    errors.append(f"{label}（{'子查询' if field_type == 'subquery' else '计算字段'}）已启用执行/回退不一致，但未配置回退表达式")
            continue

        if field_type == 'difference_condition':
            # 差异条件：分别校验新值和原值
            new_value = value_data.get('newValue')
            origin_value = value_data.get('originValue')
            err = _check_required(new_value, valid_rule, '新', label)
            if err:
                errors.append(err)
            err = _check_required(origin_value, valid_rule, '原', label)
            if err:
                errors.append(err)
            continue

        value = value_data.get('value')
        err = _check_required(value, valid_rule, '', label)
        if err:
            errors.append(err)

    # ========== 公共字段校验 ==========
    common_fields = ['filePrefix', 'onesLink', 'dynamicNo']
    for field_name in common_fields:
        value = None
        if query_values:
            value_data = query_values.get(field_name, {})
            value = value_data.get('value') if isinstance(value_data, dict) else value_data
        if not value:
            value_data = form_values.get(field_name, {})
            value = value_data.get('value')
        if _is_empty(value):
            errors.append(f"{field_name}不能为空")

    # ========== 至少填一个查询/更新条件 ==========
    query_field_has_value = False
    for item in query_items:
        binding_key = item.get('bindingKey')
        connected_tables = item.get('connectedTable', [])
        field_type = item.get('type', 'text')

        value_data = _get_query_value_data(binding_key, connected_tables, form_values)

        if field_type in ('calculated', 'subquery'):
            # 计算字段/子查询字段：只要有表达式就认为提供了查询条件
            expressions = item.get('expressions', []) or []
            if isinstance(expressions, dict):
                has_expr = any((expr or '').strip() for expr in expressions.values())
            else:
                has_expr = any((entry.get('expression') or '').strip() for entry in expressions)
            if has_expr:
                query_field_has_value = True
                break
        elif field_type == 'difference_condition':
            if not _is_empty(value_data.get('newValue')) or not _is_empty(value_data.get('originValue')):
                query_field_has_value = True
                break
        else:
            if not _is_empty(value_data.get('value')):
                query_field_has_value = True
                break

    if query_items and not query_field_has_value:
        errors.append("查询字段至少需要填写一个条件")

    has_non_empty_update = False
    for item in update_items:
        binding_key = item.get('bindingKey')
        connected_tables = item.get('connectedTable', [])
        value_data = _get_value_data(binding_key, connected_tables, form_values)
        if not _is_empty(value_data.get('newValue')) or not _is_empty(value_data.get('originValue')):
            has_non_empty_update = True
            break

    if update_items and not has_non_empty_update:
        errors.append("更新字段至少需要填写一个新值或原值")

    if errors:
        return {
            'success': False,
            'message': f'共有{len(errors)}个校验错误',
            'errors': errors
        }

    return {
        'success': True,
        'message': '校验通过',
        'errors': []
    }


def _get_query_item_value_dicts(item, headers, ws=None, row_idx=None):
    """从 Excel 或字典行数据中提取查询字段值。

    普通类型返回 {None: value_dict}。
    差异条件返回 {None: {newValue, originValue}}。
    计算字段返回 {None: {value: '', fieldType: 'calculated', expressions}}。
    若必要列缺失，返回中可能包含 '_missing': label。
    """
    label = item.get('label', '')
    binding_key = item.get('bindingKey', '')
    field_type = item.get('type', 'text')
    default_value = item.get('defaultValue', '') or ''
    valid_rule = item.get('ValidRule', '')

    def _read_value(col_key):
        if col_key not in headers:
            return None
        if ws is not None and row_idx is not None:
            col = headers[col_key]
            value = ws.cell(row=row_idx, column=col).value
        else:
            value = headers[col_key]
        return str(value).strip() if value is not None else ''

    result = {}

    if field_type in ('calculated', 'subquery'):
        result[None] = {
            'label': label,
            'value': '',
            'inputType': 'query',
            'fieldType': field_type,
            'ValidRule': valid_rule,
            'expressions': item.get('expressions', []) or [],
        }
        return result

    if field_type == 'difference_condition':
        new_label = f'新{label}'
        origin_label = f'原{label}'

        new_value = ''
        origin_value = ''

        new_read = _read_value(new_label)
        if new_read is not None:
            new_value = new_read
        elif valid_rule == 'required':
            result['_missing'] = new_label

        origin_read = _read_value(origin_label)
        if origin_read is not None:
            origin_value = origin_read
        elif valid_rule == 'required':
            result['_missing'] = origin_label

        result[None] = {
            'label': label,
            'newValue': new_value,
            'originValue': origin_value,
            'inputType': 'query',
            'fieldType': 'difference_condition',
            'ValidRule': valid_rule,
        }
        return result

    value_str = ''
    read = _read_value(label)
    if read is not None:
        value_str = read
    elif valid_rule == 'required':
        result['_missing'] = label

    if not value_str and default_value:
        value_str = str(default_value).strip()

    result[None] = {
        'label': label,
        'value': value_str,
        'inputType': 'query',
        'fieldType': field_type,
        'ValidRule': valid_rule,
        'defaultValue': default_value,
    }

    return result


def build_form_values_from_excel(ws, row_idx, headers, query_items, update_items):
    """从 Excel 行构建表单值

    关键：对于有 connectedTable 的字段，使用 bindingKey_tableName 作为 key，
    确保同名字段关联不同表时能正确查找。
    """
    form_values = {}
    missing_columns = []

    for item in query_items:
        binding_key = item.get('bindingKey', '')

        value_dicts = _get_query_item_value_dicts(item, headers, ws=ws, row_idx=row_idx)
        if '_missing' in value_dicts:
            missing_columns.append(value_dicts.pop('_missing'))

        # 查询字段使用独立的 query_{bindingKey} 键，避免与修改字段冲突
        for value_dict in value_dicts.values():
            form_values[f"query_{binding_key}"] = value_dict

    for item in update_items:
        label = item.get('label', '')
        binding_key = item.get('bindingKey', '')
        input_type = item.get('inputType', '')
        connected_tables = item.get('connectedTable', [])

        if input_type == 'supplement':
            sub_fields = item.get('subFields', [])
            new_label = f'新{label}'
            origin_label = f'原{label}'

            new_value = ''
            origin_value = ''

            new_valid_rule = item.get('newValidRule', '')
            origin_valid_rule = item.get('originValidRule', '')

            if new_label in headers:
                col = headers[new_label]
                new_value = ws.cell(row=row_idx, column=col).value
            elif new_valid_rule == 'required':
                missing_columns.append(new_label)

            if origin_label in headers:
                col = headers[origin_label]
                origin_value = ws.cell(row=row_idx, column=col).value
            elif origin_valid_rule == 'required':
                missing_columns.append(origin_label)

            nvs = str(new_value).strip() if new_value is not None else ''
            ovs = str(origin_value).strip() if origin_value is not None else ''
            if not nvs and item.get('newDefaultValue'):
                nvs = str(item.get('newDefaultValue', '')).strip()
            if not ovs and item.get('originDefaultValue'):
                ovs = str(item.get('originDefaultValue', '')).strip()

            supplement_value = {
                'label': label,
                'newValue': nvs,
                'originValue': ovs,
                'inputType': 'supplement',
                'fieldType': 'supplement',
                'newValidRule': item.get('newValidRule', ''),
                'originValidRule': item.get('originValidRule', '')
            }

            # 使用 bindingKey_tableName 格式存储补充框主字段
            if connected_tables:
                supplement_unique_keys = [f"{binding_key}_{table}" for table in connected_tables]
                for unique_key in supplement_unique_keys:
                    form_values[unique_key] = supplement_value
            else:
                supplement_unique_keys = [binding_key]
                form_values[binding_key] = supplement_value

            # ==================== 自动填充补充框子字段 ====================
            # 如果新值或原值不为空，从数据库查询对应的子字段值
            main_table = item.get('mainTable', '')
            main_field = item.get('mainField', '')

            if main_table and main_field and sub_fields:
                from django.db import connection

                # 构建要查询的字段列表
                select_fields = [main_field]
                for sub_field in sub_fields:
                    if isinstance(sub_field, dict):
                        field_name = sub_field.get('dbField') or sub_field.get('bindingKey')
                        if field_name:
                            select_fields.append(field_name)
                    elif isinstance(sub_field, str):
                        select_fields.append(sub_field)
                fields_str = ', '.join(select_fields)

                def _query_supplement_map(main_values, auxiliary_values):
                    """按主字段值和辅助条件查询，返回 {main_val: row_dict}"""
                    if not main_values:
                        return {}
                    main_values = list(set(main_values))
                    values_str = ', '.join(["'" + str(v).replace("'", "''") + "'" for v in main_values])
                    where_conditions = [f"{main_field} IN ({values_str})"]
                    where_conditions.extend(build_auxiliary_where_sql(auxiliary_values))
                    sql = f"SELECT {fields_str} FROM {main_table} WHERE {' AND '.join(where_conditions)}"
                    data_map = {}
                    try:
                        with connection.cursor() as cursor:
                            cursor.execute(sql)
                            rows = cursor.fetchall()
                            for row in rows:
                                row_dict = {}
                                for idx, field in enumerate(select_fields):
                                    row_dict[field] = row[idx]
                                main_val = row_dict.get(main_field, '')
                                data_map[main_val] = row_dict
                    except Exception as e:
                        print(f"查询补充框子字段失败: {e}")
                        import traceback
                        traceback.print_exc()
                    return data_map

                # 分别读取新值/原值的辅助条件，避免两者不一致时互相干扰
                new_aux = get_auxiliary_values_from_excel(ws, row_idx, headers, item, is_origin=False)
                origin_aux = get_auxiliary_values_from_excel(ws, row_idx, headers, item, is_origin=True)

                new_values = []
                if new_value and str(new_value).strip():
                    new_values.append(str(new_value).strip())
                origin_values = []
                if origin_value and str(origin_value).strip():
                    origin_values.append(str(origin_value).strip())

                new_data_map = _query_supplement_map(new_values, new_aux)
                origin_data_map = _query_supplement_map(origin_values, origin_aux)

                # 填充子字段（使用 connected_tables 作为 key，与 generate_update_sql 查找一致）
                parent_value_data = form_values.get(supplement_unique_keys[0], {})

                # 填充新值的子字段
                new_val_str = parent_value_data.get('newValue', '')
                if new_val_str and str(new_val_str).strip():
                    new_val_str = str(new_val_str).strip()
                    if new_val_str in new_data_map:
                        row_data = new_data_map[new_val_str]
                        for sub_field in sub_fields:
                            if isinstance(sub_field, dict):
                                sub_binding_key = sub_field.get('bindingKey', '')
                                db_field = sub_field.get('dbField') or sub_binding_key
                                sub_value = row_data.get(db_field, '')

                                # 按 connected_tables 存储，与 generate_update_sql 查找一致
                                if connected_tables:
                                    for table in connected_tables:
                                        form_values[f"{sub_binding_key}_{table}"] = {
                                            'newValue': str(sub_value).strip() if sub_value is not None else '',
                                            'originValue': '',
                                            'inputType': 'supplement-sub',
                                            'fieldType': 'supplement-sub',
                                            'parentKey': supplement_unique_keys[0],
                                            'label': sub_field.get('label', '')
                                        }
                                else:
                                    form_values[sub_binding_key] = {
                                        'newValue': str(sub_value).strip() if sub_value is not None else '',
                                        'originValue': '',
                                        'inputType': 'supplement-sub',
                                        'fieldType': 'supplement-sub',
                                        'parentKey': supplement_unique_keys[0],
                                        'label': sub_field.get('label', '')
                                    }

                # 填充原值的子字段
                origin_val_str = parent_value_data.get('originValue', '')
                if origin_val_str and str(origin_val_str).strip():
                    origin_val_str = str(origin_val_str).strip()
                    if origin_val_str in origin_data_map:
                        row_data = origin_data_map[origin_val_str]
                        for sub_field in sub_fields:
                            if isinstance(sub_field, dict):
                                sub_binding_key = sub_field.get('bindingKey', '')
                                db_field = sub_field.get('dbField') or sub_binding_key
                                sub_value = row_data.get(db_field, '')

                                if connected_tables:
                                    for table in connected_tables:
                                        sub_unique_key = f"{sub_binding_key}_{table}"
                                        if sub_unique_key in form_values:
                                            form_values[sub_unique_key]['originValue'] = str(
                                                sub_value).strip() if sub_value is not None else ''
                                        else:
                                            form_values[sub_unique_key] = {
                                                'newValue': '',
                                                'originValue': str(sub_value).strip() if sub_value is not None else '',
                                                'inputType': 'supplement-sub',
                                                'fieldType': 'supplement-sub',
                                                'parentKey': supplement_unique_keys[0],
                                                'label': sub_field.get('label', '')
                                            }
                                else:
                                    if sub_binding_key in form_values:
                                        form_values[sub_binding_key]['originValue'] = str(
                                            sub_value).strip() if sub_value is not None else ''
                                    else:
                                        form_values[sub_binding_key] = {
                                            'newValue': '',
                                            'originValue': str(sub_value).strip() if sub_value is not None else '',
                                            'inputType': 'supplement-sub',
                                            'fieldType': 'supplement-sub',
                                            'parentKey': supplement_unique_keys[0],
                                            'label': sub_field.get('label', '')
                                        }
            else:
                # 如果没有配置主表主字段，手动处理子字段
                for sub_field in sub_fields:
                    sub_label = sub_field.get('label', '')
                    sub_binding_key = sub_field.get('bindingKey', '')

                    new_sub_label = f'新{sub_label}'
                    origin_sub_label = f'原{sub_label}'

                    new_sub_value = ''
                    origin_sub_value = ''

                    if new_sub_label in headers:
                        col = headers[new_sub_label]
                        new_sub_value = ws.cell(row=row_idx, column=col).value

                    if origin_sub_label in headers:
                        col = headers[origin_sub_label]
                        origin_sub_value = ws.cell(row=row_idx, column=col).value

                    # 子字段也使用 bindingKey_tableName 格式
                    if connected_tables:
                        for table in connected_tables:
                            form_values[f"{sub_binding_key}_{table}"] = {
                                'newValue': str(new_sub_value).strip() if new_sub_value is not None else '',
                                'originValue': str(origin_sub_value).strip() if origin_sub_value is not None else '',
                                'inputType': 'supplement-sub',
                                'fieldType': 'supplement-sub',
                                'parentKey': f"{binding_key}_{table}",
                                'label': sub_label
                            }
                    else:
                        form_values[sub_binding_key] = {
                            'newValue': str(new_sub_value).strip() if new_sub_value is not None else '',
                            'originValue': str(origin_sub_value).strip() if origin_sub_value is not None else '',
                            'inputType': 'supplement-sub',
                            'fieldType': 'supplement-sub',
                            'parentKey': binding_key,
                            'label': sub_label
                        }
        else:
            # 处理普通字段和计算字段
            new_label = f'新{label}'
            origin_label = f'原{label}'

            new_value = ''
            origin_value = ''

            # 如果是计算字段，不需要从Excel读取，直接初始化为空
            if input_type == 'calculated':
                # 计算字段的值会在generate_update_sql中通过表达式计算
                calc_value = {
                    'label': label,
                    'newValue': '',
                    'originValue': '',
                    'inputType': input_type,
                    'fieldType': item.get('type', 'text'),
                    'newValidRule': item.get('newValidRule', ''),
                    'originValidRule': item.get('originValidRule', ''),
                    'expressions': item.get('expressions', []) or []  # 保留表达式配置
                }
                if connected_tables:
                    for table in connected_tables:
                        form_values[f"{binding_key}_{table}"] = calc_value
                else:
                    form_values[binding_key] = calc_value
            else:
                # 普通字段需要从Excel读取
                new_valid_rule = item.get('newValidRule', '')
                origin_valid_rule = item.get('originValidRule', '')

                if new_label in headers:
                    col = headers[new_label]
                    new_value = ws.cell(row=row_idx, column=col).value
                elif new_valid_rule == 'required':
                    missing_columns.append(new_label)

                if origin_label in headers:
                    col = headers[origin_label]
                    origin_value = ws.cell(row=row_idx, column=col).value
                elif origin_valid_rule == 'required':
                    missing_columns.append(origin_label)

                # 如果值为空且有默认值，使用默认值
                new_value_str = str(new_value).strip() if new_value is not None else ''
                origin_value_str = str(origin_value).strip() if origin_value is not None else ''

                if not new_value_str and item.get('newDefaultValue'):
                    new_value_str = str(item.get('newDefaultValue', '')).strip()

                if not origin_value_str and item.get('originDefaultValue'):
                    origin_value_str = str(item.get('originDefaultValue', '')).strip()

                regular_value = {
                    'label': label,
                    'newValue': new_value_str,
                    'originValue': origin_value_str,
                    'inputType': input_type,
                    'fieldType': item.get('type', 'text'),
                    'newValidRule': item.get('newValidRule', ''),
                    'originValidRule': item.get('originValidRule', '')
                }
                # 使用 bindingKey_tableName 格式存储
                if connected_tables:
                    for table in connected_tables:
                        form_values[f"{binding_key}_{table}"] = regular_value
                else:
                    form_values[binding_key] = regular_value

    return form_values, missing_columns





def build_form_values_from_excel_batch(ws, row_idx, headers, query_items, update_items):
    """批量导入专用：轻量版，跳过数据库查询
    
    补充框子字段填充由后续的缓存填充阶段完成。
    """
    form_values = {}
    missing_columns = []

    for item in query_items:
        binding_key = item.get('bindingKey', '')

        value_dicts = _get_query_item_value_dicts(item, headers, ws=ws, row_idx=row_idx)
        if '_missing' in value_dicts:
            missing_columns.append(value_dicts.pop('_missing'))

        # 查询字段使用独立的 query_{bindingKey} 键，避免与修改字段冲突
        for value_dict in value_dicts.values():
            form_values[f"query_{binding_key}"] = value_dict

    for item in update_items:
        label = item.get('label', '')
        binding_key = item.get('bindingKey', '')
        input_type = item.get('inputType', '')
        connected_tables = item.get('connectedTable', [])

        if input_type == 'supplement':
            new_label = f'新{label}'
            origin_label = f'原{label}'
            new_value = ''
            origin_value = ''
            new_valid_rule = item.get('newValidRule', '')
            origin_valid_rule = item.get('originValidRule', '')

            if new_label in headers:
                col = headers[new_label]
                new_value = ws.cell(row=row_idx, column=col).value
            elif new_valid_rule == 'required':
                missing_columns.append(new_label)

            if origin_label in headers:
                col = headers[origin_label]
                origin_value = ws.cell(row=row_idx, column=col).value

            nvs = str(new_value).strip() if new_value is not None else ''
            ovs = str(origin_value).strip() if origin_value is not None else ''
            if not nvs and item.get('newDefaultValue'):
                nvs = str(item.get('newDefaultValue', '')).strip()
            if not ovs and item.get('originDefaultValue'):
                ovs = str(item.get('originDefaultValue', '')).strip()

            supplement_value = {'label': label,
                'newValue': nvs,
                'originValue': ovs,
                'inputType': 'supplement', 'fieldType': 'supplement',
                'newValidRule': item.get('newValidRule', ''),
                'originValidRule': item.get('originValidRule', '')}

            if connected_tables:
                for table in connected_tables:
                    form_values[f"{binding_key}_{table}"] = supplement_value
            else:
                form_values[binding_key] = supplement_value

            # 跳过数据库查询，没有主表主字段时手动处理子字段
            main_table = item.get('mainTable', '')
            main_field = item.get('mainField', '')
            sub_fields = item.get('subFields', [])
            if not (main_table and main_field and sub_fields):
                for sf in sub_fields:
                    sub_label = sf.get('label', '')
                    sub_bk = sf.get('bindingKey', '')
                    ns = ws.cell(row=row_idx, column=headers.get(f'新{sub_label}', 0)).value if f'新{sub_label}' in headers else ''
                    os = ws.cell(row=row_idx, column=headers.get(f'原{sub_label}', 0)).value if f'原{sub_label}' in headers else ''
                    if connected_tables:
                        for table in connected_tables:
                            form_values[f"{sub_bk}_{table}"] = {
                                'newValue': str(ns).strip() if ns is not None else '',
                                'originValue': str(os).strip() if os is not None else '',
                                'inputType': 'supplement-sub', 'fieldType': 'supplement-sub',
                                'parentKey': f"{binding_key}_{table}", 'label': sub_label}
                    else:
                        form_values[sub_bk] = {
                            'newValue': str(ns).strip() if ns is not None else '',
                            'originValue': str(os).strip() if os is not None else '',
                            'inputType': 'supplement-sub', 'fieldType': 'supplement-sub',
                            'parentKey': binding_key, 'label': sub_label}
        else:
            # 普通字段/计算字段
            new_label = f'新{label}'
            origin_label = f'原{label}'

            if input_type == 'calculated':
                calc_value = {'label': label, 'newValue': '', 'originValue': '',
                    'inputType': input_type, 'fieldType': item.get('type', 'text'),
                    'newValidRule': item.get('newValidRule', ''),
                    'originValidRule': item.get('originValidRule', ''),
                    'expressions': item.get('expressions', []) or []}
                if connected_tables:
                    for table in connected_tables:
                        form_values[f"{binding_key}_{table}"] = calc_value
                else:
                    form_values[binding_key] = calc_value
            else:
                new_value = ''
                origin_value = ''
                new_valid_rule = item.get('newValidRule', '')
                origin_valid_rule = item.get('originValidRule', '')
                if new_label in headers:
                    col = headers[new_label]
                    new_value = ws.cell(row=row_idx, column=col).value
                elif new_valid_rule == 'required':
                    missing_columns.append(new_label)
                if origin_label in headers:
                    col = headers[origin_label]
                    origin_value = ws.cell(row=row_idx, column=col).value
                elif origin_valid_rule == 'required':
                    missing_columns.append(origin_label)

                nvs = str(new_value).strip() if new_value is not None else ''
                ovs = str(origin_value).strip() if origin_value is not None else ''
                if not nvs and item.get('newDefaultValue'): nvs = str(item.get('newDefaultValue', '')).strip()
                if not ovs and item.get('originDefaultValue'): ovs = str(item.get('originDefaultValue', '')).strip()

                rv = {'label': label, 'newValue': nvs, 'originValue': ovs,
                    'inputType': input_type, 'fieldType': item.get('type', 'text'),
                    'newValidRule': item.get('newValidRule', ''),
                    'originValidRule': item.get('originValidRule', '')}
                if connected_tables:
                    for table in connected_tables:
                        form_values[f"{binding_key}_{table}"] = rv
                else:
                    form_values[binding_key] = rv

    return form_values, missing_columns


def build_form_values_from_rows(row_data, query_items, update_items):
    """从在线编辑器行数据构建表单值
    
    row_data: dict, key为列名(如'新单价'), value为单元格值
    与build_form_values_from_excel_batch逻辑一致，但数据源为字典而非Excel
    """
    form_values = {}
    missing_columns = []
    headers = row_data  # 这里row_data本身就是key->value的映射

    for item in query_items:
        binding_key = item.get('bindingKey', '')

        value_dicts = _get_query_item_value_dicts(item, headers)
        if '_missing' in value_dicts:
            missing_columns.append(value_dicts.pop('_missing'))

        # 查询字段使用独立的 query_{bindingKey} 键，避免与修改字段冲突
        for value_dict in value_dicts.values():
            form_values[f"query_{binding_key}"] = value_dict

    for item in update_items:
        label = item.get('label', '')
        binding_key = item.get('bindingKey', '')
        input_type = item.get('inputType', '')
        connected_tables = item.get('connectedTable', [])

        if input_type == 'supplement':
            new_label = f'新{label}'
            origin_label = f'原{label}'
            new_value = headers.get(new_label, '')
            origin_value = headers.get(origin_label, '')

            nvs = str(new_value).strip() if new_value is not None else ''
            ovs = str(origin_value).strip() if origin_value is not None else ''
            if not nvs and item.get('newDefaultValue'):
                nvs = str(item.get('newDefaultValue', '')).strip()
            if not ovs and item.get('originDefaultValue'):
                ovs = str(item.get('originDefaultValue', '')).strip()

            supplement_value = {'label': label,
                'newValue': nvs,
                'originValue': ovs,
                'inputType': 'supplement', 'fieldType': 'supplement',
                'newValidRule': item.get('newValidRule', ''),
                'originValidRule': item.get('originValidRule', '')}

            if connected_tables:
                for table in connected_tables:
                    form_values[f"{binding_key}_{table}"] = supplement_value
            else:
                form_values[binding_key] = supplement_value

            # 跳过数据库查询，没有主表主字段时手动处理子字段
            main_table = item.get('mainTable', '')
            main_field = item.get('mainField', '')
            sub_fields = item.get('subFields', [])
            if not (main_table and main_field and sub_fields):
                for sf in sub_fields:
                    sub_label = sf.get('label', '')
                    sub_bk = sf.get('bindingKey', '')
                    ns = headers.get(f'新{sub_label}', '')
                    os = headers.get(f'原{sub_label}', '')
                    if connected_tables:
                        for table in connected_tables:
                            form_values[f"{sub_bk}_{table}"] = {
                                'newValue': str(ns).strip() if ns is not None else '',
                                'originValue': str(os).strip() if os is not None else '',
                                'inputType': 'supplement-sub', 'fieldType': 'supplement-sub',
                                'parentKey': f"{binding_key}_{table}", 'label': sub_label}
                    else:
                        form_values[sub_bk] = {
                            'newValue': str(ns).strip() if ns is not None else '',
                            'originValue': str(os).strip() if os is not None else '',
                            'inputType': 'supplement-sub', 'fieldType': 'supplement-sub',
                            'parentKey': binding_key, 'label': sub_label}
        else:
            # 普通字段/计算字段
            new_label = f'新{label}'
            origin_label = f'原{label}'

            if input_type == 'calculated':
                calc_value = {'label': label, 'newValue': '', 'originValue': '',
                    'inputType': input_type, 'fieldType': item.get('type', 'text'),
                    'newValidRule': item.get('newValidRule', ''),
                    'originValidRule': item.get('originValidRule', ''),
                    'expressions': item.get('expressions', []) or []}
                if connected_tables:
                    for table in connected_tables:
                        form_values[f"{binding_key}_{table}"] = calc_value
                else:
                    form_values[binding_key] = calc_value
            else:
                new_value = headers.get(new_label, '')
                origin_value = headers.get(origin_label, '')

                nvs = str(new_value).strip() if new_value is not None else ''
                ovs = str(origin_value).strip() if origin_value is not None else ''
                if not nvs and item.get('newDefaultValue'): nvs = str(item.get('newDefaultValue', '')).strip()
                if not ovs and item.get('originDefaultValue'): ovs = str(item.get('originDefaultValue', '')).strip()

                rv = {'label': label, 'newValue': nvs, 'originValue': ovs,
                    'inputType': input_type, 'fieldType': item.get('type', 'text'),
                    'newValidRule': item.get('newValidRule', ''),
                    'originValidRule': item.get('originValidRule', '')}
                if connected_tables:
                    for table in connected_tables:
                        form_values[f"{binding_key}_{table}"] = rv
                else:
                    form_values[binding_key] = rv

    return form_values, missing_columns


# ==================== 补充框辅助字段工具函数 ====================

def get_auxiliary_sub_fields(sub_fields):
    """从 subFields 中筛选出辅助字段（仅明确 type == 'auxiliary'）

    未指定 type 的子字段默认为普通字段，避免旧数据被误判为辅助查询条件。
    """
    aux_fields = []
    for sf in sub_fields:
        if isinstance(sf, dict) and sf.get('type') == 'auxiliary':
            aux_fields.append(sf)
    return aux_fields


def get_auxiliary_values_from_excel(ws, row_idx, headers, item, is_origin):
    """从 Excel 行中读取辅助字段值

    ws: openpyxl worksheet
    row_idx: 行号（从1开始）
    headers: dict, label -> column number
    item: supplement 配置项
    is_origin: True 读取原值，False 读取新值
    返回: [{'dbField': ..., 'value': ...}, ...]
    """
    aux_values = []
    prefix = '原' if is_origin else '新'
    sub_fields = item.get('subFields', [])
    for sf in get_auxiliary_sub_fields(sub_fields):
        sub_label = sf.get('label', '')
        db_field = sf.get('dbField') or sf.get('bindingKey', '')
        col_name = f'{prefix}{sub_label}'
        if not sub_label or not db_field or col_name not in headers:
            continue
        col = headers[col_name]
        value = ws.cell(row=row_idx, column=col).value
        value_str = str(value).strip() if value is not None else ''
        if value_str:
            aux_values.append({'dbField': db_field, 'value': value_str})
    return aux_values


def get_auxiliary_values_from_row_data(row_data, item, is_origin):
    """从行数据字典中读取辅助字段值

    row_data: dict, key 为列名（如 '新辅助字段'）
    item: supplement 配置项
    is_origin: True 读取原值，False 读取新值
    返回: [{'dbField': ..., 'value': ...}, ...]
    """
    aux_values = []
    prefix = '原' if is_origin else '新'
    sub_fields = item.get('subFields', [])
    for sf in get_auxiliary_sub_fields(sub_fields):
        sub_label = sf.get('label', '')
        db_field = sf.get('dbField') or sf.get('bindingKey', '')
        col_name = f'{prefix}{sub_label}'
        if not sub_label or not db_field or col_name not in row_data:
            continue
        value = row_data[col_name]
        value_str = str(value).strip() if value is not None else ''
        if value_str:
            aux_values.append({'dbField': db_field, 'value': value_str})
    return aux_values


def build_auxiliary_where_sql(auxiliary_values):
    """将辅助字段条件转换为 SQL WHERE 片段列表"""
    conditions = []
    if not isinstance(auxiliary_values, list):
        return conditions
    for av in auxiliary_values:
        if isinstance(av, dict):
            db_field = av.get('dbField', '').strip()
            value = av.get('value', '').strip()
            if db_field and value:
                conditions.append(db_field + " = '" + value.replace("'", "''") + "'")
    return conditions


def build_auxiliary_group_key(auxiliary_values):
    """根据辅助字段条件生成稳定的分组 key，用于批量缓存查询分组"""
    parts = []
    if isinstance(auxiliary_values, list):
        for av in sorted(auxiliary_values, key=lambda x: x.get('dbField', '') if isinstance(x, dict) else ''):
            if isinstance(av, dict):
                db_field = av.get('dbField', '').strip()
                value = av.get('value', '').strip()
                if db_field and value:
                    parts.append(f"{db_field}:{value}")
    return '|'.join(parts)


def build_auxiliary_cache_key(query_key, is_origin, auxiliary_values, main_value):
    """构建包含辅助字段条件的缓存 key"""
    aux_hash = build_auxiliary_group_key(auxiliary_values)
    origin_flag = 'origin' if is_origin else 'new'
    return f"{query_key}_{origin_flag}_{aux_hash}_{str(main_value).strip()}"


# ==================== 视图函数 ====================

@csrf_exempt
def dynamic_submit(request):
    """处理动态表单提交"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            config = data.get('config', {})
            form_values = data.get('formValues', {})

            validation_result = validate_form_data(config, form_values)

            if not validation_result['success']:
                return JsonResponse({
                    'success': False,
                    'message': validation_result['message'],
                    'errors': validation_result.get('errors', [])
                }, status=400)

            sql_result = generate_update_sql(config, form_values)

            # 检查是否有缺失字段错误
            if sql_result.get('missing_field_errors'):
                return JsonResponse({
                    'success': False,
                    'message': 'SQL 生成失败，请检查必填字段',
                    'errors': sql_result['missing_field_errors']
                }, status=400)

            dynamic_no = form_values.get('dynamicNo', {}).get('value', '')
            file_prefix = form_values.get('filePrefix', {}).get('value', '')

            # 使用路径配置获取保存路径
            save_dir = get_save_path_from_config()
            print(f"[DEBUG] SQL文件保存路径: {save_dir}")
            print(f"[DEBUG] 文件名: {dynamic_no}_{file_prefix}.sql")

            os.makedirs(save_dir, exist_ok=True)

            sql_content = []

            if sql_result['forward_sqls']:
                sql_content.append("1.执行语句")
                for i, sql_item in enumerate(sql_result['forward_sqls'], 1):
                    # sql_item可能是字典{'raw': ..., 'formatted': ...}或字符串
                    sql = sql_item['formatted'] if isinstance(sql_item, dict) else sql_item
                    # 确保SQL末尾有分号
                    if not sql.rstrip().endswith(';'):
                        sql = sql.rstrip() + ';'
                    sql_content.append(sql)
                    sql_content.append("")

            if sql_result['backward_sqls']:
                sql_content.append("2.回退语句")
                for i, sql_item in enumerate(sql_result['backward_sqls'], 1):
                    # sql_item可能是字典{'raw': ..., 'formatted': ...}或字符串
                    sql = sql_item['formatted'] if isinstance(sql_item, dict) else sql_item
                    # 确保SQL末尾有分号
                    if not sql.rstrip().endswith(';'):
                        sql = sql.rstrip() + ';'
                    sql_content.append(sql)
                    sql_content.append("")

            # 添加数据库信息
            database_ip_ids = config.get('databaseIpIds', [])
            if database_ip_ids:
                db_configs = DatabaseIPConfig.objects.filter(id__in=database_ip_ids).order_by('id')
                if db_configs:
                    sql_content.append("3.数据库")
                    for db_config in db_configs:
                        sql_content.append(f"ip：{db_config.ip_address}")
                        sql_content.append(f"库名：{db_config.database_name}")
                        sql_content.append("")

            # 文件名格式：编号_文件名.sql
            filename = f"{dynamic_no}_{file_prefix}.sql"
            filepath = os.path.join(save_dir, filename)

            with open(filepath, 'w', encoding='utf-8') as f:
                f.write('\n'.join(sql_content))

            # 记录表单调用日志
            try:
                form_id = config.get('formId')
                if form_id:
                    form_config = FormConfig.objects.get(id=form_id)
                    record_form_usage(form_config, source='dynamic')
            except Exception:
                # 调用统计不应影响主流程
                pass

            # 返回时也需要提取格式化后的SQL
            forward_formatted = [item['formatted'] if isinstance(item, dict) else item
                                 for item in sql_result['forward_sqls']]
            backward_formatted = [item['formatted'] if isinstance(item, dict) else item
                                  for item in sql_result['backward_sqls']]

            return JsonResponse({
                'success': True,
                'message': 'SQL 文件生成成功',
                'filePath': filepath,
                'forward_sqls': forward_formatted,
                'backward_sqls': backward_formatted
            })

        except json.JSONDecodeError as e:
            return JsonResponse({'success': False, 'message': f'JSON 解析失败：{str(e)}'}, status=400)
        except Exception as e:
            print(f"处理异常：{e}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'message': f'服务器错误：{str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': '仅支持 POST 请求'}, status=405)


@csrf_exempt
def download_template(request):
    """下载导入模板"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            config = data.get('config', {})

            form_name = config.get('formName', '模板')
            query_items = config.get('queryItems', [])
            update_items = config.get('updateItems', [])

            wb = Workbook()
            ws = wb.active
            ws.title = '导入模板'

            # 设置表头样式
            header_font = Font(bold=True, color='FFFFFF')
            header_alignment = Alignment(horizontal='center', vertical='center')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            default_value_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

            headers = []

            # 查询字段
            for item in query_items:
                label = item.get('label', '')
                field_type = item.get('type', 'text')

                if field_type in ('calculated', 'subquery'):
                    # 计算字段/子查询字段不显示在模板中
                    continue

                if field_type == 'difference_condition':
                    headers.append({
                        'label': f'新{label}',
                        'bindingKey': item.get('bindingKey', ''),
                        'type': 'query_new',
                        'hasDefaultValue': False,
                        'defaultValue': ''
                    })
                    headers.append({
                        'label': f'原{label}',
                        'bindingKey': item.get('bindingKey', ''),
                        'type': 'query_origin',
                        'hasDefaultValue': False,
                        'defaultValue': ''
                    })
                else:
                    has_default = bool(item.get('defaultValue'))
                    headers.append({
                        'label': label,
                        'bindingKey': item.get('bindingKey', ''),
                        'type': 'query',
                        'hasDefaultValue': has_default,
                        'defaultValue': item.get('defaultValue', '')
                    })

            # 更新字段：新值列全部在前，原值列全部在后
            new_headers = []
            origin_headers = []
            for item in update_items:
                input_type = item.get('inputType', '')
                parent_label = item.get('label', '')
                parent_binding_key = item.get('bindingKey', '')
                sub_fields = item.get('subFields', [])

                if input_type == 'calculated':
                    # 计算字段不显示在模板中
                    continue

                if input_type == 'supplement':
                    has_new_default = bool(item.get('newDefaultValue'))
                    has_origin_default = bool(item.get('originDefaultValue'))

                    new_headers.append({
                        'label': f'新{parent_label}',
                        'bindingKey': parent_binding_key,
                        'type': 'update_new',
                        'hasDefaultValue': has_new_default,
                        'defaultValue': item.get('newDefaultValue', '')
                    })
                    origin_headers.append({
                        'label': f'原{parent_label}',
                        'bindingKey': parent_binding_key,
                        'type': 'update_origin',
                        'hasDefaultValue': has_origin_default,
                        'defaultValue': item.get('originDefaultValue', '')
                    })

                    # 子字段：只显示辅助字段，普通字段由后端自动回填
                    for sub_field in sub_fields:
                        if sub_field.get('type') != 'auxiliary':
                            continue
                        sub_label = sub_field.get('label', '')
                        sub_binding_key = sub_field.get('bindingKey', '')

                        new_headers.append({
                            'label': f'新{sub_label}',
                            'bindingKey': sub_binding_key,
                            'type': 'update_new_sub',
                            'hasDefaultValue': False
                        })
                        origin_headers.append({
                            'label': f'原{sub_label}',
                            'bindingKey': sub_binding_key,
                            'type': 'update_origin_sub',
                            'hasDefaultValue': False
                        })
                else:
                    label = item.get('label', '')
                    binding_key = item.get('bindingKey', '')
                    has_new_default = bool(item.get('newDefaultValue'))
                    has_origin_default = bool(item.get('originDefaultValue'))

                    new_headers.append({
                        'label': f'新{label}',
                        'bindingKey': binding_key,
                        'type': 'update_new',
                        'hasDefaultValue': has_new_default,
                        'defaultValue': item.get('newDefaultValue', '')
                    })
                    origin_headers.append({
                        'label': f'原{label}',
                        'bindingKey': binding_key,
                        'type': 'update_origin',
                        'hasDefaultValue': has_origin_default,
                        'defaultValue': item.get('originDefaultValue', '')
                    })

            headers.extend(new_headers)
            headers.extend(origin_headers)

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header['label'])
                cell.font = header_font
                cell.alignment = header_alignment

                # 如果有默认值，使用黄色背景
                if header.get('hasDefaultValue'):
                    cell.fill = default_value_fill
                else:
                    cell.fill = header_fill

                col_letter = chr(64 + (col_num % 26)) if col_num <= 26 else chr(64 + (col_num // 26)) + chr(
                    64 + (col_num % 26))
                ws.column_dimensions[col_letter].width = 15

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{form_name}_{timestamp}.xlsx".replace('/', '-').replace('\\', '-')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            from urllib.parse import quote
            encoded_filename = quote(filename)
            response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"

            return response

        except json.JSONDecodeError as e:
            return JsonResponse({'success': False, 'message': f'JSON 解析失败：{str(e)}'}, status=400)
        except Exception as e:
            print(f"处理异常：{e}")
            return JsonResponse({'success': False, 'message': f'服务器错误：{str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': '仅支持 POST 请求'}, status=405)


def process_single_sheet_import(ws, query_items_data, update_items_data, query_values, form_name, table_name_list=None,
                                query_mode='strict', append_ops_remark=True, table_aliases=None):
    """处理单个Sheet的导入（用于多Sheet批量导入）"""
    try:

        config = {
            'formName': form_name,
            'tableNameList': table_name_list or [],  # 添加表名列表
            'tableAliases': table_aliases or {},  # 添加表别名映射
            'queryItems': query_items_data,
            'updateItems': update_items_data,
            'queryMode': query_mode,  # 添加查询模式（strict/loose）
            'appendOpsRemark': append_ops_remark  # 添加操作备注配置
        }

        # 读取表头
        headers = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value is not None and str(cell_value).strip():
                headers[str(cell_value).strip()] = col

        # 检查是否有有效数据
        required_columns = []
        for item in query_items_data:
            field_type = item.get('type', 'text')
            label = item.get('label', '')
            if field_type in ('calculated', 'subquery'):
                continue
            elif field_type == 'difference_condition':
                required_columns.append(f'新{label}')
            else:
                required_columns.append(label)

        for item in update_items_data:
            if item.get('inputType') == 'supplement':
                required_columns.append(f'新{item.get("label", "")}')
            else:
                required_columns.append(f'新{item.get("label", "")}')

        has_valid_data = any(col in headers for col in required_columns)

        if not has_valid_data or len(headers) == 0:
            return {
                'success': False,
                'message': '数据表中无有效的数据，请检查 Excel 文件格式是否正确',
                'forward_sqls': [],
                'backward_sqls': []
            }

        if ws.max_row < 2:
            return {
                'success': False,
                'message': 'Excel 文件没有数据行，请至少填写一行数据',
                'forward_sqls': [],
                'backward_sqls': []
            }

        # 统计有效数据行数
        valid_data_rows = 0
        for row_idx in range(2, ws.max_row + 1):
            has_required_value = False
            for col_name in required_columns:
                col_num = headers.get(col_name)
                if col_num:
                    cell_value = ws.cell(row=row_idx, column=col_num).value
                    if cell_value is not None and str(cell_value).strip():
                        has_required_value = True
                        break
            if has_required_value:
                valid_data_rows += 1

        if valid_data_rows == 0:
            return {
                'success': False,
                'message': f'Excel 中没有有效的必填数据（共{ws.max_row - 1}行，但都没有必填字段的值）',
                'forward_sqls': [],
                'backward_sqls': []
            }

        # ==================== 第一步：收集所有补充框主字段值（按辅助字段条件分组） ====================
        supplement_queries = {}  # {query_key: {'tableName': ..., 'mainField': ..., 'subFields': ..., 'groups': {aux_group_key: {'auxiliaryValues': [...], 'new_values': set(), 'origin_values': set()}}}}

        for row_idx in range(2, ws.max_row + 1):
            form_values, missing_columns = build_form_values_from_excel(ws, row_idx, headers, query_items_data,
                                                                        update_items_data)

            if missing_columns:
                continue

            # 收集补充框主字段值
            for item in update_items_data:
                if item.get('inputType') == 'supplement':
                    main_table = item.get('mainTable', '')
                    main_field = item.get('mainField', '')
                    parent_key = item.get('bindingKey', '')
                    connected_tables = item.get('connectedTable', [])

                    if main_table and main_field:
                        query_key = f"{main_table}_{main_field}"
                        if query_key not in supplement_queries:
                            supplement_queries[query_key] = {
                                'tableName': main_table,
                                'mainField': main_field,
                                'subFields': item.get('subFields', []),
                                'groups': {}
                            }

                        # 收集新值和原值：支持 bindingKey_tableName 查找
                        value_data = None
                        if connected_tables:
                            for table in connected_tables:
                                unique_key = f"{parent_key}_{table}"
                                if unique_key in form_values:
                                    value_data = form_values[unique_key]
                                    break
                        if value_data is None:
                            value_data = form_values.get(parent_key, {})

                        new_value = value_data.get('newValue', '')
                        origin_value = value_data.get('originValue', '')

                        # 按辅助字段条件分组收集
                        new_aux = get_auxiliary_values_from_excel(ws, row_idx, headers, item, is_origin=False)
                        origin_aux = get_auxiliary_values_from_excel(ws, row_idx, headers, item, is_origin=True)
                        new_group_key = build_auxiliary_group_key(new_aux)
                        origin_group_key = build_auxiliary_group_key(origin_aux)

                        groups = supplement_queries[query_key]['groups']
                        if new_group_key not in groups:
                            groups[new_group_key] = {'auxiliaryValues': new_aux, 'new_values': set(),
                                                     'origin_values': set()}
                        if origin_group_key not in groups:
                            groups[origin_group_key] = {'auxiliaryValues': origin_aux, 'new_values': set(),
                                                        'origin_values': set()}

                        if new_value:
                            groups[new_group_key]['new_values'].add(new_value)
                        if origin_value:
                            groups[origin_group_key]['origin_values'].add(origin_value)

        # ==================== 第二步：批量查询补充框数据 ====================
        supplement_data_cache = {}  # {cache_key: {subField: value}}

        from django.db import connection

        for query_key, query_info in supplement_queries.items():
            table_name = query_info['tableName']
            main_field = query_info['mainField']
            sub_fields = query_info['subFields']

            # 构建查询字段
            select_fields = [main_field]
            for sub_field in sub_fields:
                if isinstance(sub_field, dict):
                    field_name = sub_field.get('dbField') or sub_field.get('bindingKey')
                    if field_name:
                        select_fields.append(field_name)
                elif isinstance(sub_field, str):
                    select_fields.append(sub_field)
            fields_str = ', '.join(select_fields)

            for group_key, group_info in query_info['groups'].items():
                main_values = list(group_info['new_values'] | group_info['origin_values'])
                if not main_values:
                    continue

                auxiliary_values = group_info['auxiliaryValues']
                values_str = ', '.join(["'" + str(v).replace("'", "''") + "'" for v in main_values])
                where_conditions = [f"{main_field} IN ({values_str})"]
                where_conditions.extend(build_auxiliary_where_sql(auxiliary_values))
                sql = f"SELECT {fields_str} FROM {table_name} WHERE {' AND '.join(where_conditions)}"

                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    rows = cursor.fetchall()

                    for row in rows:
                        row_dict = {}
                        for idx, field in enumerate(select_fields):
                            row_dict[field] = row[idx]

                        main_val = row_dict.get(main_field, '')
                        # 同时缓存新值和原值两种标志，方便后续查找
                        supplement_data_cache[
                            build_auxiliary_cache_key(query_key, False, auxiliary_values, main_val)] = row_dict
                        supplement_data_cache[
                            build_auxiliary_cache_key(query_key, True, auxiliary_values, main_val)] = row_dict

        print(f"补充框数据缓存: {len(supplement_data_cache)} 条")

        # ==================== 第三步：处理每一行数据 ====================
        all_sql_statements = []
        success_count = 0
        fail_count = 0

        # 在最后一列添加"失败原因"列
        from openpyxl.styles import Font, Alignment, PatternFill
        fail_column = ws.max_column + 1
        ws.cell(row=1, column=fail_column, value='失败原因').font = Font(bold=True)
        ws.cell(row=1, column=fail_column).alignment = Alignment(horizontal='center')
        ws.cell(row=1, column=fail_column).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC",
                                                              fill_type="solid")

        for row_idx in range(2, ws.max_row + 1):

            form_values, missing_columns = build_form_values_from_excel(ws, row_idx, headers, query_items_data,
                                                                        update_items_data)

            if missing_columns:
                fail_count += 1
                missing_cols_str = ', '.join(missing_columns)
                ws.cell(row=row_idx, column=fail_column, value=f'缺少必需的列：{missing_cols_str}')
                ws.cell(row=row_idx, column=fail_column).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC",
                                                                            fill_type="solid")
                continue

            # ==================== 添加公共字段到form_values ====================
            # 将query_values中的公共字段添加到form_values中，以便校验通过
            if query_values:
                for field_name in ['filePrefix', 'onesLink', 'dynamicNo', 'ops_remark']:
                    if field_name in query_values:
                        value_data = query_values[field_name]
                        if isinstance(value_data, dict):
                            form_values[field_name] = value_data
                        else:
                            form_values[field_name] = {'value': str(value_data)}

            # ==================== 使用缓存的补充框数据填充子字段 ====================
            for item in update_items_data:
                if item.get('inputType') == 'supplement':
                    main_table = item.get('mainTable', '')
                    main_field = item.get('mainField', '')
                    parent_key = item.get('bindingKey', '')
                    sub_fields = item.get('subFields', [])
                    connected_tables = item.get('connectedTable', [])

                    if main_table and main_field and sub_fields:
                        query_key = f"{main_table}_{main_field}"

                        # 支持 bindingKey_tableName 查找主字段值
                        parent_value_data = None
                        parent_unique_key = None
                        if connected_tables:
                            for table in connected_tables:
                                unique_key = f"{parent_key}_{table}"
                                if unique_key in form_values:
                                    parent_value_data = form_values[unique_key]
                                    parent_unique_key = unique_key
                                    break
                        if parent_value_data is None:
                            parent_value_data = form_values.get(parent_key, {})
                            parent_unique_key = parent_key

                        new_value = parent_value_data.get('newValue', '')
                        origin_value = parent_value_data.get('originValue', '')

                        # 读取当前行的辅助字段条件
                        new_aux = get_auxiliary_values_from_excel(ws, row_idx, headers, item, is_origin=False)
                        origin_aux = get_auxiliary_values_from_excel(ws, row_idx, headers, item, is_origin=True)

                        # 填充新值的子字段
                        if new_value and str(new_value).strip():
                            cache_key = build_auxiliary_cache_key(query_key, False, new_aux, new_value)
                            if cache_key in supplement_data_cache:
                                row_data = supplement_data_cache[cache_key]
                                for sub_field in sub_fields:
                                    if isinstance(sub_field, dict):
                                        sub_binding_key = sub_field.get('bindingKey', '')
                                        db_field = sub_field.get('dbField') or sub_binding_key
                                        sub_value = row_data.get(db_field, '')

                                        # 子字段按 connected_tables 存储，与 generate_update_sql 查找一致
                                        if connected_tables:
                                            for sub_table in connected_tables:
                                                form_values[f"{sub_binding_key}_{sub_table}"] = {
                                                    'newValue': str(sub_value).strip() if sub_value is not None else '',
                                                    'originValue': '',
                                                    'inputType': 'supplement-sub',
                                                    'fieldType': 'supplement-sub',
                                                    'parentKey': parent_unique_key,
                                                    'label': sub_field.get('label', '')
                                                }
                                        else:
                                            form_values[sub_binding_key] = {
                                                'newValue': str(sub_value).strip() if sub_value is not None else '',
                                                'originValue': '',
                                                'inputType': 'supplement-sub',
                                                'fieldType': 'supplement-sub',
                                                'parentKey': parent_unique_key,
                                                'label': sub_field.get('label', '')
                                            }

                        # 填充原值的子字段
                        if origin_value and str(origin_value).strip():
                            cache_key = build_auxiliary_cache_key(query_key, True, origin_aux, origin_value)
                            if cache_key in supplement_data_cache:
                                row_data = supplement_data_cache[cache_key]
                                for sub_field in sub_fields:
                                    if isinstance(sub_field, dict):
                                        sub_binding_key = sub_field.get('bindingKey', '')
                                        db_field = sub_field.get('dbField') or sub_binding_key
                                        sub_value = row_data.get(db_field, '')

                                        if connected_tables:
                                            for sub_table in connected_tables:
                                                sub_unique_key = f"{sub_binding_key}_{sub_table}"
                                                if sub_unique_key in form_values:
                                                    form_values[sub_unique_key]['originValue'] = str(
                                                        sub_value).strip() if sub_value is not None else ''
                                                else:
                                                    form_values[sub_unique_key] = {
                                                        'newValue': '',
                                                        'originValue': str(sub_value).strip() if sub_value is not None else '',
                                                        'inputType': 'supplement-sub',
                                                        'fieldType': 'supplement-sub',
                                                        'parentKey': parent_unique_key,
                                                        'label': sub_field.get('label', '')
                                                    }
                                        else:
                                            if sub_binding_key in form_values:
                                                form_values[sub_binding_key]['originValue'] = str(
                                                    sub_value).strip() if sub_value is not None else ''
                                            else:
                                                form_values[sub_binding_key] = {
                                                    'newValue': '',
                                                    'originValue': str(sub_value).strip() if sub_value is not None else '',
                                                    'inputType': 'supplement-sub',
                                                    'fieldType': 'supplement-sub',
                                                    'parentKey': parent_unique_key,
                                                    'label': sub_field.get('label', '')
                                                }

            validation_result = validate_form_data(config, form_values, query_values)

            if not validation_result['success']:
                fail_count += 1
                # 标记具体的校验错误
                fail_reason = '; '.join(validation_result.get('errors', []))
                ws.cell(row=row_idx, column=fail_column, value=fail_reason)
                ws.cell(row=row_idx, column=fail_column).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC",
                                                                            fill_type="solid")
            else:
                sql_result = generate_update_sql(config, form_values)

                # 检查是否有缺失字段错误
                if sql_result.get('missing_field_errors'):
                    fail_count += 1
                    fail_reason = '; '.join(sql_result.get('missing_field_errors', []))
                    ws.cell(row=row_idx, column=fail_column, value=fail_reason)
                    ws.cell(row=row_idx, column=fail_column).fill = PatternFill(start_color="FFFFCC",
                                                                                end_color="FFFFCC", fill_type="solid")
                elif sql_result['forward_sqls'] and sql_result['backward_sqls']:
                    all_sql_statements.append({
                        'row': row_idx,
                        'forward_sqls': sql_result['forward_sqls'],
                        'backward_sqls': sql_result['backward_sqls']
                    })
                    success_count += 1
                else:
                    fail_count += 1
                    ws.cell(row=row_idx, column=fail_column, value='未生成有效的 SQL 语句')
                    ws.cell(row=row_idx, column=fail_column).fill = PatternFill(start_color="FFFFCC",
                                                                                end_color="FFFFCC", fill_type="solid")

        if success_count == 0:
            return {
                'success': False,
                'message': f'没有成功处理任何数据（失败{fail_count}条）',
                'forward_sqls': [],
                'backward_sqls': []
            }

        # 合并相同修改条件的SQL
        merged_result = merge_sql_statements(all_sql_statements)

        return {
            'success': True,
            'message': f'成功处理{success_count}条数据',
            'forward_sqls': merged_result['forward_sqls'],
            'backward_sqls': merged_result['backward_sqls'],
            'success_count': success_count,
            'fail_count': fail_count
        }

    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"处理Sheet失败: {error_detail}")
        return {
            'success': False,
            'message': f'处理失败：{str(e)}',
            'forward_sqls': [],
            'backward_sqls': []
        }


@csrf_exempt
def batch_import(request):
    """批量导入数据（支持文件上传和在线编辑器JSON数据）"""
    if request.method == 'POST':
        try:
            # 判断是文件上传还是JSON数据
            content_type = request.content_type or ''
            is_json = 'application/json' in content_type

            if is_json:
                # 在线编辑器模式：直接接收JSON数据
                data = json.loads(request.body)
                config = data.get('config', {})
                query_values = data.get('queryValues', {})
                rows = data.get('rows', [])

                if not rows:
                    return JsonResponse({'success': False, 'message': '没有数据可导入'}, status=400)

                form_name = config.get('formName', '模板')
                query_items = config.get('queryItems', [])
                update_items = config.get('updateItems', [])

                # 校验公共字段
                common_fields = ['filePrefix', 'onesLink', 'dynamicNo']
                for field_name in common_fields:
                    value_data = query_values.get(field_name, {})
                    if isinstance(value_data, dict):
                        value = value_data.get('value', '')
                    elif isinstance(value_data, str):
                        value = value_data
                    else:
                        value = ''
                    if not value or str(value).strip() == '':
                        return JsonResponse({'success': False, 'message': f'{field_name}不能为空'}, status=400)

                # 构建headers映射（用于兼容现有逻辑）
                # 收集所有可能的列名：查询字段 -> 所有新值列 -> 所有原值列
                all_labels = []
                for item in query_items:
                    field_type = item.get('type', 'text')
                    label = item.get('label', '')
                    if field_type in ('calculated', 'subquery'):
                        continue
                    elif field_type == 'difference_condition':
                        all_labels.append(f'新{label}')
                        all_labels.append(f'原{label}')
                    else:
                        all_labels.append(label)
                for item in update_items:
                    if item.get('inputType') == 'supplement':
                        all_labels.append(f'新{item.get("label", "")}')
                        for sf in item.get('subFields', []):
                            if sf.get('type') != 'auxiliary':
                                continue
                            all_labels.append(f'新{sf.get("label", "")}')
                    else:
                        all_labels.append(f'新{item.get("label", "")}')
                for item in update_items:
                    if item.get('inputType') == 'supplement':
                        all_labels.append(f'原{item.get("label", "")}')
                        for sf in item.get('subFields', []):
                            if sf.get('type') != 'auxiliary':
                                continue
                            all_labels.append(f'原{sf.get("label", "")}')
                    else:
                        all_labels.append(f'原{item.get("label", "")}')

                # 创建一个虚拟的Workbook用于失败报告
                wb = Workbook()
                ws = wb.active
                ws.title = '导入数据'

                # 写入表头
                for col_num, label in enumerate(all_labels, 1):
                    if label:
                        ws.cell(row=1, column=col_num, value=label)

                # 写入数据行
                for row_idx, row_data in enumerate(rows, 2):
                    for col_num, label in enumerate(all_labels, 1):
                        if label and label in row_data:
                            ws.cell(row=row_idx, column=col_num, value=row_data[label])

                # 构建headers字典（label -> col_num）
                headers = {}
                for col_num, label in enumerate(all_labels, 1):
                    if label:
                        headers[label] = col_num

                # 校验数据
                required_columns = []
                for item in query_items:
                    field_type = item.get('type', 'text')
                    label = item.get('label', '')
                    if field_type in ('calculated', 'subquery'):
                        continue
                    elif field_type == 'difference_condition':
                        required_columns.append(f'新{label}')
                    else:
                        required_columns.append(label)
                for item in update_items:
                    if item.get('inputType') == 'supplement':
                        required_columns.append(f'新{item.get("label", "")}')
                    else:
                        required_columns.append(f'新{item.get("label", "")}')

                has_valid_data = any(col in headers for col in required_columns)
                if not has_valid_data or len(headers) == 0:
                    return JsonResponse({'success': False, 'message': '数据表中无有效的数据，请检查数据格式是否正确'}, status=400)

                if len(rows) == 0:
                    return JsonResponse({'success': False, 'message': '没有数据行，请至少填写一行数据'}, status=400)

                valid_data_rows = 0
                for row_data in rows:
                    has_required_value = False
                    for col_name in required_columns:
                        if col_name and col_name in row_data and row_data[col_name] is not None and str(row_data[col_name]).strip():
                            has_required_value = True
                            break
                    if has_required_value:
                        valid_data_rows += 1

                if valid_data_rows == 0:
                    return JsonResponse({'success': False, 'message': '没有有效的必填数据'}, status=400)

                # 添加失败原因列
                fail_column = ws.max_column + 1
                ws.cell(row=1, column=fail_column, value='失败原因')
                ws.cell(row=1, column=fail_column).font = Font(bold=True)
                ws.cell(row=1, column=fail_column).alignment = Alignment(horizontal='center')
                ws.cell(row=1, column=fail_column).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

                total_rows = len(rows)
                success_count = 0
                fail_count = 0
                all_sql_statements = []

                # 收集补充框主字段值（按辅助字段条件分组）
                supplement_queries = {}
                for row_idx, row_data in enumerate(rows, 2):
                    for item in update_items:
                        if item.get('inputType') == 'supplement':
                            main_table = item.get('mainTable', '')
                            main_field = item.get('mainField', '')
                            query_key = f"{main_table}_{main_field}"
                            if main_table and main_field:
                                if query_key not in supplement_queries:
                                    supplement_queries[query_key] = {
                                        'tableName': main_table,
                                        'mainField': main_field,
                                        'subFields': item.get('subFields', []),
                                        'groups': {}
                                    }
                                label = item.get('label', '')
                                new_value = row_data.get(f'新{label}', '')
                                origin_value = row_data.get(f'原{label}', '')

                                new_aux = get_auxiliary_values_from_row_data(row_data, item, is_origin=False)
                                origin_aux = get_auxiliary_values_from_row_data(row_data, item, is_origin=True)
                                new_group_key = build_auxiliary_group_key(new_aux)
                                origin_group_key = build_auxiliary_group_key(origin_aux)

                                groups = supplement_queries[query_key]['groups']
                                if new_group_key not in groups:
                                    groups[new_group_key] = {'auxiliaryValues': new_aux, 'new_values': set(),
                                                             'origin_values': set()}
                                if origin_group_key not in groups:
                                    groups[origin_group_key] = {'auxiliaryValues': origin_aux, 'new_values': set(),
                                                                'origin_values': set()}

                                if new_value is not None and str(new_value).strip():
                                    groups[new_group_key]['new_values'].add(str(new_value).strip())
                                if origin_value is not None and str(origin_value).strip():
                                    groups[origin_group_key]['origin_values'].add(str(origin_value).strip())

                # 批量查询补充框数据
                supplement_data_cache = {}
                from django.db import connection
                for query_key, query_info in supplement_queries.items():
                    table_name = query_info['tableName']
                    main_field = query_info['mainField']
                    sub_fields = query_info['subFields']
                    select_fields = [main_field]
                    for sub_field in sub_fields:
                        if isinstance(sub_field, dict):
                            field_name = sub_field.get('dbField') or sub_field.get('bindingKey')
                            if field_name:
                                select_fields.append(field_name)
                        elif isinstance(sub_field, str):
                            select_fields.append(sub_field)
                    fields_str = ', '.join(select_fields)
                    for group_key, group_info in query_info['groups'].items():
                        main_values = list(group_info['new_values'] | group_info['origin_values'])
                        if not main_values:
                            continue
                        auxiliary_values = group_info['auxiliaryValues']
                        values_str = ', '.join(["'" + str(v).replace("'", "''") + "'" for v in main_values])
                        where_conditions = [f"{main_field} IN ({values_str})"]
                        where_conditions.extend(build_auxiliary_where_sql(auxiliary_values))
                        sql = f"SELECT {fields_str} FROM {table_name} WHERE {' AND '.join(where_conditions)}"
                        with connection.cursor() as cursor:
                            cursor.execute(sql)
                            db_rows = cursor.fetchall()
                        for row in db_rows:
                            row_dict = {}
                            for idx, field in enumerate(select_fields):
                                row_dict[field] = row[idx]
                            main_val = row_dict.get(main_field, '')
                            supplement_data_cache[
                                build_auxiliary_cache_key(query_key, False, auxiliary_values, main_val)] = row_dict
                            supplement_data_cache[
                                build_auxiliary_cache_key(query_key, True, auxiliary_values, main_val)] = row_dict

                # 预编译补充框item
                supplement_items = []
                for item in update_items:
                    if item.get('inputType') == 'supplement':
                        main_table = item.get('mainTable', '')
                        main_field = item.get('mainField', '')
                        parent_key = item.get('bindingKey', '')
                        sub_fields = item.get('subFields', [])
                        connected_tables = item.get('connectedTable', [])
                        sub_fields_dict = {}
                        for sf in sub_fields:
                            if isinstance(sf, dict):
                                bk = sf.get('bindingKey', '')
                                sub_fields_dict[bk] = sf
                        if main_table and main_field and sub_fields_dict:
                            supplement_items.append({
                                'main_table': main_table,
                                'main_field': main_field,
                                'parent_key': parent_key,
                                'sub_fields_dict': sub_fields_dict,
                                'connected_tables': connected_tables,
                                'query_key': f"{main_table}_{main_field}",
                                'item': item
                            })

                # 处理每一行数据
                row_data_list = []
                for row_idx, row_data in enumerate(rows, 2):
                    form_values, missing_columns = build_form_values_from_rows(row_data, query_items, update_items)
                    if query_values:
                        for field_name in ['filePrefix', 'onesLink', 'dynamicNo', 'ops_remark']:
                            if field_name in query_values:
                                value_data = query_values[field_name]
                                if isinstance(value_data, dict):
                                    form_values[field_name] = value_data
                                else:
                                    form_values[field_name] = {'value': str(value_data)}
                    row_data_list.append(
                        {'row_idx': row_idx, 'form_values': form_values, 'missing_columns': missing_columns,
                         'row_data': row_data})

                # 多线程并行处理
                max_workers = min(8, total_rows)
                from concurrent.futures import ThreadPoolExecutor, as_completed

                def process_row(row_data):
                    row_idx = row_data['row_idx']
                    form_values = row_data['form_values']
                    missing_columns = row_data['missing_columns']
                    current_row_data = row_data.get('row_data', {})
                    if missing_columns:
                        return {'row_idx': row_idx, 'status': 'missing', 'missing_columns': missing_columns}
                    for s_item in supplement_items:
                        query_key = s_item['query_key']
                        parent_key = s_item['parent_key']
                        sub_fields_dict = s_item['sub_fields_dict']
                        connected_tables = s_item['connected_tables']
                        item = s_item.get('item')
                        parent_value_data = None
                        parent_unique_key = None
                        if connected_tables:
                            for table in connected_tables:
                                unique_key = f"{parent_key}_{table}"
                                if unique_key in form_values:
                                    parent_value_data = form_values[unique_key]
                                    parent_unique_key = unique_key
                                    break
                        if parent_value_data is None:
                            parent_value_data = form_values.get(parent_key, {})
                            parent_unique_key = parent_key
                        if not parent_value_data:
                            continue
                        new_value = parent_value_data.get('newValue', '')
                        origin_value = parent_value_data.get('originValue', '')

                        # 读取当前行的辅助字段条件
                        new_aux = get_auxiliary_values_from_row_data(current_row_data, item, is_origin=False) if item else []
                        origin_aux = get_auxiliary_values_from_row_data(current_row_data, item, is_origin=True) if item else []

                        if new_value and str(new_value).strip():
                            cache_key = build_auxiliary_cache_key(query_key, False, new_aux, new_value)
                            row_data_cache = supplement_data_cache.get(cache_key)
                            if row_data_cache:
                                for sub_binding_key, sub_field in sub_fields_dict.items():
                                    db_field = sub_field.get('dbField') or sub_binding_key
                                    sub_value = row_data_cache.get(db_field, '')
                                    if connected_tables:
                                        for sub_table in connected_tables:
                                            form_values[f"{sub_binding_key}_{sub_table}"] = {
                                                'newValue': str(sub_value) if sub_value is not None else '',
                                                'originValue': '',
                                                'inputType': 'supplement-sub',
                                                'fieldType': 'supplement-sub',
                                                'parentKey': parent_unique_key,
                                                'label': sub_field.get('label', '')}
                                    else:
                                        form_values[sub_binding_key] = {
                                            'newValue': str(sub_value) if sub_value is not None else '',
                                            'originValue': '',
                                            'inputType': 'supplement-sub',
                                            'fieldType': 'supplement-sub',
                                            'parentKey': parent_unique_key,
                                            'label': sub_field.get('label', '')}
                        if origin_value and str(origin_value).strip():
                            cache_key = build_auxiliary_cache_key(query_key, True, origin_aux, origin_value)
                            row_data_cache = supplement_data_cache.get(cache_key)
                            if row_data_cache:
                                for sub_binding_key, sub_field in sub_fields_dict.items():
                                    db_field = sub_field.get('dbField') or sub_binding_key
                                    sub_value = row_data_cache.get(db_field, '')
                                    if connected_tables:
                                        for sub_table in connected_tables:
                                            sub_unique_key = f"{sub_binding_key}_{sub_table}"
                                            if sub_unique_key in form_values:
                                                form_values[sub_unique_key]['originValue'] = str(sub_value) if sub_value is not None else ''
                                            else:
                                                form_values[sub_unique_key] = {
                                                    'newValue': '',
                                                    'originValue': str(sub_value) if sub_value is not None else '',
                                                    'inputType': 'supplement-sub',
                                                    'fieldType': 'supplement-sub',
                                                    'parentKey': parent_unique_key,
                                                    'label': sub_field.get('label', '')}
                                    else:
                                        if sub_binding_key in form_values:
                                            form_values[sub_binding_key]['originValue'] = str(sub_value) if sub_value is not None else ''
                                        else:
                                            form_values[sub_binding_key] = {
                                                'newValue': '',
                                                'originValue': str(sub_value) if sub_value is not None else '',
                                                'inputType': 'supplement-sub',
                                                'fieldType': 'supplement-sub',
                                                'parentKey': parent_unique_key,
                                                'label': sub_field.get('label', '')}
                    validation_result = validate_form_data(config, form_values, query_values)
                    if not validation_result['success']:
                        return {'row_idx': row_idx, 'status': 'validation_fail',
                                'fail_reason': '; '.join(validation_result.get('errors', []))}
                    sql_result = generate_update_sql(config, form_values)
                    if sql_result.get('missing_field_errors'):
                        return {'row_idx': row_idx, 'status': 'missing_field',
                                'fail_reason': '; '.join(sql_result.get('missing_field_errors', []))}
                    if sql_result['forward_sqls'] and sql_result['backward_sqls']:
                        return {'row_idx': row_idx, 'status': 'success',
                                'forward_sqls': sql_result['forward_sqls'],
                                'backward_sqls': sql_result['backward_sqls']}
                    return {'row_idx': row_idx, 'status': 'no_sql', 'fail_reason': '未生成有效的 SQL 语句'}

                results = []
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = {executor.submit(process_row, row_data): row_data['row_idx'] for row_data in row_data_list}
                    for future in as_completed(futures):
                        result = future.result()
                        results.append(result)
                results.sort(key=lambda r: r['row_idx'])

                row_errors = []
                for result in results:
                    row_idx = result['row_idx']
                    status = result['status']
                    if status == 'missing':
                        fail_count += 1
                        missing_cols_str = ', '.join(result['missing_columns'])
                        fail_reason = f'缺少必需的列：{missing_cols_str}'
                        row_errors.append({'rowIndex': row_idx, 'failReason': fail_reason})
                    elif status == 'validation_fail':
                        fail_count += 1
                        row_errors.append({'rowIndex': row_idx, 'failReason': result['fail_reason']})
                    elif status == 'missing_field':
                        fail_count += 1
                        row_errors.append({'rowIndex': row_idx, 'failReason': result['fail_reason']})
                    elif status == 'success':
                        all_sql_statements.append({
                            'row': row_idx,
                            'forward_sqls': result['forward_sqls'],
                            'backward_sqls': result['backward_sqls']
                        })
                        success_count += 1
                    elif status == 'no_sql':
                        fail_count += 1
                        row_errors.append({'rowIndex': row_idx, 'failReason': result['fail_reason']})

                dynamic_no = ''
                file_prefix = form_name
                if query_values:
                    dynamic_no_data = query_values.get('dynamicNo', '')
                    if isinstance(dynamic_no_data, dict):
                        dynamic_no = dynamic_no_data.get('value', '')
                    elif isinstance(dynamic_no_data, str):
                        dynamic_no = dynamic_no_data
                    file_prefix_data = query_values.get('filePrefix', '')
                    if isinstance(file_prefix_data, dict):
                        file_prefix = file_prefix_data.get('value', '') or form_name
                    elif isinstance(file_prefix_data, str) and file_prefix_data.strip():
                        file_prefix = file_prefix_data.strip()

                save_dir = get_save_path_from_config()
                os.makedirs(save_dir, exist_ok=True)
                all_success = (fail_count == 0 and success_count > 0)
                sql_file_path = None
                if all_success and all_sql_statements:
                    sql_filename = f"{dynamic_no}_{file_prefix}.sql"
                    sql_filepath = os.path.join(save_dir, sql_filename)
                    merged_result = merge_sql_statements(all_sql_statements)
                    sql_content = []
                    sql_content.append("1.执行语句")
                    for idx, sql in enumerate(merged_result['forward_sqls'], 1):
                        if not sql.rstrip().endswith(';'):
                            sql = sql.rstrip() + ';'
                        sql_content.append(sql)
                        sql_content.append("")
                    sql_content.append("2.回退语句")
                    for idx, sql in enumerate(merged_result['backward_sqls'], 1):
                        if not sql.rstrip().endswith(';'):
                            sql = sql.rstrip() + ';'
                        sql_content.append(sql)
                        sql_content.append("")
                    database_ip_ids = config.get('databaseIpIds', [])
                    if database_ip_ids:
                        db_configs = DatabaseIPConfig.objects.filter(id__in=database_ip_ids).order_by('id')
                        if db_configs:
                            sql_content.append("3.数据库")
                            for db_config in db_configs:
                                sql_content.append(f"ip：{db_config.ip_address}")
                                sql_content.append(f"库名：{db_config.database_name}")
                                sql_content.append("")
                    with open(sql_filepath, 'w', encoding='utf-8') as f:
                        f.write('\n'.join(sql_content))
                    sql_file_path = sql_filepath

                # 记录表单调用日志（在线编辑批量导入）
                try:
                    form_id = config.get('formId')
                    if form_id:
                        form_config = FormConfig.objects.get(id=form_id)
                        record_form_usage(form_config, source='dynamic')
                except Exception:
                    # 调用统计不应影响主流程
                    pass

                # 在线编辑模式不生成 Excel 错误文件，错误信息通过 rowErrors 返回给前端回显
                return JsonResponse({
                    'success': True,
                    'message': f'批量导入完成，成功{success_count}条，失败{fail_count}条',
                    'sqlFilePath': sql_file_path,
                    'excelFilePath': None,
                    'totalRows': total_rows,
                    'successCount': success_count,
                    'failCount': fail_count,
                    'rowErrors': row_errors
                })

            # ==================== 原有文件上传模式 ====================
            file = request.FILES.get('file')
            config_json = request.POST.get('config')
            query_values_json = request.POST.get('queryValues')

            if not file or not config_json:
                return JsonResponse({'success': False, 'message': '缺少文件或配置参数'}, status=400)

            config = json.loads(config_json)
            form_name = config.get('formName', '模板')
            query_items = config.get('queryItems', [])
            update_items = config.get('updateItems', [])

            query_values = {}
            if query_values_json:
                try:
                    query_values = json.loads(query_values_json)
                except json.JSONDecodeError:
                    query_values = {}

            common_fields = ['filePrefix', 'onesLink', 'dynamicNo']
            for field_name in common_fields:
                value_data = query_values.get(field_name, {})
                if isinstance(value_data, dict):
                    value = value_data.get('value', '')
                elif isinstance(value_data, str):
                    value = value_data
                else:
                    value = ''
                if not value or str(value).strip() == '':
                    return JsonResponse({'success': False, 'message': f'{field_name}不能为空'}, status=400)

            wb = load_workbook(file)
            ws = wb.active

            headers = {}
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    headers[str(cell_value).strip()] = col

            required_columns = []
            for item in query_items:
                field_type = item.get('type', 'text')
                label = item.get('label', '')
                if field_type in ('calculated', 'subquery'):
                    continue
                elif field_type == 'difference_condition':
                    required_columns.append(f'新{label}')
                else:
                    required_columns.append(label)
            for item in update_items:
                if item.get('inputType') == 'supplement':
                    required_columns.append(f'新{item.get("label", "")}')
                else:
                    required_columns.append(f'新{item.get("label", "")}')

            has_valid_data = any(col in headers for col in required_columns)
            if not has_valid_data or len(headers) == 0:
                return JsonResponse({'success': False, 'message': '数据表中无有效的数据，请检查 Excel 文件格式是否正确'}, status=400)
            if ws.max_row < 2:
                return JsonResponse({'success': False, 'message': 'Excel 文件没有数据行，请至少填写一行数据'}, status=400)

            valid_data_rows = 0
            for row_idx in range(2, ws.max_row + 1):
                has_required_value = False
                for col_name in required_columns:
                    col_num = headers.get(col_name)
                    if col_num:
                        cell_value = ws.cell(row=row_idx, column=col_num).value
                        if cell_value is not None and str(cell_value).strip():
                            has_required_value = True
                            break
                if has_required_value:
                    valid_data_rows += 1
            if valid_data_rows == 0:
                return JsonResponse({'success': False, 'message': f'Excel 中没有有效的必填数据（共{ws.max_row - 1}行，但都没有必填字段的值）'}, status=400)

            fail_column = ws.max_column + 1
            ws.cell(row=1, column=fail_column, value='失败原因')
            ws.cell(row=1, column=fail_column).font = Font(bold=True)
            ws.cell(row=1, column=fail_column).alignment = Alignment(horizontal='center')
            ws.cell(row=1, column=fail_column).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            total_rows = ws.max_row - 1
            success_count = 0
            fail_count = 0
            all_sql_statements = []

            # 原有文件上传处理逻辑保持不变...
            # 收集所有补充框主字段值（按辅助字段条件分组）
            supplement_queries = {}
            for row_idx in range(2, ws.max_row + 1):
                for item in update_items:
                    if item.get('inputType') == 'supplement':
                        main_table = item.get('mainTable', '')
                        main_field = item.get('mainField', '')
                        parent_key = item.get('bindingKey', '')
                        connected_tables = item.get('connectedTable', [])
                        if main_table and main_field:
                            query_key = f"{main_table}_{main_field}"
                            if query_key not in supplement_queries:
                                supplement_queries[query_key] = {
                                    'tableName': main_table,
                                    'mainField': main_field,
                                    'subFields': item.get('subFields', []),
                                    'groups': {}
                                }
                            label = item.get('label', '')
                            new_col = headers.get(f'新{label}')
                            origin_col = headers.get(f'原{label}')
                            new_value = ''
                            origin_value = ''
                            if new_col:
                                new_value = ws.cell(row=row_idx, column=new_col).value
                            if origin_col:
                                origin_value = ws.cell(row=row_idx, column=origin_col).value

                            new_aux = get_auxiliary_values_from_excel(ws, row_idx, headers, item, is_origin=False)
                            origin_aux = get_auxiliary_values_from_excel(ws, row_idx, headers, item, is_origin=True)
                            new_group_key = build_auxiliary_group_key(new_aux)
                            origin_group_key = build_auxiliary_group_key(origin_aux)

                            groups = supplement_queries[query_key]['groups']
                            if new_group_key not in groups:
                                groups[new_group_key] = {'auxiliaryValues': new_aux, 'new_values': set(),
                                                         'origin_values': set()}
                            if origin_group_key not in groups:
                                groups[origin_group_key] = {'auxiliaryValues': origin_aux, 'new_values': set(),
                                                            'origin_values': set()}

                            if new_value is not None and str(new_value).strip():
                                groups[new_group_key]['new_values'].add(str(new_value).strip())
                            if origin_value is not None and str(origin_value).strip():
                                groups[origin_group_key]['origin_values'].add(str(origin_value).strip())

            # 批量查询补充框数据
            supplement_data_cache = {}
            from django.db import connection
            for query_key, query_info in supplement_queries.items():
                table_name = query_info['tableName']
                main_field = query_info['mainField']
                sub_fields = query_info['subFields']
                select_fields = [main_field]
                for sub_field in sub_fields:
                    if isinstance(sub_field, dict):
                        field_name = sub_field.get('dbField') or sub_field.get('bindingKey')
                        if field_name:
                            select_fields.append(field_name)
                    elif isinstance(sub_field, str):
                        select_fields.append(sub_field)
                fields_str = ', '.join(select_fields)
                for group_key, group_info in query_info['groups'].items():
                    main_values = list(group_info['new_values'] | group_info['origin_values'])
                    if not main_values:
                        continue
                    auxiliary_values = group_info['auxiliaryValues']
                    values_str = ', '.join(["'" + str(v).replace("'", "''") + "'" for v in main_values])
                    where_conditions = [f"{main_field} IN ({values_str})"]
                    where_conditions.extend(build_auxiliary_where_sql(auxiliary_values))
                    sql = f"SELECT {fields_str} FROM {table_name} WHERE {' AND '.join(where_conditions)}"
                    with connection.cursor() as cursor:
                        cursor.execute(sql)
                        rows = cursor.fetchall()
                    for row in rows:
                        row_dict = {}
                        for idx, field in enumerate(select_fields):
                            row_dict[field] = row[idx]
                        main_val = row_dict.get(main_field, '')
                        supplement_data_cache[
                            build_auxiliary_cache_key(query_key, False, auxiliary_values, main_val)] = row_dict
                        supplement_data_cache[
                            build_auxiliary_cache_key(query_key, True, auxiliary_values, main_val)] = row_dict

            print(f"补充框数据缓存: {len(supplement_data_cache)} 条")

            # 预编译补充框item
            supplement_items = []
            for item in update_items:
                if item.get('inputType') == 'supplement':
                    main_table = item.get('mainTable', '')
                    main_field = item.get('mainField', '')
                    parent_key = item.get('bindingKey', '')
                    sub_fields = item.get('subFields', [])
                    connected_tables = item.get('connectedTable', [])
                    sub_fields_dict = {}
                    for sf in sub_fields:
                        if isinstance(sf, dict):
                            bk = sf.get('bindingKey', '')
                            sub_fields_dict[bk] = sf
                    if main_table and main_field and sub_fields_dict:
                        supplement_items.append({
                            'main_table': main_table,
                            'main_field': main_field,
                            'parent_key': parent_key,
                            'sub_fields_dict': sub_fields_dict,
                            'connected_tables': connected_tables,
                            'query_key': f"{main_table}_{main_field}",
                            'item': item
                        })

            # 读取所有行数据
            row_data_list = []
            for row_idx in range(2, ws.max_row + 1):
                form_values, missing_columns = build_form_values_from_excel_batch(ws, row_idx, headers, query_items, update_items)
                if query_values and not missing_columns:
                    for field_name in ['filePrefix', 'onesLink', 'dynamicNo', 'ops_remark']:
                        if field_name in query_values:
                            value_data = query_values[field_name]
                            if isinstance(value_data, dict):
                                form_values[field_name] = value_data
                            else:
                                form_values[field_name] = {'value': str(value_data)}
                # 预读取辅助字段条件，避免多线程中读取 Excel
                aux_cache = {}
                for item in update_items:
                    if item.get('inputType') == 'supplement':
                        parent_key = item.get('bindingKey', '')
                        aux_cache[parent_key] = {
                            'new': get_auxiliary_values_from_excel(ws, row_idx, headers, item, is_origin=False),
                            'origin': get_auxiliary_values_from_excel(ws, row_idx, headers, item, is_origin=True),
                        }
                row_data_list.append(
                    {'row_idx': row_idx, 'form_values': form_values, 'missing_columns': missing_columns,
                     'aux_cache': aux_cache})

            # 多线程并行处理
            max_workers = min(8, total_rows)
            from concurrent.futures import ThreadPoolExecutor, as_completed

            def process_row(row_data):
                row_idx = row_data['row_idx']
                form_values = row_data['form_values']
                missing_columns = row_data['missing_columns']
                aux_cache = row_data.get('aux_cache', {})
                if missing_columns:
                    return {'row_idx': row_idx, 'status': 'missing', 'missing_columns': missing_columns}
                for s_item in supplement_items:
                    query_key = s_item['query_key']
                    parent_key = s_item['parent_key']
                    sub_fields_dict = s_item['sub_fields_dict']
                    connected_tables = s_item['connected_tables']
                    item = s_item.get('item')
                    parent_value_data = None
                    parent_unique_key = None
                    if connected_tables:
                        for table in connected_tables:
                            unique_key = f"{parent_key}_{table}"
                            if unique_key in form_values:
                                parent_value_data = form_values[unique_key]
                                parent_unique_key = unique_key
                                break
                    if parent_value_data is None:
                        parent_value_data = form_values.get(parent_key, {})
                        parent_unique_key = parent_key
                    if not parent_value_data:
                        continue
                    new_value = parent_value_data.get('newValue', '')
                    origin_value = parent_value_data.get('originValue', '')

                    # 读取当前行的辅助字段条件
                    new_aux = aux_cache.get(parent_key, {}).get('new', [])
                    origin_aux = aux_cache.get(parent_key, {}).get('origin', [])

                    if new_value and str(new_value).strip():
                        cache_key = build_auxiliary_cache_key(query_key, False, new_aux, new_value)
                        row_data_cache = supplement_data_cache.get(cache_key)
                        if row_data_cache:
                            for sub_binding_key, sub_field in sub_fields_dict.items():
                                db_field = sub_field.get('dbField') or sub_binding_key
                                sub_value = row_data_cache.get(db_field, '')
                                if connected_tables:
                                    for sub_table in connected_tables:
                                        form_values[f"{sub_binding_key}_{sub_table}"] = {
                                            'newValue': str(sub_value) if sub_value is not None else '',
                                            'originValue': '',
                                            'inputType': 'supplement-sub',
                                            'fieldType': 'supplement-sub',
                                            'parentKey': parent_unique_key,
                                            'label': sub_field.get('label', '')}
                                else:
                                    form_values[sub_binding_key] = {
                                        'newValue': str(sub_value) if sub_value is not None else '',
                                        'originValue': '',
                                        'inputType': 'supplement-sub',
                                        'fieldType': 'supplement-sub',
                                        'parentKey': parent_unique_key,
                                        'label': sub_field.get('label', '')}
                    if origin_value and str(origin_value).strip():
                        cache_key = build_auxiliary_cache_key(query_key, True, origin_aux, origin_value)
                        row_data_cache = supplement_data_cache.get(cache_key)
                        if row_data_cache:
                            for sub_binding_key, sub_field in sub_fields_dict.items():
                                db_field = sub_field.get('dbField') or sub_binding_key
                                sub_value = row_data_cache.get(db_field, '')
                                if connected_tables:
                                    for sub_table in connected_tables:
                                        sub_unique_key = f"{sub_binding_key}_{sub_table}"
                                        if sub_unique_key in form_values:
                                            form_values[sub_unique_key]['originValue'] = str(sub_value) if sub_value is not None else ''
                                        else:
                                            form_values[sub_unique_key] = {
                                                'newValue': '',
                                                'originValue': str(sub_value) if sub_value is not None else '',
                                                'inputType': 'supplement-sub',
                                                'fieldType': 'supplement-sub',
                                                'parentKey': parent_unique_key,
                                                'label': sub_field.get('label', '')}
                                else:
                                    if sub_binding_key in form_values:
                                        form_values[sub_binding_key]['originValue'] = str(sub_value) if sub_value is not None else ''
                                    else:
                                        form_values[sub_binding_key] = {
                                            'newValue': '',
                                            'originValue': str(sub_value) if sub_value is not None else '',
                                            'inputType': 'supplement-sub',
                                            'fieldType': 'supplement-sub',
                                            'parentKey': parent_unique_key,
                                            'label': sub_field.get('label', '')}
                validation_result = validate_form_data(config, form_values, query_values)
                if not validation_result['success']:
                    return {'row_idx': row_idx, 'status': 'validation_fail',
                            'fail_reason': '; '.join(validation_result.get('errors', []))}
                sql_result = generate_update_sql(config, form_values)
                if sql_result.get('missing_field_errors'):
                    return {'row_idx': row_idx, 'status': 'missing_field',
                            'fail_reason': '; '.join(sql_result.get('missing_field_errors', []))}
                if sql_result['forward_sqls'] and sql_result['backward_sqls']:
                    return {'row_idx': row_idx, 'status': 'success',
                            'forward_sqls': sql_result['forward_sqls'],
                            'backward_sqls': sql_result['backward_sqls']}
                return {'row_idx': row_idx, 'status': 'no_sql', 'fail_reason': '未生成有效的 SQL 语句'}

            results = []
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {executor.submit(process_row, row_data): row_data['row_idx'] for row_data in row_data_list}
                for future in as_completed(futures):
                    result = future.result()
                    results.append(result)
            results.sort(key=lambda r: r['row_idx'])

            fail_count = 0
            success_count = 0
            all_sql_statements = []
            for result in results:
                row_idx = result['row_idx']
                status = result['status']
                if status == 'missing':
                    fail_count += 1
                    missing_cols_str = ', '.join(result['missing_columns'])
                    ws.cell(row=row_idx, column=fail_column, value=f'缺少必需的列：{missing_cols_str}')
                    ws.cell(row=row_idx, column=fail_column).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                elif status == 'validation_fail':
                    fail_count += 1
                    ws.cell(row=row_idx, column=fail_column, value=result['fail_reason'])
                    ws.cell(row=row_idx, column=fail_column).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                elif status == 'missing_field':
                    fail_count += 1
                    ws.cell(row=row_idx, column=fail_column, value=result['fail_reason'])
                    ws.cell(row=row_idx, column=fail_column).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                elif status == 'success':
                    all_sql_statements.append({
                        'row': row_idx,
                        'forward_sqls': result['forward_sqls'],
                        'backward_sqls': result['backward_sqls']
                    })
                    success_count += 1
                elif status == 'no_sql':
                    fail_count += 1
                    ws.cell(row=row_idx, column=fail_column, value=result['fail_reason'])
                    ws.cell(row=row_idx, column=fail_column).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

            stats_ws = wb.create_sheet(title='导入统计')
            stats_ws.cell(row=1, column=1, value='总行数').font = Font(bold=True)
            stats_ws.cell(row=1, column=2, value=total_rows).font = Font(bold=True)
            stats_ws.cell(row=2, column=1, value='成功数').font = Font(bold=True)
            stats_ws.cell(row=2, column=2, value=success_count).font = Font(bold=True)
            stats_ws.cell(row=3, column=1, value='失败数').font = Font(bold=True)
            stats_ws.cell(row=3, column=2, value=fail_count).font = Font(bold=True)
            stats_ws.cell(row=4, column=1, value='成功率').font = Font(bold=True)
            stats_ws.cell(row=4, column=2, value=f'{success_count / total_rows * 100:.2f}%' if total_rows > 0 else '0%').font = Font(bold=True)

            dynamic_no = ''
            file_prefix = form_name
            if query_values:
                dynamic_no_data = query_values.get('dynamicNo', '')
                if isinstance(dynamic_no_data, dict):
                    dynamic_no = dynamic_no_data.get('value', '')
                elif isinstance(dynamic_no_data, str):
                    dynamic_no = dynamic_no_data
                file_prefix_data = query_values.get('filePrefix', '')
                if isinstance(file_prefix_data, dict):
                    file_prefix = file_prefix_data.get('value', '') or form_name
                elif isinstance(file_prefix_data, str) and file_prefix_data.strip():
                    file_prefix = file_prefix_data.strip()

            save_dir = get_save_path_from_config()
            os.makedirs(save_dir, exist_ok=True)
            all_success = (fail_count == 0 and success_count > 0)
            sql_file_path = None
            if all_success and all_sql_statements:
                sql_filename = f"{dynamic_no}_{file_prefix}.sql"
                sql_filepath = os.path.join(save_dir, sql_filename)
                merged_result = merge_sql_statements(all_sql_statements)
                sql_content = []
                sql_content.append("1.执行语句")
                for idx, sql in enumerate(merged_result['forward_sqls'], 1):
                    if not sql.rstrip().endswith(';'):
                        sql = sql.rstrip() + ';'
                    sql_content.append(sql)
                    sql_content.append("")
                sql_content.append("2.回退语句")
                for idx, sql in enumerate(merged_result['backward_sqls'], 1):
                    if not sql.rstrip().endswith(';'):
                        sql = sql.rstrip() + ';'
                    sql_content.append(sql)
                    sql_content.append("")
                database_ip_ids = config.get('databaseIpIds', [])
                if database_ip_ids:
                    db_configs = DatabaseIPConfig.objects.filter(id__in=database_ip_ids).order_by('id')
                    if db_configs:
                        sql_content.append("3.数据库")
                        for db_config in db_configs:
                            sql_content.append(f"ip：{db_config.ip_address}")
                            sql_content.append(f"库名：{db_config.database_name}")
                            sql_content.append("")
                with open(sql_filepath, 'w', encoding='utf-8') as f:
                    f.write('\n'.join(sql_content))
                sql_file_path = sql_filepath

            # 记录表单调用日志（文件上传批量导入）
            try:
                form_id = config.get('formId')
                if form_id:
                    form_config = FormConfig.objects.get(id=form_id)
                    record_form_usage(form_config, source='dynamic')
            except Exception:
                # 调用统计不应影响主流程
                pass

            excel_file_path = None
            if fail_count > 0:
                now = datetime.now()
                excel_filename = f"{dynamic_no}_导入结果_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
                excel_filepath = os.path.join(save_dir, excel_filename)
                wb.save(excel_filepath)
                excel_file_path = excel_filepath

            return JsonResponse({
                'success': True,
                'message': f'批量导入完成，成功{success_count}条，失败{fail_count}条',
                'sqlFilePath': sql_file_path,
                'excelFilePath': excel_file_path,
                'totalRows': total_rows,
                'successCount': success_count,
                'failCount': fail_count
            })

        except Exception as e:
            print(f"批量导入异常：{e}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'message': f'服务器错误：{str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': '仅支持 POST 请求'}, status=405)


@csrf_exempt
def download_failed_file(request):
    """下载失败的 Excel 结果文件"""
    if request.method == 'GET':
        try:
            file_path = request.GET.get('path', '')

            if not file_path or not os.path.exists(file_path):
                return JsonResponse({'success': False, 'message': '文件不存在'}, status=404)

            response = FileResponse(open(file_path, 'rb'))
            response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            filename = os.path.basename(file_path)
            from urllib.parse import quote
            encoded_filename = quote(filename)
            response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'

            return response

        except Exception as e:
            print(f"下载文件异常：{e}")
            return JsonResponse({'success': False, 'message': f'服务器错误：{str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': '仅支持 GET 请求'}, status=405)
