import json
import openpyxl
import xlrd
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from ..models import ComponentConfig


# ... existing code ...

def get_components(request):
    """获取组件配置列表（支持分页和搜索）"""
    try:
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 5))
        search_term = request.GET.get('search', '').strip()

        # 构建查询集
        queryset = ComponentConfig.objects.all()

        # 如果有关键词搜索
        if search_term:
            queryset = queryset.filter(name__icontains=search_term)

        # 分页
        paginator = Paginator(queryset, page_size)
        try:
            components_page = paginator.page(page)
        except PageNotAnInteger:
            components_page = paginator.page(1)
        except EmptyPage:
            components_page = paginator.page(paginator.num_pages)

        # 序列化数据
        components_data = []
        for component in components_page:
            components_data.append({
                'id': component.id,
                'name': component.name,
                'type': component.component_type,
                'options': component.options,
                'option_count': component.get_option_count(),
                'usage_count': component.usage_count,
                'created_at': component.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': component.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            })

        return JsonResponse({
            'success': True,
            'data': components_data,
            'total': paginator.count,
            'page': components_page.number,
            'page_size': page_size,
            'total_pages': paginator.num_pages
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)




def get_component_detail(request, component_id):
    """获取单个组件配置详情"""
    try:
        component = ComponentConfig.objects.get(id=component_id)

        return JsonResponse({
            'success': True,
            'data': {
                'id': component.id,
                'name': component.name,
                'type': component.component_type,
                'options': component.options,
                'usage_count': component.usage_count,
                'created_at': component.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': component.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        })

    except ComponentConfig.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': '配置项不存在'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@require_http_methods(["POST"])
@csrf_exempt
def save_component(request):
    """创建或更新组件配置"""
    try:
        data = json.loads(request.body)

        name = data.get('name', '').strip()
        component_type = data.get('type', 'select')
        options = data.get('options', [])
        component_id = data.get('id')

        # 验证必填字段
        if not name:
            return JsonResponse({
                'success': False,
                'error': '配置项名称不能为空'
            }, status=400)

        # 验证组件类型
        if component_type not in ['select', 'radio']:
            return JsonResponse({
                'success': False,
                'error': '无效的组件类型'
            }, status=400)

        # 验证选项
        if not isinstance(options, list) or len(options) == 0:
            return JsonResponse({
                'success': False,
                'error': '至少需要配置一个选项'
            }, status=400)

        # 单选框必须有2个选项
        if component_type == 'radio' and len(options) != 2:
            return JsonResponse({
                'success': False,
                'error': '单选框必须有且仅有2个选项'
            }, status=400)

        # 验证选项格式
        for opt in options:
            if not isinstance(opt, dict) or 'value' not in opt or 'label' not in opt:
                return JsonResponse({
                    'success': False,
                    'error': '选项格式不正确，每个选项必须包含value和label'
                }, status=400)

        if component_id:
            # 更新现有配置
            try:
                component = ComponentConfig.objects.get(id=component_id)
                old_name = component.name  # 保存旧名称

                # 检查名称是否已被其他配置使用
                if ComponentConfig.objects.filter(name=name).exclude(id=component_id).exists():
                    return JsonResponse({
                        'success': False,
                        'error': '配置项名称已存在'
                    }, status=400)

                # 如果名称被修改，需要更新所有引用的表单
                updated_forms_count = 0
                if old_name != name:
                    from ..models import FormUpdateItem
                    
                    # 查询所有引用了旧名称的字段
                    referenced_items = FormUpdateItem.objects.filter(component_name=old_name)
                    updated_forms_count = referenced_items.count()
                    
                    if updated_forms_count > 0:
                        # 批量更新为新的名称
                        referenced_items.update(component_name=name)
                        print(f"[配置项重命名] '{old_name}' -> '{name}'，已更新 {updated_forms_count} 个引用字段")

                component.name = name
                component.component_type = component_type
                component.options = options
                component.save()

                # 构建返回消息
                message = '更新成功'
                if updated_forms_count > 0:
                    message = f'更新成功，已同步更新 {updated_forms_count} 个引用该配置项的表单字段'

                return JsonResponse({
                    'success': True,
                    'data': {'id': component.id},
                    'message': message
                })

            except ComponentConfig.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': '配置项不存在'
                }, status=404)
        else:
            # 创建新配置
            # 检查名称是否已存在
            if ComponentConfig.objects.filter(name=name).exists():
                return JsonResponse({
                    'success': False,
                    'error': '配置项名称已存在'
                }, status=400)

            component = ComponentConfig.objects.create(
                name=name,
                component_type=component_type,
                options=options
            )

            return JsonResponse({
                'success': True,
                'data': {'id': component.id},
                'message': '创建成功'
            })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': '无效的JSON数据'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@require_http_methods(["DELETE"])
@csrf_exempt
def delete_component(request, component_id):
    """删除组件配置"""
    try:
        component = ComponentConfig.objects.get(id=component_id)

        # 检查是否有引用
        if component.usage_count > 0:
            return JsonResponse({
                'success': False,
                'error': f'该配置项正在被使用（{component.usage_count}次），无法删除'
            }, status=400)

        component.delete()

        return JsonResponse({
            'success': True,
            'message': '删除成功'
        })

    except ComponentConfig.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': '配置项不存在'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


def get_component_usage(request, component_id):
    """获取组件配置的使用情况（引用的表单）"""
    try:
        component = ComponentConfig.objects.get(id=component_id)

        # 这里可以查询引用了该组件的表单配置
        # 由于目前FormUpdateItem中只有component_name字段,我们需要通过名称匹配
        from ..models import FormUpdateItem

        # 查询引用了该组件的更新字段配置
        referenced_items = FormUpdateItem.objects.filter(component_name=component.name)

        usage_data = []
        for item in referenced_items:
            usage_data.append({
                'form_config_id': item.form_config.id,
                'form_name': item.form_config.form_name,
                'field_label': item.label,
                'field_binding_key': item.binding_key
            })

        return JsonResponse({
            'success': True,
            'data': {
                'component_name': component.name,
                'usage_count': component.usage_count,
                'references': usage_data
            }
        })

    except ComponentConfig.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': '配置项不存在'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@require_http_methods(["POST"])
@csrf_exempt
def import_options_from_excel(request):
    """从Excel文件导入选项配置"""
    try:
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({
                'success': False,
                'error': '请上传文件'
            }, status=400)

        # 检查文件类型
        if not file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({
                'success': False,
                'error': '仅支持Excel文件格式(.xlsx, .xls)'
            }, status=400)

        options = []
        
        # 根据文件类型选择不同的解析方式
        if file.name.endswith('.xlsx'):
            # 使用 openpyxl 解析 .xlsx 文件
            options = parse_xlsx_file(file)
        else:
            # 使用 xlrd 解析 .xls 文件
            options = parse_xls_file(file)

        if len(options) == 0:
            return JsonResponse({
                'success': False,
                'error': 'Excel文件中没有有效的选项数据'
            }, status=400)

        return JsonResponse({
            'success': True,
            'data': options,
            'message': f'成功读取 {len(options)} 个选项'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'解析文件失败:{str(e)}'
        }, status=500)


def parse_xlsx_file(file):
    """解析 .xlsx 文件（使用 openpyxl）"""
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # 读取表头
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value is not None:
            headers.append(str(cell_value).strip())

    # 查找列
    value_col_idx, label_col_idx = find_columns(headers)

    # 读取数据行
    options = []
    seen_values = set()

    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=value_col_idx).value
        label = ws.cell(row=row, column=label_col_idx).value

        if value is None and label is None:
            continue

        value_str = str(value).strip() if value is not None else ''
        label_str = str(label).strip() if label is not None else ''

        if not value_str or value_str in seen_values:
            continue

        seen_values.add(value_str)
        options.append({
            'value': value_str,
            'label': label_str if label_str else value_str
        })

    return options


def parse_xls_file(file):
    """解析 .xls 文件（使用 xlrd）"""
    # 读取文件内容到内存
    file_content = file.read()
    
    # 使用 xlrd 打开工作簿
    wb = xlrd.open_workbook(file_contents=file_content)
    ws = wb.sheet_by_index(0)

    if ws.nrows < 2:
        return []

    # 读取表头
    headers = []
    for col in range(ws.ncols):
        cell_value = ws.cell_value(0, col)
        if cell_value is not None:
            headers.append(str(cell_value).strip())

    # 查找列
    value_col_idx, label_col_idx = find_columns(headers)

    # 读取数据行
    options = []
    seen_values = set()

    for row in range(1, ws.nrows):
        value = ws.cell_value(row, value_col_idx - 1)  # xlrd列从0开始
        label = ws.cell_value(row, label_col_idx - 1)

        if value is None and label is None:
            continue

        # 处理xlrd的单元格类型
        value_str = format_xls_cell(value, ws.cell_type(row, value_col_idx - 1))
        label_str = format_xls_cell(label, ws.cell_type(row, label_col_idx - 1))

        if not value_str or value_str in seen_values:
            continue

        seen_values.add(value_str)
        options.append({
            'value': value_str,
            'label': label_str if label_str else value_str
        })

    return options


def find_columns(headers):
    """查找代码项编号和代码项中文名称列（严格匹配）"""
    value_col_idx = None
    label_col_idx = None

    for idx, header in enumerate(headers):
        if header == '代码项编号':
            value_col_idx = idx + 1  # 列从1开始
        elif header == '代码项中文名称':
            label_col_idx = idx + 1

    if not value_col_idx or not label_col_idx:
        missing = []
        if not value_col_idx:
            missing.append('代码项编号')
        if not label_col_idx:
            missing.append('代码项中文名称')
        raise ValueError(f'未找到必需的列：{", ".join(missing)}')

    return value_col_idx, label_col_idx


def format_xls_cell(value, cell_type):
    """格式化xls单元格值"""
    if value is None:
        return ''
    
    # 处理数字类型（Excel中的数字可能被识别为float）
    if cell_type == xlrd.XL_CELL_NUMBER:
        # 如果是整数，去掉小数部分
        if isinstance(value, float) and value == int(value):
            return str(int(value))
        return str(value)
    
    return str(value).strip()